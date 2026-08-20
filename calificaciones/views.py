from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from decimal import Decimal

from colegios.models import Colegio, CursoColegio, SeccionCurso, Asignatura, Estudiante, ConfiguracionAcademica
from solicitudes.models import MiembroColegio
from .models import Evaluacion, Nota, ObservacionBoletin


def obtener_datos_base_calificaciones(request):
    user = request.user
    colegio = user.colegios_administrados.order_by('-fecha_creacion').first()
    miembro = None
    if not colegio:
        miembro = MiembroColegio.objects.filter(usuario=user, activo=True).order_by('-fecha_ingreso').first()
        if miembro:
            colegio = miembro.colegio
    else:
        miembro = MiembroColegio.objects.filter(usuario=user, colegio=colegio, activo=True).first()

    if colegio and not miembro:
        miembro = MiembroColegio.objects.filter(usuario=user, colegio=colegio).first()

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first() if colegio else None

    is_admin = False
    if colegio:
        is_admin = (
            user.colegios_administrados.filter(id=colegio.id).exists()
            or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
        )

    return colegio, miembro, periodo, is_admin


@login_required
def libro_calificaciones_view(request):
    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    if not colegio:
        messages.warning(request, "No estás asociado a ningún colegio para ver calificaciones.")
        return redirect('solicitar_acceso')

    # Secciones disponibles del colegio
    if is_admin:
        secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('curso__nivel', 'curso__nombre', 'letra')
    else:
        # Docente: Secciones donde es Profesor Jefe + Secciones donde dicta alguna Asignatura
        from django.db.models import Q
        asignaturas_docente = Asignatura.objects.filter(colegio=colegio, docente=request.user, activo=True)
        cursos_ids = asignaturas_docente.values_list('curso_id', flat=True).distinct()
        
        secciones = SeccionCurso.objects.filter(
            Q(curso__colegio=colegio, activo=True) &
            (Q(profesor_jefe=request.user) | Q(curso_id__in=cursos_ids))
        ).distinct().order_by('curso__nivel', 'curso__nombre', 'letra')

    seccion_id = request.GET.get('seccion')
    asignatura_id = request.GET.get('asignatura')

    seccion_seleccionada = None
    if seccion_id and seccion_id.isdigit():
        seccion_seleccionada = secciones.filter(id=int(seccion_id)).first()
    if not seccion_seleccionada and secciones.exists():
        seccion_seleccionada = secciones.first()

    # Asignaturas del curso de la sección seleccionada
    asignaturas = []
    asignatura_seleccionada = None
    is_profesor_jefe = False

    if seccion_seleccionada:
        is_profesor_jefe = (seccion_seleccionada.profesor_jefe == request.user)
        # Si es Admin o es el Profesor Jefe de este curso -> puede ver TODAS las asignaturas del curso
        if is_admin or is_profesor_jefe:
            asignaturas = Asignatura.objects.filter(curso=seccion_seleccionada.curso, activo=True).order_by('nombre')
        else:
            # Si solo es docente de asignatura -> SOLO ve su propia materia
            asignaturas = Asignatura.objects.filter(curso=seccion_seleccionada.curso, docente=request.user, activo=True).order_by('nombre')

        if asignatura_id and asignatura_id.isdigit():
            asignatura_seleccionada = asignaturas.filter(id=int(asignatura_id)).first()
        if not asignatura_seleccionada and asignaturas.exists():
            asignatura_seleccionada = asignaturas.first()



    # Cargar datos de la matriz si hay sección y asignatura seleccionadas
    estudiantes = []
    evaluaciones = []
    matriz_notas = []
    
    promedio_curso = None
    tasa_aprobacion = None
    alumnos_riesgo_count = 0
    evaluaciones_count = 0

    if seccion_seleccionada and asignatura_seleccionada:
        estudiantes = Estudiante.objects.filter(seccion=seccion_seleccionada, activo=True).order_by('nombre_completo')
        evaluaciones = Evaluacion.objects.filter(
            colegio=colegio,
            seccion=seccion_seleccionada,
            asignatura=asignatura_seleccionada
        ).order_by('fecha', 'id')

        evaluaciones_count = evaluaciones.count()

        # Construir matriz por estudiante
        total_promedios = []
        aprobados_count = 0

        for est in estudiantes:
            notas_dict = {}
            suma_ponderada = Decimal('0.0')
            suma_ponderaciones = Decimal('0.0')
            cantidad_notas = 0

            nota_minima = float(periodo.nota_minima_aprobacion) if (periodo and periodo.nota_minima_aprobacion) else 4.0
            regla_redondeo = periodo.regla_redondeo if periodo else 'un_decimal'

            for ev in evaluaciones:
                nota_obj = Nota.objects.filter(evaluacion=ev, estudiante=est).first()
                if nota_obj:
                    val = float(nota_obj.valor)
                    notas_dict[ev.id] = {
                        'valor': val,
                        'nota_id': nota_obj.id,
                        'rojo': val < nota_minima
                    }
                    val_dec = Decimal(str(val))
                    pond_dec = Decimal(str(ev.ponderacion))
                    suma_ponderada += val_dec * pond_dec
                    suma_ponderaciones += pond_dec
                    cantidad_notas += 1
                else:
                    notas_dict[ev.id] = None

            # Calcular promedio individual según la regla de redondeo configurada
            promedio_estudiante = None
            es_rojo_promedio = False
            if cantidad_notas > 0 and suma_ponderaciones > 0:
                calc = float(suma_ponderada / suma_ponderaciones)
                if regla_redondeo == 'dos_decimales':
                    promedio_estudiante = round(calc, 2)
                elif regla_redondeo == 'truncado':
                    promedio_estudiante = int(calc * 10) / 10.0
                else:
                    promedio_estudiante = round(calc, 1)

                total_promedios.append(promedio_estudiante)
                if promedio_estudiante < nota_minima:
                    es_rojo_promedio = True
                    alumnos_riesgo_count += 1
                else:
                    aprobados_count += 1


            matriz_notas.append({
                'estudiante': est,
                'notas': notas_dict,
                'promedio': promedio_estudiante,
                'es_rojo': es_rojo_promedio,
                'total_notas': cantidad_notas
            })

        if total_promedios:
            promedio_curso = round(sum(total_promedios) / len(total_promedios), 1)
            tasa_aprobacion = round((aprobados_count / len(total_promedios)) * 100, 1)

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'secciones': secciones,
        'seccion_seleccionada': seccion_seleccionada,
        'asignaturas': asignaturas,
        'asignatura_seleccionada': asignatura_seleccionada,
        'estudiantes': estudiantes,
        'evaluaciones': evaluaciones,
        'matriz_notas': matriz_notas,
        'promedio_curso': promedio_curso,
        'tasa_aprobacion': tasa_aprobacion,
        'alumnos_riesgo_count': alumnos_riesgo_count,
        'evaluaciones_count': evaluaciones_count,
        'is_profesor_jefe': is_profesor_jefe,
        'hoy': timezone.now(),
    }
    return render(request, 'calificaciones/libro_clases.html', context)



@login_required
def crear_evaluacion_view(request):
    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    if request.method == 'POST':
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')
        nombre = request.POST.get('nombre', '').strip()
        fecha_str = request.POST.get('fecha')
        ponderacion_str = request.POST.get('ponderacion', '1.0')
        periodo_nombre = request.POST.get('periodo_nombre', '1° Semestre')

        seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
        asignatura = get_object_or_404(Asignatura, id=asignatura_id, curso=seccion.curso)

        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
        except ValueError:
            fecha_obj = timezone.now().date()

        try:
            ponderacion = float(ponderacion_str)
        except ValueError:
            ponderacion = 1.0

        Evaluacion.objects.create(
            colegio=colegio,
            seccion=seccion,
            asignatura=asignatura,
            nombre=nombre if nombre else "Evaluación",
            fecha=fecha_obj,
            ponderacion=ponderacion,
            periodo_nombre=periodo_nombre,
            creado_por=request.user
        )

        messages.success(request, f"Evaluación '{nombre}' creada exitosamente.")
        return redirect(f"/calificaciones/?seccion={seccion.id}&asignatura={asignatura.id}")

    return redirect('libro_calificaciones')


@login_required
def eliminar_evaluacion_view(request, evaluacion_id):
    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    evaluacion = get_object_or_404(Evaluacion, id=evaluacion_id, colegio=colegio)
    seccion_id = evaluacion.seccion_id
    asignatura_id = evaluacion.asignatura_id
    
    nombre_ev = evaluacion.nombre
    evaluacion.delete()
    messages.success(request, f"Evaluación '{nombre_ev}' eliminada.")
    return redirect(f"/calificaciones/?seccion={seccion_id}&asignatura={asignatura_id}")


@login_required
@login_required
def guardar_notas_view(request):
    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    if request.method == 'POST':
        CONCEPTUAL_MAP = {
            'MB': 6.5, 'B': 5.5, 'S': 4.0, 'I': 3.0,
            'PL': 7.0, 'L': 5.5, 'EP': 4.0, 'NL': 3.0, 'NT': 2.0
        }

        # Caso 1: Envío individual por AJAX
        estudiante_id_single = request.POST.get('estudiante_id')
        evaluacion_id_single = request.POST.get('evaluacion_id')
        valor_single = request.POST.get('valor')

        if estudiante_id_single and evaluacion_id_single:
            val_clean = valor_single.strip().upper().replace(',', '.') if valor_single else ''
            if val_clean in CONCEPTUAL_MAP:
                val_clean = str(CONCEPTUAL_MAP[val_clean])
            
            if val_clean:
                try:
                    val_float = float(val_clean)
                    if 1.0 <= val_float <= 7.0:
                        val_dec = Decimal(f"{val_float:.1f}")
                        ev_obj = Evaluacion.objects.filter(id=evaluacion_id_single, colegio=colegio).first()
                        est_obj = Estudiante.objects.filter(id=estudiante_id_single, colegio=colegio).first()
                        if ev_obj and est_obj:
                            Nota.objects.update_or_create(
                                evaluacion=ev_obj,
                                estudiante=est_obj,
                                defaults={'valor': val_dec}
                            )
                            return JsonResponse({'status': 'success', 'valor': str(val_dec)})
                except ValueError:
                    pass
            else:
                Nota.objects.filter(evaluacion_id=evaluacion_id_single, estudiante_id=estudiante_id_single).delete()
                return JsonResponse({'status': 'deleted'})
            return JsonResponse({'status': 'invalid'}, status=400)

        # Caso 2: Envío masivo por formulario (Botón Guardar Calificaciones)
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')

        count_guardadas = 0
        for key, val_str in request.POST.items():
            if key.startswith('nota_'):
                parts = key.split('_')
                if len(parts) == 3:
                    evaluacion_id = parts[1]
                    estudiante_id = parts[2]
                    val_clean = val_str.strip().upper().replace(',', '.')
                    
                    if val_clean in CONCEPTUAL_MAP:
                        val_clean = str(CONCEPTUAL_MAP[val_clean])

                    if val_clean:
                        try:
                            val_float = float(val_clean)
                            if 1.0 <= val_float <= 7.0:
                                val_dec = Decimal(f"{val_float:.1f}")
                                ev_obj = Evaluacion.objects.filter(id=evaluacion_id, colegio=colegio).first()
                                est_obj = Estudiante.objects.filter(id=estudiante_id, colegio=colegio).first()
                                
                                if ev_obj and est_obj:
                                    Nota.objects.update_or_create(
                                        evaluacion=ev_obj,
                                        estudiante=est_obj,
                                        defaults={'valor': val_dec}
                                    )
                                    count_guardadas += 1
                        except ValueError:
                            pass
                    else:
                        Nota.objects.filter(evaluacion_id=evaluacion_id, estudiante_id=estudiante_id).delete()

        messages.success(request, f"¡Calificaciones guardadas exitosamente ({count_guardadas} notas actualizadas)!")
        return redirect(f"/calificaciones/?seccion={seccion_id}&asignatura={asignatura_id}")

    return redirect('libro_calificaciones')



@login_required
def exportar_notas_excel_view(request, seccion_id, asignatura_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    asignatura = get_object_or_404(Asignatura, id=asignatura_id, curso=seccion.curso)

    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')
    evaluaciones = Evaluacion.objects.filter(colegio=colegio, seccion=seccion, asignatura=asignatura).order_by('fecha', 'id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro de Notas"
    ws.views.sheetView[0].showGridLines = True

    # Encabezado
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Libro de Calificaciones - {asignatura.nombre}"
    ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="7C5CFC")

    ws["A2"] = f"Colegio: {colegio.nombre} | Sección: {seccion.nombre}"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    ws["A3"] = f"Fecha de reporte: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(name="Segoe UI", size=9, color="555555")

    # Cabecera de la tabla
    headers = ["Estudiante", "RUT"]
    for ev in evaluaciones:
        headers.append(f"{ev.nombre} ({ev.fecha.strftime('%d/%m')})")
    headers.append("Promedio")

    ws.append([]) # Fila 4 vacía
    ws.append(headers) # Fila 5 es la cabecera

    fill_header = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center if col_num > 2 else align_left
        cell.border = thin_border

    fill_red = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")
    font_red = Font(name="Segoe UI", size=9, color="E11D48", bold=True)
    fill_green = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    font_green = Font(name="Segoe UI", size=9, color="059669")

    current_row = 6
    for est in estudiantes:
        row_data = [est.nombre_completo, est.rut or "-"]
        suma_pond = Decimal('0.0')
        suma_coef = Decimal('0.0')
        cnt = 0

        for ev in evaluaciones:
            nota_obj = Nota.objects.filter(evaluacion=ev, estudiante=est).first()
            if nota_obj:
                v = float(nota_obj.valor)
                row_data.append(v)
                v_dec = Decimal(str(v))
                p_dec = Decimal(str(ev.ponderacion))
                suma_pond += v_dec * p_dec
                suma_coef += p_dec
                cnt += 1
            else:
                row_data.append("-")

        prom_val = None
        if cnt > 0 and suma_coef > 0:
            prom_val = round(float(suma_pond / suma_coef), 1)
            row_data.append(prom_val)
        else:
            row_data.append("-")

        ws.append(row_data)

        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=9)

            if col_num > 2 and col_num <= len(row_data) - 1:
                val = row_data[col_num - 1]
                cell.alignment = align_center
                if isinstance(val, (int, float)):
                    if val < 4.0:
                        cell.fill = fill_red
                        cell.font = font_red
                    else:
                        cell.fill = fill_green
                        cell.font = font_green
            elif col_num == len(row_data):
                cell.alignment = align_center
                if isinstance(prom_val, (int, float)):
                    if prom_val < 4.0:
                        cell.fill = fill_red
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="E11D48")
                    else:
                        cell.fill = fill_green
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="059669")
            else:
                cell.alignment = align_left

        current_row += 1

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(5, current_row):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws.column_dimensions['A'].width = 32


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="notas_{seccion.nombre}_{asignatura.nombre}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def boletin_estudiante_view(request, estudiante_id):
    colegio, miembro, periodo, is_admin = obtener_datos_base_calificaciones(request)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)
    seccion = estudiante.seccion

    if not seccion:
        messages.warning(request, "El estudiante no tiene una sección o curso asignado.")
        return redirect('listar_estudiantes')

    asignaturas = Asignatura.objects.filter(curso=seccion.curso, activo=True).order_by('nombre')

    asig_docente_ids = []
    if not is_admin:
        asig_docente_ids = Asignatura.objects.filter(colegio=colegio, docente=request.user, activo=True).values_list('id', flat=True)

    nota_minima = float(periodo.nota_minima_aprobacion) if (periodo and periodo.nota_minima_aprobacion) else 4.0
    regla_redondeo = periodo.regla_redondeo if periodo else 'un_decimal'

    resumen_asignaturas = []
    total_promedios_validos = []

    for asig in asignaturas:
        evaluaciones = Evaluacion.objects.filter(colegio=colegio, seccion=seccion, asignatura=asig).order_by('fecha', 'id')
        notas_list = []
        suma_ponderada = Decimal('0.0')
        suma_ponderaciones = Decimal('0.0')
        cant_notas = 0

        for ev in evaluaciones:
            nota_obj = Nota.objects.filter(evaluacion=ev, estudiante=estudiante).first()
            if nota_obj:
                val = float(nota_obj.valor)
                notas_list.append({
                    'evaluacion': ev,
                    'valor': val,
                    'rojo': val < nota_minima
                })
                val_dec = Decimal(str(val))
                pond_dec = Decimal(str(ev.ponderacion))
                suma_ponderada += val_dec * pond_dec
                suma_ponderaciones += pond_dec
                cant_notas += 1
            else:
                notas_list.append({'evaluacion': ev, 'valor': None})

        promedio = None
        es_rojo = False
        if cant_notas > 0 and suma_ponderaciones > 0:
            calc = float(suma_ponderada / suma_ponderaciones)
            if regla_redondeo == 'dos_decimales':
                promedio = round(calc, 2)
            elif regla_redondeo == 'truncado':
                promedio = int(calc * 10) / 10.0
            else:
                promedio = round(calc, 1)

            total_promedios_validos.append(promedio)
            if promedio < nota_minima:
                es_rojo = True

        resumen_asignaturas.append({
            'asignatura': asig,
            'notas': notas_list,
            'promedio': promedio,
            'es_rojo': es_rojo,
            'es_docente': (is_admin or asig.id in asig_docente_ids)
        })

    promedio_general = None
    if total_promedios_validos:
        calc_gen = sum(total_promedios_validos) / len(total_promedios_validos)
        if regla_redondeo == 'dos_decimales':
            promedio_general = round(calc_gen, 2)
        elif regla_redondeo == 'truncado':
            promedio_general = int(calc_gen * 10) / 10.0
        else:
            promedio_general = round(calc_gen, 1)

    from asistencia.models import DetalleAsistencia
    detalles_asig = DetalleAsistencia.objects.filter(estudiante=estudiante)
    total_clases = detalles_asig.count()
    presentes = detalles_asig.filter(estado__in=['presente', 'tarde', 'justificado']).count()
    ausentes = detalles_asig.filter(estado='ausente').count()
    porcentaje_asistencia = round((presentes / total_clases * 100), 1) if total_clases > 0 else 100.0

    pct_min_asistencia = float(periodo.porcentaje_asistencia_minima) if (periodo and periodo.porcentaje_asistencia_minima) else 85.0
    en_riesgo_asistencia = porcentaje_asistencia < pct_min_asistencia

    # Manejo de guardado de observación del profesor
    if request.method == 'POST':
        texto_obs = request.POST.get('observacion_texto', '').strip()
        obs_obj, _ = ObservacionBoletin.objects.get_or_create(
            colegio=colegio,
            estudiante=estudiante,
            periodo_nombre="1° Semestre",
            defaults={'creado_por': request.user}
        )
        obs_obj.texto = texto_obs
        obs_obj.creado_por = request.user
        obs_obj.save()
        messages.success(request, f"¡Observación para {estudiante.nombre_completo} guardada exitosamente!")
        return redirect('boletin_estudiante', estudiante_id=estudiante.id)

    observacion_obj = ObservacionBoletin.objects.filter(estudiante=estudiante, periodo_nombre="1° Semestre").first()
    estudiantes_seccion = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'estudiante': estudiante,
        'seccion': seccion,
        'estudiantes_seccion': estudiantes_seccion,
        'resumen_asignaturas': resumen_asignaturas,
        'promedio_general': promedio_general,
        'total_clases': total_clases,
        'presentes': presentes,
        'ausentes': ausentes,
        'porcentaje_asistencia': porcentaje_asistencia,
        'en_riesgo_asistencia': en_riesgo_asistencia,
        'pct_min_asistencia': pct_min_asistencia,
        'observacion_obj': observacion_obj,
        'hoy': timezone.now(),
    }
    return render(request, 'calificaciones/boletin_estudiante.html', context)


