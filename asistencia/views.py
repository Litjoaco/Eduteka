from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Prefetch
from datetime import datetime
from colegios.models import Colegio, CursoColegio, SeccionCurso, Estudiante
from solicitudes.models import MiembroColegio
from .models import RegistroAsistencia, DetalleAsistencia

def obtener_colegio_usuario(user):
    """
    Obtiene el colegio del cual el usuario es administrador o miembro activo.
    """
    colegio = user.colegios_administrados.order_by('-fecha_creacion').first()
    if not colegio:
        miembro = MiembroColegio.objects.filter(usuario=user, activo=True).order_by('-fecha_ingreso').first()
        if miembro:
            colegio = miembro.colegio
    return colegio

def obtener_datos_base_asistencia(request):
    user = request.user
    colegio = user.colegios_administrados.order_by('-fecha_creacion').first()
    miembro = None
    if not colegio:
        miembro = MiembroColegio.objects.filter(usuario=user, activo=True).order_by('-fecha_ingreso').first()
        if miembro:
            colegio = miembro.colegio
    else:
        miembro = MiembroColegio.objects.filter(usuario=user, colegio=colegio, activo=True).first()

    if colegio and not miembro:
        miembro = MiembroColegio.objects.filter(usuario=user, colegio=colegio).first()

    from colegios.models import ConfiguracionAcademica
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first() if colegio else None

    is_admin = False
    if colegio:
        rol_nombre_str = (miembro.rol.nombre.lower() if miembro and miembro.rol else '')
        is_admin = (
            user.is_superuser
            or user.colegios_administrados.filter(id=colegio.id).exists()
            or any(r in rol_nombre_str for r in ['administrador', 'director', 'inspector', 'utp', 'coordinador'])
        )

    return colegio, miembro, periodo, is_admin

from colegios.models import Colegio, CursoColegio, SeccionCurso, Estudiante, Asignatura

@login_required
def registrar_asistencia_view(request):
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    if not colegio:
        messages.warning(request, "No tienes un colegio asociado para registrar asistencia.")
        return redirect('solicitar_acceso')

    modalidad = periodo.modalidad_asistencia if periodo else 'asignatura'

    # Obtener los cursos con sus secciones activas pre-cargadas para este colegio
    secciones_activas = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('letra')
    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).prefetch_related(
        Prefetch('secciones', queryset=secciones_activas)
    ).order_by('nivel', 'nombre')


    # Asignaturas disponibles si es modalidad por asignatura
    asignaturas_disponibles = []
    if modalidad == 'asignatura':
        if is_admin:
            asignaturas_disponibles = Asignatura.objects.filter(colegio=colegio, activo=True).select_related('curso')
        else:
            asignaturas_disponibles = Asignatura.objects.filter(colegio=colegio, docente=request.user, activo=True).select_related('curso')

    # Procesar selección inicial
    fecha_str = request.GET.get('fecha') or request.POST.get('fecha')
    if not fecha_str:
        fecha_str = timezone.now().strftime('%Y-%m-%d')

    if request.method == 'POST' and 'seccion_id' in request.POST:
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')
        if modalidad == 'asignatura' and asignatura_id:
            return redirect(f'/asistencia/registrar/{seccion_id}/?fecha={fecha_str}&asignatura={asignatura_id}')
        return redirect(f'/asistencia/registrar/{seccion_id}/?fecha={fecha_str}')

    if is_admin:
        secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__nivel', 'curso__nombre', 'letra')
    else:
        asigs_prof = Asignatura.objects.filter(colegio=colegio, docente=request.user, activo=True)
        cursos_prof = asigs_prof.values_list('curso_id', flat=True).distinct()
        if cursos_prof.exists():
            secciones = SeccionCurso.objects.filter(curso_id__in=cursos_prof, activo=True).select_related('curso').order_by('curso__nivel', 'curso__nombre', 'letra')
        else:
            secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__nivel', 'curso__nombre', 'letra')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'cursos': cursos,
        'secciones': secciones,
        'modalidad': modalidad,
        'asignaturas_disponibles': asignaturas_disponibles,
        'fecha_actual': fecha_str,
        'hoy': timezone.now().date(),
    }
    return render(request, 'asistencia/registrar.html', context)



@login_required
def registrar_asistencia_seccion_view(request, seccion_id):
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    modalidad = periodo.modalidad_asistencia if periodo else 'asignatura'

    fecha_str = request.GET.get('fecha')
    if not fecha_str:
        fecha_str = timezone.now().strftime('%Y-%m-%d')

    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        fecha_obj = timezone.now().date()
        fecha_str = fecha_obj.strftime('%Y-%m-%d')

    # Asignaturas de la sección si aplica la modalidad
    asignaturas_disponibles = []
    asignatura_seleccionada = None
    if modalidad == 'asignatura':
        if is_admin:
            asignaturas_disponibles = Asignatura.objects.filter(curso=seccion.curso, activo=True).order_by('nombre')
        else:
            asignaturas_disponibles = Asignatura.objects.filter(curso=seccion.curso, docente=request.user, activo=True).order_by('nombre')
        
        asignatura_id = request.GET.get('asignatura')
        if asignatura_id and asignatura_id.isdigit():
            asignatura_seleccionada = asignaturas_disponibles.filter(id=int(asignatura_id)).first()
        if not asignatura_seleccionada and asignaturas_disponibles.exists():
            asignatura_seleccionada = asignaturas_disponibles.first()

    # Obtener alumnos de la sección
    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    if request.method == 'POST':
        # Crear o actualizar el registro principal de asistencia
        registro, created = RegistroAsistencia.objects.get_or_create(
            seccion=seccion,
            asignatura=asignatura_seleccionada,
            fecha=fecha_obj,
            defaults={'creado_por': request.user}
        )
        if not created:
            registro.creado_por = request.user
            registro.save()

        # Guardar el detalle de cada alumno
        for estudiante in estudiantes:
            estado = request.POST.get(f'estado_{estudiante.id}', 'presente')
            observacion = request.POST.get(f'obs_{estudiante.id}', '').strip()

            DetalleAsistencia.objects.update_or_create(
                registro=registro,
                estudiante=estudiante,
                defaults={
                    'estado': estado,
                    'observacion': observacion if observacion else None
                }
            )

        asig_nombre = f" en {asignatura_seleccionada.nombre}" if asignatura_seleccionada else ""
        messages.success(request, f"¡Asistencia de {seccion.nombre}{asig_nombre} para el {fecha_obj.strftime('%d/%m/%Y')} guardada con éxito!")
        if asignatura_seleccionada:
            return redirect(f'/asistencia/registrar/{seccion.id}/?fecha={fecha_str}&asignatura={asignatura_seleccionada.id}')
        return redirect(f'/asistencia/registrar/?fecha={fecha_str}')

    # Si es GET, ver si ya existe un registro de asistencia
    registro_existente = RegistroAsistencia.objects.filter(
        seccion=seccion,
        asignatura=asignatura_seleccionada,
        fecha=fecha_obj
    ).first()
    
    # Mapear estados existentes si los hay
    estados_alumnos = {}
    obs_alumnos = {}
    if registro_existente:
        detalles = registro_existente.detalles.all()
        for det in detalles:
            estados_alumnos[det.estudiante_id] = det.estado
            obs_alumnos[det.estudiante_id] = det.observacion or ''


    # Asignar estado actual para renderizado en plantilla
    for est in estudiantes:
        est.estado_actual = estados_alumnos.get(est.id, 'presente')
        est.obs_actual = obs_alumnos.get(est.id, '')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'seccion': seccion,
        'fecha': fecha_obj,
        'fecha_str': fecha_str,
        'estudiantes': estudiantes,
        'registro_existente': registro_existente,
        'modalidad': modalidad,
        'asignatura_seleccionada': asignatura_seleccionada,
        'asignaturas_disponibles': asignaturas_disponibles,
    }
    return render(request, 'asistencia/registrar_seccion.html', context)


@login_required
def historial_asistencia_view(request):
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    if not colegio:
        messages.warning(request, "No tienes un colegio asociado para ver historial.")
        return redirect('solicitar_acceso')

    # Obtener los cursos con sus secciones activas pre-cargadas para este colegio
    secciones_activas = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('letra')
    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).prefetch_related(
        Prefetch('secciones', queryset=secciones_activas)
    ).order_by('nivel', 'nombre')


    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'cursos': cursos,
    }
    return render(request, 'asistencia/historial.html', context)

@login_required
def historial_seccion_view(request, seccion_id):
    from django.db.models import Count, Q
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)

    # Obtener mes de filtro (GET)
    mes_str = request.GET.get('mes')
    mes_seleccionado = None
    if mes_str and mes_str.isdigit():
        mes_seleccionado = int(mes_str)
        if 1 <= mes_seleccionado <= 12:
            registros = RegistroAsistencia.objects.filter(seccion=seccion, fecha__month=mes_seleccionado).order_by('-fecha')
        else:
            mes_seleccionado = None
            registros = RegistroAsistencia.objects.filter(seccion=seccion).order_by('-fecha')
    else:
        registros = RegistroAsistencia.objects.filter(seccion=seccion).order_by('-fecha')

    # Calcular estadísticas por cada registro
    historial_data = []
    for reg in registros:
        detalles = reg.detalles.all()
        total = detalles.count()
        if total > 0:
            presentes = detalles.filter(estado='presente').count()
            tardes = detalles.filter(estado='tarde').count()
            ausentes = detalles.filter(estado='ausente').count()
            justificados = detalles.filter(estado='justificado').count()
            
            # Asistencia es presentes + tarde + justificado
            tasa_asistencia = ((presentes + tardes + justificados) / total) * 100
        else:
            presentes = 0
            tardes = 0
            ausentes = 0
            justificados = 0
            tasa_asistencia = 0

        historial_data.append({
            'registro': reg,
            'total': total,
            'presentes': presentes,
            'tardes': tardes,
            'ausentes': ausentes,
            'justificados': justificados,
            'tasa': round(tasa_asistencia, 1)
        })

    # 1. Datos para el gráfico cronológico (de más antiguo a más reciente)
    registros_cronologicos = list(registros)[::-1]
    chart_labels = []
    chart_data = []
    for h in registros_cronologicos:
        chart_labels.append(h.fecha.strftime('%d/%m'))
        detalles = h.detalles.all()
        tot = detalles.count()
        if tot > 0:
            pres = detalles.filter(estado__in=['presente', 'tarde', 'justificado']).count()
            tasa_dia = (pres / tot) * 100
        else:
            tasa_dia = 100.0
        chart_data.append(round(tasa_dia, 1))

    # 2. Resumen por estudiante
    filter_q = Q(detalles_asistencia__registro__seccion=seccion)
    if mes_seleccionado:
        filter_q &= Q(detalles_asistencia__registro__fecha__month=mes_seleccionado)
        
    estudiantes_data = Estudiante.objects.filter(seccion=seccion, activo=True).annotate(
        total_asistencias=Count('detalles_asistencia', filter=filter_q),
        presentes=Count('detalles_asistencia', filter=filter_q & Q(detalles_asistencia__estado='presente')),
        tardes=Count('detalles_asistencia', filter=filter_q & Q(detalles_asistencia__estado='tarde')),
        ausentes=Count('detalles_asistencia', filter=filter_q & Q(detalles_asistencia__estado='ausente')),
        justificados=Count('detalles_asistencia', filter=filter_q & Q(detalles_asistencia__estado='justificado')),
    ).order_by('nombre_completo')

    alumnos_resumen = []
    for est in estudiantes_data:
        total = est.total_asistencias
        if total > 0:
            tasa = ((est.presentes + est.tardes + est.justificados) / total) * 100
        else:
            tasa = 100.0
            
        alumnos_resumen.append({
            'estudiante': est,
            'presentes': est.presentes,
            'tardes': est.tardes,
            'ausentes': est.ausentes,
            'justificados': est.justificados,
            'total': total,
            'tasa': round(tasa, 1),
            'critico': tasa < 85.0
        })

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'seccion': seccion,
        'historial_data': historial_data,
        'mes_seleccionado': mes_seleccionado,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'alumnos_resumen': alumnos_resumen,
    }
    return render(request, 'asistencia/historial_seccion.html', context)


@login_required
def exportar_asistencia_excel(request, seccion_id):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse

    colegio = obtener_colegio_usuario(request.user)
    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    
    # Obtener alumnos de la sección
    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')
    
    # Obtener registros de asistencia
    mes_str = request.GET.get('mes')
    if mes_str and mes_str.isdigit():
        mes_val = int(mes_str)
        if 1 <= mes_val <= 12:
            registros = RegistroAsistencia.objects.filter(seccion=seccion, fecha__month=mes_val).order_by('fecha')
        else:
            registros = RegistroAsistencia.objects.filter(seccion=seccion).order_by('fecha')
    else:
        registros = RegistroAsistencia.objects.filter(seccion=seccion).order_by('fecha')
        
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla Asistencia"
    
    # Activar gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Título institucional
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Reporte de Asistencia - {seccion.curso.nombre} {seccion.letra}"
    ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="7C5CFC")
    
    ws["A2"] = f"Colegio: {colegio.nombre}"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    ws["A3"] = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(name="Segoe UI", size=9, color="555555")
    
    # Fila de Cabecera de la tabla
    headers = ["Estudiante", "RUT"]
    # Columnas de fechas
    for reg in registros:
        headers.append(reg.fecha.strftime('%d/%m'))
    headers.append("% Asistencia")
    
    ws.append([]) # Fila 4 vacía
    ws.append(headers) # Fila 5 es la cabecera
    
    # Estilos de cabecera
    fill_header = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # Estilos de estados
    state_fills = {
        'presente': PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),
        'tarde': PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
        'ausente': PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid"),
        'justificado': PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
    }
    state_fonts = {
        'presente': Font(name="Segoe UI", size=9, color="059669", bold=True),
        'tarde': Font(name="Segoe UI", size=9, color="D97706", bold=True),
        'ausente': Font(name="Segoe UI", size=9, color="E11D48", bold=True),
        'justificado': Font(name="Segoe UI", size=9, color="3B82F6", bold=True),
    }
    state_abbrev = {
        'presente': 'P',
        'tarde': 'A',
        'ausente': 'F',
        'justificado': 'J',
    }
    
    # Formatear la fila de cabecera
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center if col_num > 2 else align_left
        cell.border = thin_border
        
    # Llenar datos de estudiantes
    current_row = 6
    for est in estudiantes:
        row_data = [est.nombre_completo, est.rut or "-"]
        
        total_dias = len(registros)
        asistencias_validas = 0
        
        for reg in registros:
            detalle = DetalleAsistencia.objects.filter(registro=reg, estudiante=est).first()
            if detalle:
                row_data.append(state_abbrev.get(detalle.estado, 'P'))
                if detalle.estado in ['presente', 'tarde', 'justificado']:
                    asistencias_validas += 1
            else:
                row_data.append("P")
                asistencias_validas += 1
                
        # Porcentaje de asistencia
        pct = 100.0 if total_dias == 0 else (asistencias_validas / total_dias) * 100
        row_data.append(f"{round(pct, 1)}%")
        
        ws.append(row_data)
        
        # Aplicar estilos a la fila de este estudiante
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = Font(name="Segoe UI", size=9)
            cell.border = thin_border
            
            if 3 <= col_num <= len(row_data) - 1:
                reg_index = col_num - 3
                reg = registros[reg_index]
                detalle = DetalleAsistencia.objects.filter(registro=reg, estudiante=est).first()
                estado = detalle.estado if detalle else 'presente'
                
                cell.fill = state_fills.get(estado, state_fills['presente'])
                cell.font = state_fonts.get(estado, state_fonts['presente'])
                cell.alignment = align_center
            elif col_num == len(row_data):
                cell.alignment = align_center
                cell.font = Font(name="Segoe UI", size=9, bold=True, color="7C5CFC" if pct >= 85 else "E11D48")
                if pct < 85:
                    cell.fill = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")
            else:
                cell.alignment = align_left
                
        current_row += 1
        
    # Autoajustar ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
    ws.column_dimensions['A'].width = 30
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="asistencia_{seccion.curso.nombre}_{seccion.letra}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


# ==============================================================================
# MÓDULO INTELIGENTE DE ASISTENCIA POR CÓDIGOS QR
# ==============================================================================
import io
import base64
import qrcode
from datetime import timedelta
from django.urls import reverse

@login_required
def asistencia_qr_sala_view(request, seccion_id):
    """
    Motor inteligente de escaneo de QR por Sala.
    Detecta el docente, hora actual, sección y horario para redirigir directamente
    a la lista de asistencia correspondiente o mostrar selector rápido.
    """
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    if not colegio:
        messages.warning(request, "No tienes un colegio asociado para registrar asistencia.")
        return redirect('solicitar_acceso')

    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    now = timezone.localtime(timezone.now()) if timezone.is_aware(timezone.now()) else timezone.now()
    hoy_date = now.date()
    dia_actual = now.weekday() # 0=Lunes .. 4=Viernes
    hora_actual = now.time()

    from colegios.models import EventoAgenda
    
    # 1. Buscar si el docente actual tiene clase programada para esta sección hoy
    clases_docente_hoy = EventoAgenda.objects.filter(
        colegio=colegio,
        tipo='clase',
        es_recurrente=True,
        dia_semana=dia_actual,
        asignado_a=request.user
    ).select_related('asignatura', 'curso')

    # Filtrar clase activa en ventana de tiempo (-25 min antes de inicio hasta +20 min después de fin)
    clase_activa = None
    for ev in clases_docente_hoy:
        if ev.fecha_inicio and ev.fecha_fin:
            ev_ini_t = timezone.localtime(ev.fecha_inicio).time() if timezone.is_aware(ev.fecha_inicio) else ev.fecha_inicio.time()
            ev_fin_t = timezone.localtime(ev.fecha_fin).time() if timezone.is_aware(ev.fecha_fin) else ev.fecha_fin.time()
            
            dt_ini_c = datetime.combine(hoy_date, ev_ini_t) - timedelta(minutes=25)
            dt_fin_c = datetime.combine(hoy_date, ev_fin_t) + timedelta(minutes=20)
            
            dt_actual = datetime.combine(hoy_date, hora_actual)
            if dt_ini_c <= dt_actual <= dt_fin_c:
                clase_activa = ev
                break

    # Si coincide con esta sección o tiene una clase activa
    if clase_activa and (clase_activa.curso_id == seccion.curso_id):
        asig_param = f"&asignatura={clase_activa.asignatura_id}" if clase_activa.asignatura_id else ""
        asig_nom = f" de {clase_activa.asignatura.nombre}" if clase_activa.asignatura else ""
        messages.success(request, f"⚡ ¡Identificado por QR! Accediendo a clase{asig_nom} con {seccion.nombre}.")
        return redirect(f"/asistencia/registrar/{seccion.id}/?fecha={hoy_date.strftime('%Y-%m-%d')}{asig_param}&qr_scan=1")

    # Si no hubo redirección directa automática, mostrar selector rápido 1-touch
    asignaturas_seccion = Asignatura.objects.filter(curso=seccion.curso, activo=True)
    asignaturas_docente = asignaturas_seccion.filter(docente=request.user)

    context = {
        'colegio': colegio,
        'seccion': seccion,
        'hoy': hoy_date,
        'hora_actual': hora_actual.strftime('%H:%M'),
        'clase_activa': clase_activa,
        'clases_docente_hoy': clases_docente_hoy,
        'asignaturas_seccion': asignaturas_seccion,
        'asignaturas_docente': asignaturas_docente,
        'is_admin': is_admin,
    }
    return render(request, 'asistencia/qr_selector.html', context)


@login_required
def asistencia_qr_express_view(request):
    """
    Escaneo rápido general / express.
    Encuentra la clase que le toca al profesor logueado en este minuto y lo lleva a tomar lista.
    """
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    if not colegio:
        messages.warning(request, "No tienes un colegio asociado.")
        return redirect('solicitar_acceso')

    now = timezone.localtime(timezone.now()) if timezone.is_aware(timezone.now()) else timezone.now()
    hoy_date = now.date()
    dia_actual = now.weekday()
    hora_actual = now.time()

    from colegios.models import EventoAgenda
    clases_docente_hoy = EventoAgenda.objects.filter(
        colegio=colegio,
        tipo='clase',
        es_recurrente=True,
        dia_semana=dia_actual,
        asignado_a=request.user
    ).select_related('asignatura', 'curso')

    clase_activa = None
    for ev in clases_docente_hoy:
        if ev.fecha_inicio and ev.fecha_fin:
            ev_ini_t = timezone.localtime(ev.fecha_inicio).time() if timezone.is_aware(ev.fecha_inicio) else ev.fecha_inicio.time()
            ev_fin_t = timezone.localtime(ev.fecha_fin).time() if timezone.is_aware(ev.fecha_fin) else ev.fecha_fin.time()
            
            dt_ini_c = datetime.combine(hoy_date, ev_ini_t) - timedelta(minutes=25)
            dt_fin_c = datetime.combine(hoy_date, ev_fin_t) + timedelta(minutes=20)
            
            dt_actual = datetime.combine(hoy_date, hora_actual)
            if dt_ini_c <= dt_actual <= dt_fin_c:
                clase_activa = ev
                break

    if clase_activa:
        seccion = SeccionCurso.objects.filter(curso=clase_activa.curso, activo=True).first()
        if seccion:
            asig_param = f"&asignatura={clase_activa.asignatura_id}" if clase_activa.asignatura_id else ""
            asig_nom = f" de {clase_activa.asignatura.nombre}" if clase_activa.asignatura else ""
            messages.success(request, f"⚡ ¡Horario detectado! Accediendo a tu clase{asig_nom} ({seccion.nombre}).")
            return redirect(f"/asistencia/registrar/{seccion.id}/?fecha={hoy_date.strftime('%Y-%m-%d')}{asig_param}&qr_scan=1")

    messages.info(request, "No se detectó un bloque de clase activo en este momento. Selecciona el curso manualmente.")
    return redirect('registrar_asistencia')


@login_required
def asistencia_qr_carteles_view(request):
    """
    Genera carteles imprimibles oficiales en alta resolución con códigos QR
    para cada una de las salas / secciones del colegio.
    """
    colegio, miembro, periodo, is_admin = obtener_datos_base_asistencia(request)
    if not colegio:
        messages.warning(request, "No tienes un colegio asociado.")
        return redirect('solicitar_acceso')

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')

    carteles = []
    for sec in secciones:
        qr_url = request.build_absolute_uri(reverse('asistencia_qr_sala', args=[sec.id]))

        # Generar imagen QR en base64
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="#1E1B4B", back_color="white")
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        qr_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

        carteles.append({
            'seccion': sec,
            'nombre_sala': f"Sala {sec.nombre}",
            'qr_url': qr_url,
            'qr_b64': qr_b64,
        })

    context = {
        'colegio': colegio,
        'carteles': carteles,
        'hoy': timezone.now().date(),
        'is_admin': is_admin,
    }
    return render(request, 'asistencia/qr_carteles.html', context)

