import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth import login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.utils import timezone

from usuarios.models import PerfilUsuario
from solicitudes.models import MiembroColegio
from .models import (
    Colegio, Suscripcion, ColegioModulo, ConfiguracionAcademica, 
    RolColegio, Permiso, RolPermiso, CursoColegio, SeccionCurso, 
    Asignatura, ConfiguracionModulos
)
from .forms import RegistroColegioPaso1Form, RegistroColegioPaso2Form



def registro_colegio_paso1_view(request):
    # Si hay un usuario logueado, cerramos su sesión inmediatamente para evitar 
    # que el colegio se asocie a una cuenta antigua por accidente.
    if request.user.is_authenticated:
        logout(request)

    if request.method == 'POST':
        form = RegistroColegioPaso1Form(request.POST)
        if form.is_valid():
            colegio = form.save(commit=False)
            
            # Como ya cerramos sesión arriba, siempre entraremos por aquí
            correo = form.cleaned_data['correo_institucional']
            password = form.cleaned_data['password']
            nombre_admin = form.cleaned_data['nombre_administrador']
            
            user = User.objects.create_user(username=correo, email=correo, password=password)
            PerfilUsuario.objects.create(
                usuario=user, 
                nombre_completo=nombre_admin,
                rut=form.cleaned_data.get('rut_administrador', ''),
                telefono=form.cleaned_data.get('telefono_administrador', '')
            )
            
            login(request, user)
            colegio.administrador = user
                
            colegio.save()
            return redirect('registro_colegio_paso2', colegio_id=colegio.id)
    else:
        form = RegistroColegioPaso1Form()
    
    return render(request, 'registrocolegio.html', {'form': form})

def registro_colegio_paso2_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)
    
    if request.method == 'POST':
        form = RegistroColegioPaso2Form(request.POST)
        if form.is_valid():
            plan = form.cleaned_data['plan']
            tipo_facturacion = form.cleaned_data['tipo_facturacion']
            
            # Calcular monto
            monto = plan.precio_anual if tipo_facturacion == 'anual' else plan.precio_mensual
            
            # 1. Crear la suscripción
            Suscripcion.objects.update_or_create(
                colegio=colegio,
                defaults={
                    'plan': plan,
                    'tipo_facturacion': tipo_facturacion,
                    'monto': monto,
                    'estado': 'pendiente_pago'
                }
            )
            
            # 2. Guardar los módulos activos
            modulos_seleccionados = form.cleaned_data['modulos']
            # Primero desactivamos todos los actuales si existen
            ColegioModulo.objects.filter(colegio=colegio).delete()
            for modulo in modulos_seleccionados:
                ColegioModulo.objects.create(colegio=colegio, modulo=modulo, activo=True)
            
            # 3. Clonar Roles Base para el colegio si no existen
            if not RolColegio.objects.filter(colegio=colegio).exists():
                roles_base = RolColegio.objects.filter(es_base=True, colegio=None)
                for rb in roles_base:
                    nuevo_rol = RolColegio.objects.create(
                        colegio=colegio,
                        nombre=rb.nombre,
                        descripcion=rb.descripcion,
                        es_base=False,
                        activo=True
                    )
                    # Clonar permisos del rol base
                    for rp in rb.permisos.all():
                        RolPermiso.objects.create(
                            rol=nuevo_rol,
                            modulo=rp.modulo,
                            puede_ver=rp.puede_ver,
                            puede_crear=rp.puede_crear,
                            puede_editar=rp.puede_editar,
                            puede_eliminar=rp.puede_eliminar,
                            puede_exportar=rp.puede_exportar,
                            puede_aprobar=rp.puede_aprobar,
                            puede_enviar_mensajes=rp.puede_enviar_mensajes,
                            puede_administrar=rp.puede_administrar
                        )
                    
                    # Si el rol es Administrador, vincular al usuario actual
                    if rb.nombre == 'Administrador':
                        MiembroColegio.objects.get_or_create(
                            usuario=request.user,
                            colegio=colegio,
                            defaults={'rol': nuevo_rol, 'activo': True}
                        )
            
            return redirect('configuracion_colegio_paso1', colegio_id=colegio.id)
    else:
        form = RegistroColegioPaso2Form()
        
    from planes.models import Plan, Modulo
    planes = Plan.objects.filter(activo=True).prefetch_related('modulos').order_by('precio_mensual')
    modulos = Modulo.objects.filter(activo=True).order_by('id')

    return render(request, 'registrocolegiopaso2.html', {
        'form': form, 
        'colegio': colegio,
        'planes': planes,
        'modulos': modulos,
        'modulos_disponibles': modulos
    })

def api_buscar_colegios(request):
    q = request.GET.get('q', '').strip()
    if q:
        colegios = Colegio.objects.filter(nombre__icontains=q)[:10]
    else:
        # Si no hay query, mostrar colegios activos recientes como sugerencia inicial
        colegios = Colegio.objects.all().order_by('-id')[:8]

    resultados = []
    for c in colegios:
        logo_url = c.logo.url if (c.logo and hasattr(c.logo, 'url')) else ''
        tipo_str = c.get_tipo_institucion_display() if hasattr(c, 'get_tipo_institucion_display') else (c.tipo_institucion or 'Colegio')
        resultados.append({
            'id': c.id,
            'nombre': c.nombre,
            'nombre_corto': c.nombre_corto or '',
            'eslogan': c.eslogan or '',
            'logo_url': logo_url,
            'color_principal': c.color_principal or '#7C5CFC',
            'ciudad_comuna': c.ciudad_comuna or '',
            'region': c.region or '',
            'tipo_institucion': tipo_str,
        })
    return JsonResponse(resultados, safe=False)

def configuracion_colegio_paso1_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    # Validación de seguridad: el usuario debe ser el administrador de ESTE colegio.
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    if request.method == 'POST':
        nombre = request.POST.get('nombre_oficial', '').strip()
        nombre_corto = request.POST.get('nombre_corto', '').strip()

        if not nombre or not nombre_corto:
            messages.error(request, "El nombre oficial y el nombre corto son campos obligatorios.")
            return render(request, 'configuracion_colegio_paso1.html', {'colegio': colegio})
        
        colegio.nombre = nombre
        colegio.nombre_corto = nombre_corto
        colegio.eslogan = request.POST.get('eslogan', '')

        if 'logo' in request.FILES:
            colegio.logo = request.FILES['logo']
        if 'imagen' in request.FILES:
            colegio.imagen_portada = request.FILES['imagen']
        
        # Guardar colores si vienen
        if request.POST.get('color_principal'):
            colegio.color_principal = request.POST.get('color_principal')
        
        colegio.paso_configuracion_actual = 2
        colegio.save()
        return redirect('configuracion_colegio_paso2', colegio_id=colegio.id)
        
    return render(request, 'configuracion_colegio_paso1.html', {'colegio': colegio})

def configuracion_colegio_paso2_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    # Validación de seguridad
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    if request.method == 'POST':
        correo = request.POST.get('correo', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        telefono = request.POST.get('telefono_principal', '').strip()
        comuna = request.POST.get('comuna', '').strip()

        if not correo or not direccion or not telefono or not comuna:
            messages.error(request, "Por favor complete todos los campos requeridos (*).")
            return render(request, 'configuracion_colegio_paso2.html', {'colegio': colegio})

        try:
            validate_email(correo)
        except ValidationError:
            messages.error(request, "Ingrese una dirección de correo institucional válida.")
            return render(request, 'configuracion_colegio_paso2.html', {'colegio': colegio})

        colegio.correo_institucional = correo
        colegio.direccion = direccion
        colegio.telefono = telefono
        colegio.ciudad_comuna = comuna
        colegio.telefono_alternativo = request.POST.get('telefono_alternativo', '')
        colegio.region = request.POST.get('region', '')
        colegio.sitio_web = request.POST.get('sitio_web', '')
        colegio.pais = request.POST.get('pais', '')
        colegio.referencia_direccion = request.POST.get('referencia', '')
        
        colegio.paso_configuracion_actual = 3
        colegio.save()
        return redirect('configuracion_colegio_paso3', colegio_id=colegio.id)
        
    return render(request, 'configuracion_colegio_paso2.html', {'colegio': colegio})

def configuracion_colegio_paso3_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    # Validación de seguridad
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    if request.method == 'POST':
        anio_academico = request.POST.get('anio_academico')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_termino = request.POST.get('fecha_termino')
        
        # Procesar cursos y secciones SIEMPRE, incluso si faltan fechas
        datos_cursos_json = request.POST.get('datos_cursos')
        if datos_cursos_json:
            try:
                lista_cursos = json.loads(datos_cursos_json)
                # Extraer cursos válidos y evitar borrar y recrear (para mantener IDs relacionales)
                cursos_mantener = []
                
                for c in lista_cursos:
                    nivel_key = 'basica'
                    if 'parvularia' in c['nivel'].lower(): nivel_key = 'parvularia'
                    elif 'media' in c['nivel'].lower(): nivel_key = 'media'
                    elif 'tp' in c['nivel'].lower() or 'téc' in c['nivel'].lower(): nivel_key = 'tecnico_profesional'
                    elif 'especial' in c['nivel'].lower(): nivel_key = 'especial'
                    
                    curso_nombre = c['curso'].strip()
                    curso_jornada = c.get('jornada', 'Mañana').strip()
                    
                    curso_obj, _ = CursoColegio.objects.update_or_create(
                        colegio=colegio, 
                        nombre=curso_nombre, 
                        jornada=curso_jornada,
                        defaults={
                            'nivel': nivel_key, 
                            'desde_letra': c['letra_desde'].upper(), 
                            'hasta_letra': c['letra_hasta'].upper(),
                            'activo': True
                        }
                    )
                    cursos_mantener.append(curso_obj.id)
                    
                    # Generar secciones
                    start = ord(curso_obj.desde_letra)
                    end = ord(curso_obj.hasta_letra)
                    
                    if start <= end:
                        # Desactivar secciones viejas que ya no estén en el rango
                        SeccionCurso.objects.filter(curso=curso_obj).update(activo=False)
                        
                        for i in range(start, end + 1):
                            letra = chr(i)
                            SeccionCurso.objects.update_or_create(
                                curso=curso_obj,
                                letra=letra,
                                defaults={'nombre': f"{curso_obj.nombre} {letra}"[:150], 'activo': True}
                            )

                
                # Eliminar cursos que el usuario quitó explícitamente de la tabla
                CursoColegio.objects.filter(colegio=colegio).exclude(id__in=cursos_mantener).delete()

            except json.JSONDecodeError:
                pass
                
        if anio_academico and fecha_inicio and fecha_termino:
            periodo, created = ConfiguracionAcademica.objects.update_or_create(
                colegio=colegio,
                defaults={
                    'anio_academico': anio_academico,
                    'fecha_inicio': fecha_inicio,
                    'fecha_termino': fecha_termino,
                    'periodo_academico': request.POST.get('tipo_periodo', 'semestres').lower(),
                    'horario_referencial': request.POST.get('horario_referencia', '')
                }
            )
            
            # Guardar flags de niveles
            niveles_raw = request.POST.get('niveles_educativos', '').lower()
            periodo.nivel_parvularia = 'parvularia' in niveles_raw
            periodo.nivel_basica = 'básica' in niveles_raw
            periodo.nivel_media = 'media' in niveles_raw
            periodo.nivel_tecnico_profesional = 'técnico' in niveles_raw
            periodo.nivel_especial = 'especial' in niveles_raw
            periodo.nivel_otro = 'otro' in niveles_raw
            
            # Guardar flags de jornadas
            jornadas_raw = request.POST.get('jornadas', '').lower()
            periodo.jornada_manana = 'mañana' in jornadas_raw
            periodo.jornada_tarde = 'tarde' in jornadas_raw
            periodo.jornada_completa = 'completa' in jornadas_raw
            periodo.jornada_vespertina = 'vespertina' in jornadas_raw
            periodo.jornada_flexible = 'flexible' in jornadas_raw
            periodo.save()
            
            colegio.paso_configuracion_actual = 4
            colegio.save()
            return redirect('configuracion_colegio_paso4', colegio_id=colegio.id)
        else:
            messages.error(request, "El año, fecha de inicio y término son obligatorios.")

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    cursos = CursoColegio.objects.filter(colegio=colegio).order_by('nivel', 'nombre')
    
    # Preparar el JSON de cursos de forma segura para evitar errores de parseo en JS
    cursos_data = []
    for c in cursos:
        nivel_label = 'TP'
        if c.nivel == 'parvularia': nivel_label = 'Parvularia'
        elif c.nivel == 'basica': nivel_label = 'Básica'
        elif c.nivel == 'media': nivel_label = 'Media'
        
        cursos_data.append({
            'curso': c.nombre.replace('\n', ' ').replace('\r', '').strip(),
            'nivel': nivel_label,
            'jornada': c.jornada,
            'desde_letra': c.desde_letra,
            'hasta_letra': c.hasta_letra
        })

    return render(request, 'configuracion_colegio_paso3.html', {'colegio': colegio, 'periodo': periodo, 'cursos': cursos, 'cursos_data': cursos_data})

def configuracion_colegio_paso4_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    # Validación de seguridad
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    # Asegurar que existan los roles base institucionales
    asegurar_roles_base_colegio(colegio, request.user)

    # Asegurar módulos activos base si no existen aún
    if not ColegioModulo.objects.filter(colegio=colegio).exists():
        from planes.models import Modulo
        for mod in Modulo.objects.all():
            ColegioModulo.objects.get_or_create(colegio=colegio, modulo=mod, defaults={'activo': True})

    roles = RolColegio.objects.filter(colegio=colegio).exclude(nombre__iexact='Apoderado').order_by('id')
    modulos = ColegioModulo.objects.filter(colegio=colegio, activo=True).select_related('modulo').order_by('modulo__nombre')

    if request.method == 'POST':
        permissions_data = request.POST.get('permissions_data')
        if permissions_data:
            try:
                data = json.loads(permissions_data)
                for rol_nombre, rol_info in data.items():
                    rol_obj = RolColegio.objects.filter(colegio=colegio, nombre=rol_nombre).first()
                    if rol_obj:
                        modules_data = rol_info.get('modules', {})
                        for mod_nombre, perms in modules_data.items():
                            from planes.models import Modulo
                            modulo_obj = Modulo.objects.filter(nombre__iexact=mod_nombre).first()
                            if modulo_obj:
                                rp, _ = RolPermiso.objects.get_or_create(rol=rol_obj, modulo=modulo_obj)
                                rp.puede_ver = perms.get('view', False)
                                rp.puede_crear = perms.get('create', False)
                                rp.puede_editar = perms.get('edit', False)
                                rp.puede_exportar = perms.get('export', False)
                                rp.save()
            except json.JSONDecodeError:
                messages.error(request, "Error procesando la configuración de permisos.")
                return redirect('configuracion_colegio_paso4', colegio_id=colegio.id)

        colegio.paso_configuracion_actual = 5
        colegio.save()
        return redirect('configuracion_colegio_paso5', colegio_id=colegio.id)
        
    # Preparar permisos guardados en JSON para inicializar la vista de forma reactiva
    saved_permissions = {}
    for rol in roles:
        rol_perms = {}
        for rp in rol.permisos.all():
            rol_perms[rp.modulo.nombre] = {
                'view': rp.puede_ver,
                'create': rp.puede_crear,
                'edit': rp.puede_editar,
                'export': rp.puede_exportar,
            }
        if rol_perms:
            saved_permissions[rol.nombre] = rol_perms
            
    return render(request, 'configuracion_colegio_paso4.html', {
        'colegio': colegio, 
        'roles': roles, 
        'modulos': modulos,
        'saved_permissions_json': json.dumps(saved_permissions)
    })

def configuracion_colegio_paso5_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    # Validación de seguridad
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    if request.method == 'POST':
        modulos_colegio = ColegioModulo.objects.filter(colegio=colegio)
        for mod_col in modulos_colegio:
            estado = request.POST.get(f'modulo_{mod_col.modulo.id}')
            mod_col.activo = True if estado else False
            mod_col.save()
        
        colegio.paso_configuracion_actual = 6
        colegio.save()
        return redirect('configuracion_colegio_paso6', colegio_id=colegio.id)
    
    modulos = ColegioModulo.objects.filter(colegio=colegio)
    return render(request, 'configuracion_colegio_paso5.html', {'modulos_colegio': modulos, 'colegio': colegio})


def configuracion_colegio_paso6_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)

    periodo, _ = ConfiguracionAcademica.objects.get_or_create(
        colegio=colegio,
        defaults={
            'anio_academico': timezone.now().year,
            'fecha_inicio': timezone.now().date(),
            'fecha_termino': timezone.now().date(),
        }
    )

    if request.method == 'POST':
        periodo.modalidad_asistencia = request.POST.get('modalidad_asistencia', 'asignatura')
        try:
            periodo.porcentaje_asistencia_minima = int(request.POST.get('porcentaje_asistencia_minima', 85))
        except ValueError:
            periodo.porcentaje_asistencia_minima = 85

        periodo.tipo_calificacion = request.POST.get('tipo_calificacion', 'numerica')
        try:
            periodo.nota_minima_aprobacion = float(request.POST.get('nota_minima_aprobacion', 4.0))
        except ValueError:
            periodo.nota_minima_aprobacion = 4.0

        try:
            periodo.porcentaje_exigencia = int(request.POST.get('porcentaje_exigencia', 60))
        except ValueError:
            periodo.porcentaje_exigencia = 60

        periodo.regla_redondeo = request.POST.get('regla_redondeo', 'un_decimal')
        periodo.tipo_calculo_promedio = request.POST.get('tipo_calculo_promedio', 'ponderado')
        periodo.visibilidad_notas_apoderados = request.POST.get('visibilidad_notas_apoderados', 'inmediata')
        periodo.notificar_ausencias = (request.POST.get('notificar_ausencias') == 'on')
        periodo.notificar_notas_rojas = (request.POST.get('notificar_notas_rojas') == 'on')
        periodo.save()

        # Completar la configuración institucional del colegio
        colegio.configuracion_completa = True
        colegio.estado = 'activo'
        colegio.save()

        return redirect('configuracion_colegio_finalizando', colegio_id=colegio.id)

    return render(request, 'configuracion_colegio_paso6.html', {
        'colegio': colegio,
        'periodo': periodo,
    })


def configuracion_colegio_finalizando_view(request, colegio_id):
    if not request.user.is_authenticated:
        return redirect('registro_colegio')
    colegio = get_object_or_404(Colegio, id=colegio_id, administrador=request.user)
    return render(request, 'configuracion_colegio_finalizando.html', {
        'colegio': colegio,
    })


from django.db.models import Q
from .models import Estudiante, Asignatura

def obtener_colegio_usuario(user):
    colegio = user.colegios_administrados.order_by('-fecha_creacion').first()
    if not colegio:
        miembro = MiembroColegio.objects.filter(usuario=user, activo=True).order_by('-fecha_ingreso').first()
        if miembro:
            colegio = miembro.colegio
    if not colegio and (getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False)):
        colegio = Colegio.objects.filter(estado='activo').order_by('-fecha_creacion').first()
        if not colegio:
            colegio = Colegio.objects.order_by('-fecha_creacion').first()
    if colegio:
        asegurar_roles_base_colegio(colegio, user)
    return colegio


@login_required
def listar_estudiantes_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    # Filtrado por curso/seccion
    seccion_id = request.GET.get('seccion')
    busqueda = request.GET.get('q', '').strip()

    estudiantes = Estudiante.objects.filter(colegio=colegio).order_by('seccion__curso__nombre', 'nombre_completo')

    if seccion_id and seccion_id.isdigit():
        estudiantes = estudiantes.filter(seccion_id=int(seccion_id))

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(nombre_completo__icontains=busqueda) | 
            Q(rut__icontains=busqueda)
        )

    # Secciones para el dropdown de filtros
    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('curso__nombre', 'nombre')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])

    total_matriculados = estudiantes.count()
    total_activos = estudiantes.filter(activo=True).count()
    total_pie = estudiantes.filter(es_pie=True).count()
    total_baja = estudiantes.filter(activo=False).count()

    paginator = Paginator(estudiantes, 10)
    page_number = request.GET.get('page', 1)
    estudiantes_page = paginator.get_page(page_number)

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'estudiantes': estudiantes_page,
        'secciones': secciones,
        'seccion_seleccionada': int(seccion_id) if seccion_id and seccion_id.isdigit() else None,
        'busqueda': busqueda,
        'total_matriculados': total_matriculados,
        'total_activos': total_activos,
        'total_pie': total_pie,
        'total_baja': total_baja,
    }
    return render(request, 'colegios/listar_estudiantes.html', context)




@login_required
def matricular_estudiante_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes permisos de administrador.")
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    rol_str = (miembro.rol.nombre.lower() if miembro and miembro.rol else '')
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or any(r in rol_str for r in ['administrador', 'director', 'secretari', 'administrativ', 'admision', 'admisión'])
    )
    if not is_admin:
        messages.error(request, "No tienes permisos para matricular estudiantes.")
        return redirect('listar_estudiantes')

    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo', '').strip()
        rut = request.POST.get('rut', '').strip()
        seccion_id = request.POST.get('seccion')

        # Personal
        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        genero = request.POST.get('genero', 'no_informa')
        nacionalidad = request.POST.get('nacionalidad', 'Chilena').strip()
        direccion = request.POST.get('direccion', '').strip()
        comuna = request.POST.get('comuna', '').strip()

        # Salud y emergencia
        prevision_salud = request.POST.get('prevision_salud', 'fonasa')
        grupo_sanguineo = request.POST.get('grupo_sanguineo', '').strip()
        alergias_enfermedades = request.POST.get('alergias_enfermedades', '').strip()
        contacto_emergencia_nombre = request.POST.get('contacto_emergencia_nombre', '').strip()
        contacto_emergencia_parentesco = request.POST.get('contacto_emergencia_parentesco', '').strip()
        contacto_emergencia_telefono = request.POST.get('contacto_emergencia_telefono', '').strip()

        # Apoderado
        nombre_apoderado = request.POST.get('nombre_apoderado', '').strip()
        rut_apoderado = request.POST.get('rut_apoderado', '').strip()
        telefono_apoderado = request.POST.get('telefono_apoderado', '').strip()
        email_apoderado = request.POST.get('email_apoderado', '').strip()
        parentesco_apoderado = request.POST.get('parentesco_apoderado', '').strip()

        # PIE
        es_pie = request.POST.get('es_pie') == 'on' or request.POST.get('es_pie') == 'true'
        tipo_pie = request.POST.get('tipo_pie', '')
        diagnostico_pie = request.POST.get('diagnostico_pie', '').strip()
        observaciones_pie = request.POST.get('observaciones_pie', '').strip()

        if not nombre_completo or not seccion_id:
            messages.error(request, "El nombre completo y el curso/sección son requeridos.")
        else:
            seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
            Estudiante.objects.create(
                colegio=colegio,
                seccion=seccion,
                nombre_completo=nombre_completo,
                rut=rut,
                fecha_nacimiento=fecha_nacimiento,
                genero=genero,
                nacionalidad=nacionalidad,
                direccion=direccion,
                comuna=comuna,
                prevision_salud=prevision_salud,
                grupo_sanguineo=grupo_sanguineo,
                alergias_enfermedades=alergias_enfermedades,
                contacto_emergencia_nombre=contacto_emergencia_nombre,
                contacto_emergencia_parentesco=contacto_emergencia_parentesco,
                contacto_emergencia_telefono=contacto_emergencia_telefono,
                nombre_apoderado=nombre_apoderado,
                rut_apoderado=rut_apoderado,
                telefono_apoderado=telefono_apoderado,
                email_apoderado=email_apoderado,
                parentesco_apoderado=parentesco_apoderado,
                es_pie=es_pie,
                tipo_pie=tipo_pie if es_pie else None,
                diagnostico_pie=diagnostico_pie if es_pie else None,
                observaciones_pie=observaciones_pie if es_pie else None,
                activo=True
            )
            messages.success(request, f"Estudiante {nombre_completo} matriculado con éxito.")
            return redirect('listar_estudiantes')

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('curso__nombre', 'nombre')
    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'secciones': secciones,
        'titulo_pagina': 'Matricular Estudiante',
        'boton_texto': 'Matricular Estudiante',
    }
    return render(request, 'colegios/matricular_estudiante.html', context)


# ── CARGA MASIVA DE ESTUDIANTES (EXCEL INTELIGENTE Y PRE-VALIDACIÓN) ──────────

def validar_rut_chileno(rut_str):
    """Valida y formatea un RUT chileno usando el algoritmo Módulo 11."""
    if not rut_str:
        return False, ""
    
    limpio = str(rut_str).strip().replace('.', '').replace('-', '').upper()
    if len(limpio) < 8 or len(limpio) > 9:
        return False, str(rut_str).strip()
    
    cuerpo = limpio[:-1]
    dv = limpio[-1]
    
    if not cuerpo.isdigit():
        return False, str(rut_str).strip()
    
    suma = 0
    multiplicador = 2
    for c in reversed(cuerpo):
        suma += int(c) * multiplicador
        multiplicador = 2 if multiplicador == 7 else multiplicador + 1
    
    esperado = 11 - (suma % 11)
    if esperado == 11:
        dv_esperado = '0'
    elif esperado == 10:
        dv_esperado = 'K'
    else:
        dv_esperado = str(esperado)
    
    valido = (dv == dv_esperado)
    cuerpo_fmt = f"{int(cuerpo):,}".replace(',', '.')
    rut_formateado = f"{cuerpo_fmt}-{dv}"
    return valido, rut_formateado


def normalizar_fecha_excel(fecha_val):
    """Normaliza fechas desde Excel o string a formato YYYY-MM-DD."""
    if not fecha_val:
        return None
    from datetime import date as dt_date, datetime as dt_datetime
    if isinstance(fecha_val, dt_datetime):
        return fecha_val.date().strftime('%Y-%m-%d')
    if isinstance(fecha_val, dt_date):
        return fecha_val.strftime('%Y-%m-%d')
    
    s = str(fecha_val).strip()
    formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%Y/%m/%d', '%d.%m.%Y']
    for fmt in formatos:
        try:
            return dt_datetime.strptime(s, fmt).date().strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def buscar_seccion_colegio(colegio, nombre_seccion_str):
    """Encuentra la sección más adecuada en base al texto del Excel."""
    if not nombre_seccion_str:
        return None
    
    texto = str(nombre_seccion_str).strip()
    # 1. Búsqueda exacta
    seccion = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True, nombre__iexact=texto).first()
    if seccion:
        return seccion
    
    # 2. Búsqueda por partes: "1° Básico A" -> Curso "1° Básico" y Letra "A"
    partes = texto.rsplit(' ', 1)
    if len(partes) == 2:
        nombre_c, letra_c = partes[0].strip(), partes[1].strip().upper()
        sec = SeccionCurso.objects.filter(
            curso__colegio=colegio, 
            activo=True, 
            curso__nombre__iexact=nombre_c, 
            letra__iexact=letra_c
        ).first()
        if sec:
            return sec
    
    # 3. Búsqueda flexible sin acentos/símbolos
    import unicodedata
    def limpiar_str(s):
        s = s.replace('°', '').replace('º', '').replace('-', ' ')
        return unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8').lower().strip()
    
    texto_norm = limpiar_str(texto)
    for sec in SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso'):
        if limpiar_str(sec.nombre) == texto_norm or limpiar_str(f"{sec.curso.nombre} {sec.letra}") == texto_norm:
            return sec
            
    return None


@login_required
def descargar_plantilla_estudiantes_excel_view(request):
    """Genera dinámicamente un archivo Excel .xlsx inteligente con listas desplegables y formato institucional."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes permisos de acceso.")
        return redirect('dashboard_usuario')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carga_Estudiantes"
    ws.views.sheetView[0].showGridLines = True

    # Estilos Visuales Premium
    FILL_HEADER = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FONT_REQUIRED = Font(name="Segoe UI", size=10, bold=True, color="FDE047") # Amarillo dorado
    
    FILL_EXAMPLE = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    FONT_EXAMPLE = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    
    BORDER_THIN = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    headers = [
        ("* RUT Estudiante", 18, True),
        ("* Nombres", 24, True),
        ("* Apellidos", 24, True),
        ("* Curso y Sección", 22, True),
        ("Fecha Nacimiento (DD/MM/AAAA)", 28, False),
        ("Género", 18, False),
        ("Dirección", 30, False),
        ("Comuna", 20, False),
        ("Nombre Apoderado", 26, False),
        ("RUT Apoderado", 18, False),
        ("Teléfono Apoderado", 20, False),
        ("Email Apoderado", 26, False),
        ("Parentesco Apoderado", 20, False),
        ("¿Es PIE? (Sí/No)", 18, False),
        ("Diagnóstico PIE", 24, False),
    ]

    ws.row_dimensions[1].height = 36

    for col_idx, (titulo, ancho, es_obligatorio) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = FILL_HEADER
        cell.font = FONT_REQUIRED if es_obligatorio else FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Secciones disponibles en el colegio
    secciones = list(SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra'))
    nombres_secciones = [s.nombre if s.nombre else f"{s.curso.nombre} {s.letra}".strip() for s in secciones]
    if not nombres_secciones:
        nombres_secciones = ["1° Básico A", "1° Básico B", "2° Básico A"]

    # Filas de Ejemplo
    ejemplo_1 = [
        "21.432.543-9", "Martina Sofía", "González Pérez",
        nombres_secciones[0] if nombres_secciones else "1° Básico A",
        "15/03/2014", "Femenino", "Av. Los Pajaritos 1234", "Maipú",
        "Carlos González Ramos", "14.234.567-8", "+56 9 9123 4567", "carlos.gonzalez@correo.cl", "Papá",
        "No", ""
    ]
    ejemplo_2 = [
        "22.198.765-4", "Mateo Ignacio", "Silva Contreras",
        nombres_secciones[0] if nombres_secciones else "1° Básico A",
        "22/07/2014", "Masculino", "Calle Las Flores 456", "Santiago",
        "Andrea Contreras Soto", "15.987.654-2", "+56 9 8765 4321", "andrea.contreras@correo.cl", "Mamá",
        "Sí", "TDAH"
    ]

    ws.append(ejemplo_1)
    ws.append(ejemplo_2)

    for r in range(2, 4):
        ws.row_dimensions[r].height = 24
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_EXAMPLE
            cell.font = FONT_EXAMPLE
            cell.alignment = Alignment(horizontal="center" if c in [1, 4, 5, 6, 10, 11, 14] else "left", vertical="center")
            cell.border = BORDER_THIN

    # Pestaña Oculta con Parámetros para desplegables de Excel
    ws_params = wb.create_sheet(title="_Parametros")
    ws_params['A1'] = "Cursos_Disponibles"
    for i, nom in enumerate(nombres_secciones, start=2):
        ws_params[f'A{i}'] = nom

    ws_params['B1'] = "Generos"
    for i, g in enumerate(["Masculino", "Femenino", "Otro", "Prefiero no decir"], start=2):
        ws_params[f'B{i}'] = g

    ws_params['C1'] = "PIE"
    for i, p in enumerate(["Sí", "No"], start=2):
        ws_params[f'C{i}'] = p

    ws_params.sheet_state = 'hidden'

    # Validaciones en hoja principal
    max_sec = len(nombres_secciones) + 1
    dv_curso = DataValidation(type="list", formula1=f"=_Parametros!$A$2:$A${max_sec}", allow_blank=True)
    ws.add_data_validation(dv_curso)
    dv_curso.add("D4:D500")

    dv_genero = DataValidation(type="list", formula1="=_Parametros!$B$2:$B$5", allow_blank=True)
    ws.add_data_validation(dv_genero)
    dv_genero.add("F4:F500")

    dv_pie = DataValidation(type="list", formula1="=_Parametros!$C$2:$C$3", allow_blank=True)
    ws.add_data_validation(dv_pie)
    dv_pie.add("N4:N500")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    slug_nombre = colegio.nombre.lower().replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="Plantilla_Estudiantes_{slug_nombre}.xlsx"'
    wb.save(response)
    return response


@login_required
def carga_masiva_estudiantes_view(request):
    """Renderiza la vista principal del Asistente de Carga Masiva de Estudiantes."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes permisos de acceso.")
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para realizar cargas masivas.")
        return redirect('listar_estudiantes')

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'secciones': secciones,
        'secciones_count': secciones.count(),
    }
    return render(request, 'colegios/carga_masiva_estudiantes.html', context)


@login_required
@require_POST
def api_analizar_archivo_estudiantes(request):
    """Endpoint AJAX: Analiza y pre-valida el archivo Excel/CSV subido antes de guardar."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    archivo = request.FILES.get('archivo')
    if not archivo:
        return JsonResponse({'error': 'No se adjuntó ningún archivo.'}, status=400)

    nombre_archivo = archivo.name.lower()
    filas_raw = []

    try:
        if nombre_archivo.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            import openpyxl
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if any(row):
                    filas_raw.append([str(c).strip() if c is not None else '' for c in row])
        elif nombre_archivo.endswith('.csv'):
            import csv
            import io
            decoded_file = archivo.read().decode('utf-8-sig', errors='replace')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string, delimiter=';' if ';' in decoded_file[:200] else ',')
            for row in reader:
                if any(row):
                    filas_raw.append([c.strip() for c in row])
        else:
            return JsonResponse({'error': 'Formato no soportado. Por favor sube un archivo Excel (.xlsx) o CSV.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error al leer el archivo: {str(e)}'}, status=400)

    if len(filas_raw) < 2:
        return JsonResponse({'error': 'El archivo no contiene filas de datos suficientes.'}, status=400)

    # Identificar fila de cabeceras
    header_idx = 0
    headers = [c.lower() for c in filas_raw[0]]
    if not any('rut' in h or 'nombre' in h for h in headers) and len(filas_raw) > 1:
        header_idx = 1
        headers = [c.lower() for c in filas_raw[1]]

    def get_col_index(keywords):
        for i, h in enumerate(headers):
            for kw in keywords:
                if kw in h:
                    return i
        return -1

    idx_rut = get_col_index(['rut estudiante', 'rut alumno', 'rut'])
    idx_nombres = get_col_index(['nombres', 'nombre'])
    idx_apellidos = get_col_index(['apellidos', 'apellido'])
    idx_curso = get_col_index(['curso', 'seccion', 'sección', 'grado', 'nivel'])
    idx_nac = get_col_index(['nacimiento', 'fecha nac', 'f. nac'])
    idx_genero = get_col_index(['genero', 'género', 'sexo'])
    idx_direccion = get_col_index(['direccion', 'dirección', 'domicilio'])
    idx_comuna = get_col_index(['comuna', 'ciudad'])
    idx_apoderado = get_col_index(['nombre apoderado', 'apoderado'])
    idx_rut_apo = get_col_index(['rut apoderado'])
    idx_tel_apo = get_col_index(['telefono', 'teléfono', 'celular', 'fono', 'contacto'])
    idx_email_apo = get_col_index(['email', 'correo', 'mail'])
    idx_parentesco = get_col_index(['parentesco', 'relacion'])
    idx_pie = get_col_index(['pie', 'es pie', 'integracion'])
    idx_diag_pie = get_col_index(['diagnostico', 'diagnóstico', 'tipo pie'])

    ruts_en_bd = set(Estudiante.objects.filter(colegio=colegio, activo=True).exclude(rut__isnull=True).exclude(rut='').values_list('rut', flat=True))
    ruts_limpios_bd = {r.replace('.', '').replace('-', '').upper(): r for r in ruts_en_bd}

    filas_analizadas = []
    ruts_vistos_en_archivo = {}
    filas_datos = filas_raw[header_idx + 1:]

    # Descartar filas de ejemplo si el usuario no las borró pero agregó datos propios
    filas_reales = [f for f in filas_datos if '21.432.543-9' not in str(f) and '22.198.765-4' not in str(f)]
    if filas_reales:
        filas_datos = filas_reales

    for index, row in enumerate(filas_datos, start=1):
        if not any(row):
            continue

        def get_val(col_i):
            return row[col_i].strip() if 0 <= col_i < len(row) else ''

        rut_raw = get_val(idx_rut)
        nombres_raw = get_val(idx_nombres)
        apellidos_raw = get_val(idx_apellidos)
        curso_raw = get_val(idx_curso)
        fecha_nac_raw = get_val(idx_nac)
        genero_raw = get_val(idx_genero)
        direccion_raw = get_val(idx_direccion)
        comuna_raw = get_val(idx_comuna)
        nombre_apo_raw = get_val(idx_apoderado)
        rut_apo_raw = get_val(idx_rut_apo)
        tel_apo_raw = get_val(idx_tel_apo)
        email_apo_raw = get_val(idx_email_apo)
        parentesco_raw = get_val(idx_parentesco)
        pie_raw = get_val(idx_pie)
        diag_pie_raw = get_val(idx_diag_pie)

        if nombres_raw and apellidos_raw:
            nombre_completo = f"{nombres_raw} {apellidos_raw}".strip()
        elif nombres_raw:
            nombre_completo = nombres_raw.strip()
        else:
            nombre_completo = ''

        errores = []
        advertencias = []

        # 1. Nombre
        if not nombre_completo:
            errores.append("El nombre del estudiante es obligatorio.")

        # 2. RUT Estudiante
        rut_formateado = rut_raw
        if rut_raw:
            es_valido_rut, rut_formateado = validar_rut_chileno(rut_raw)
            if not es_valido_rut:
                advertencias.append(f"El RUT '{rut_raw}' no coincide con el dígito verificador chileno.")
            
            rut_limpio = rut_raw.replace('.', '').replace('-', '').upper()
            if rut_limpio in ruts_vistos_en_archivo:
                advertencias.append(f"RUT duplicado en la fila {ruts_vistos_en_archivo[rut_limpio]} del archivo.")
            else:
                ruts_vistos_en_archivo[rut_limpio] = index

            if rut_limpio in ruts_limpios_bd:
                advertencias.append(f"RUT ya registrado ({ruts_limpios_bd[rut_limpio]}). Se actualizará la matrícula.")
        else:
            advertencias.append("Sin RUT registrado.")

        # 3. Sección
        seccion_obj = buscar_seccion_colegio(colegio, curso_raw)
        seccion_id = seccion_obj.id if seccion_obj else None
        seccion_nombre = (seccion_obj.nombre if seccion_obj.nombre else f"{seccion_obj.curso.nombre} {seccion_obj.letra}".strip()) if seccion_obj else curso_raw

        if not seccion_obj:
            if curso_raw:
                errores.append(f"Curso '{curso_raw}' no encontrado en el colegio. Selecciónalo manualmente.")
            else:
                errores.append("Falta asignar el curso y sección.")

        # 4. Género
        genero_norm = 'no_informa'
        g_lower = genero_raw.lower()
        if 'masc' in g_lower or g_lower == 'm' or g_lower == 'hombre':
            genero_norm = 'masculino'
        elif 'fem' in g_lower or g_lower == 'f' or g_lower == 'mujer':
            genero_norm = 'femenino'
        elif 'otro' in g_lower or 'no bin' in g_lower:
            genero_norm = 'otro'

        # 5. Fecha Nacimiento
        fecha_nac_norm = normalizar_fecha_excel(fecha_nac_raw)
        if fecha_nac_raw and not fecha_nac_norm:
            advertencias.append(f"Fecha '{fecha_nac_raw}' no válida (use DD/MM/AAAA).")

        # 6. PIE
        es_pie = True if pie_raw.lower() in ['si', 'sí', 'true', '1', 's', 'x'] else False

        if errores:
            estado = 'error'
        elif advertencias:
            estado = 'advertencia'
        else:
            estado = 'valido'

        filas_analizadas.append({
            'index': index,
            'estado': estado,
            'errores': errores,
            'advertencias': advertencias,
            'rut': rut_formateado,
            'nombre_completo': nombre_completo,
            'seccion_id': seccion_id,
            'seccion_nombre': seccion_nombre,
            'fecha_nacimiento': fecha_nac_norm,
            'genero': genero_norm,
            'direccion': direccion_raw,
            'comuna': comuna_raw,
            'nombre_apoderado': nombre_apo_raw,
            'rut_apoderado': rut_apo_raw,
            'telefono_apoderado': tel_apo_raw,
            'email_apoderado': email_apo_raw,
            'parentesco_apoderado': parentesco_raw or 'Apoderado',
            'es_pie': es_pie,
            'diagnostico_pie': diag_pie_raw,
        })

    secciones_disponibles = [{'id': s.id, 'nombre': s.nombre if s.nombre else f"{s.curso.nombre} {s.letra}".strip()} for s in SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')]

    return JsonResponse({
        'total': len(filas_analizadas),
        'validos': sum(1 for f in filas_analizadas if f['estado'] == 'valido'),
        'advertencias': sum(1 for f in filas_analizadas if f['estado'] == 'advertencia'),
        'errores': sum(1 for f in filas_analizadas if f['estado'] == 'error'),
        'filas': filas_analizadas,
        'secciones_disponibles': secciones_disponibles
    })


@login_required
@require_POST
def api_procesar_carga_masiva_estudiantes(request):
    """Endpoint AJAX: Guarda masivamente en base de datos los estudiantes confirmados."""
    import json
    from django.db import transaction

    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
        filas = data.get('filas', [])
        politica_duplicados = data.get('politica_duplicados', 'actualizar')
    except Exception as e:
        return JsonResponse({'error': 'Datos inválidos.'}, status=400)

    if not filas:
        return JsonResponse({'error': 'No hay filas para procesar.'}, status=400)

    creados_count = 0
    actualizados_count = 0
    omitidos_count = 0

    try:
        with transaction.atomic():
            for item in filas:
                nombre = item.get('nombre_completo', '').strip()
                seccion_id = item.get('seccion_id')
                rut = item.get('rut', '').strip()

                if not nombre or not seccion_id:
                    omitidos_count += 1
                    continue

                seccion = SeccionCurso.objects.filter(id=seccion_id, curso__colegio=colegio).first()
                if not seccion:
                    omitidos_count += 1
                    continue

                defaults_data = {
                    'nombre_completo': nombre,
                    'seccion': seccion,
                    'fecha_nacimiento': item.get('fecha_nacimiento') or None,
                    'genero': item.get('genero', 'no_informa'),
                    'direccion': item.get('direccion', '').strip() or None,
                    'comuna': item.get('comuna', '').strip() or None,
                    'nombre_apoderado': item.get('nombre_apoderado', '').strip() or None,
                    'rut_apoderado': item.get('rut_apoderado', '').strip() or None,
                    'telefono_apoderado': item.get('telefono_apoderado', '').strip() or None,
                    'email_apoderado': item.get('email_apoderado', '').strip() or None,
                    'parentesco_apoderado': item.get('parentesco_apoderado', '').strip() or 'Apoderado',
                    'es_pie': bool(item.get('es_pie')),
                    'diagnostico_pie': item.get('diagnostico_pie', '').strip() or None,
                    'activo': True,
                }

                if rut:
                    est_existente = Estudiante.objects.filter(colegio=colegio, rut=rut).first()
                    if est_existente:
                        if politica_duplicados == 'actualizar':
                            for k, v in defaults_data.items():
                                setattr(est_existente, k, v)
                            est_existente.save()
                            actualizados_count += 1
                        else:
                            omitidos_count += 1
                        continue
                    
                    Estudiante.objects.create(
                        colegio=colegio,
                        rut=rut,
                        **defaults_data
                    )
                    creados_count += 1
                else:
                    Estudiante.objects.create(
                        colegio=colegio,
                        rut=None,
                        **defaults_data
                    )
                    creados_count += 1

    except Exception as e:
        return JsonResponse({'error': f'Error durante el guardado: {str(e)}'}, status=500)

    return JsonResponse({
        'success': True,
        'creados': creados_count,
        'actualizados': actualizados_count,
        'omitidos': omitidos_count,
        'total_procesados': creados_count + actualizados_count
    })


@login_required
def editar_estudiante_view(request, estudiante_id):
    colegio = obtener_colegio_usuario(request.user)
    estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    rol_str = (miembro.rol.nombre.lower() if miembro and miembro.rol else '')
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or any(r in rol_str for r in ['administrador', 'director', 'secretari', 'administrativ', 'admision', 'admisión'])
    )
    if not is_admin:
        messages.error(request, "No tienes permisos para editar estudiantes.")
        return redirect('listar_estudiantes')

    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo', '').strip()
        rut = request.POST.get('rut', '').strip()
        seccion_id = request.POST.get('seccion')
        activo = request.POST.get('activo') == 'true'

        # Personal
        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        genero = request.POST.get('genero', 'no_informa')
        nacionalidad = request.POST.get('nacionalidad', 'Chilena').strip()
        direccion = request.POST.get('direccion', '').strip()
        comuna = request.POST.get('comuna', '').strip()

        # Salud y emergencia
        prevision_salud = request.POST.get('prevision_salud', 'fonasa')
        grupo_sanguineo = request.POST.get('grupo_sanguineo', '').strip()
        alergias_enfermedades = request.POST.get('alergias_enfermedades', '').strip()
        contacto_emergencia_nombre = request.POST.get('contacto_emergencia_nombre', '').strip()
        contacto_emergencia_parentesco = request.POST.get('contacto_emergencia_parentesco', '').strip()
        contacto_emergencia_telefono = request.POST.get('contacto_emergencia_telefono', '').strip()

        # Apoderado
        nombre_apoderado = request.POST.get('nombre_apoderado', '').strip()
        rut_apoderado = request.POST.get('rut_apoderado', '').strip()
        telefono_apoderado = request.POST.get('telefono_apoderado', '').strip()
        email_apoderado = request.POST.get('email_apoderado', '').strip()
        parentesco_apoderado = request.POST.get('parentesco_apoderado', '').strip()

        # PIE
        es_pie = request.POST.get('es_pie') == 'on' or request.POST.get('es_pie') == 'true'
        tipo_pie = request.POST.get('tipo_pie', '')
        diagnostico_pie = request.POST.get('diagnostico_pie', '').strip()
        observaciones_pie = request.POST.get('observaciones_pie', '').strip()

        if not nombre_completo or not seccion_id:
            messages.error(request, "El nombre completo y el curso/sección son requeridos.")
        else:
            seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
            estudiante.nombre_completo = nombre_completo
            estudiante.rut = rut
            estudiante.seccion = seccion
            estudiante.fecha_nacimiento = fecha_nacimiento
            estudiante.genero = genero
            estudiante.nacionalidad = nacionalidad
            estudiante.direccion = direccion
            estudiante.comuna = comuna
            estudiante.prevision_salud = prevision_salud
            estudiante.grupo_sanguineo = grupo_sanguineo
            estudiante.alergias_enfermedades = alergias_enfermedades
            estudiante.contacto_emergencia_nombre = contacto_emergencia_nombre
            estudiante.contacto_emergencia_parentesco = contacto_emergencia_parentesco
            estudiante.contacto_emergencia_telefono = contacto_emergencia_telefono
            estudiante.nombre_apoderado = nombre_apoderado
            estudiante.rut_apoderado = rut_apoderado
            estudiante.telefono_apoderado = telefono_apoderado
            estudiante.email_apoderado = email_apoderado
            estudiante.parentesco_apoderado = parentesco_apoderado
            estudiante.es_pie = es_pie
            estudiante.tipo_pie = tipo_pie if es_pie else None
            estudiante.diagnostico_pie = diagnostico_pie if es_pie else None
            estudiante.observaciones_pie = observaciones_pie if es_pie else None
            estudiante.activo = activo
            estudiante.save()
            messages.success(request, f"Ficha de {nombre_completo} modificada con éxito.")
            return redirect('listar_estudiantes')

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).order_by('curso__nombre', 'nombre')
    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'estudiante': estudiante,
        'secciones': secciones,
        'titulo_pagina': 'Editar Ficha de Estudiante',
        'boton_texto': 'Guardar Cambios',
    }
    return render(request, 'colegios/matricular_estudiante.html', context)


@login_required
def baja_estudiante_view(request, estudiante_id):
    colegio = obtener_colegio_usuario(request.user)
    estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para dar de baja estudiantes.")
        return redirect('listar_estudiantes')

    estudiante.activo = not estudiante.activo
    estudiante.save()
    estado = "activado" if estudiante.activo else "desactivado (baja)"
    messages.success(request, f"Estudiante {estudiante.nombre_completo} ha sido {estado} correctamente.")
    return redirect('listar_estudiantes')

@login_required
def listar_asignaturas_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    rol_str = (miembro.rol.nombre.lower() if miembro and miembro.rol else '')
    is_admin = request.user.is_superuser or request.user.colegios_administrados.filter(id=colegio.id).exists() or any(r in rol_str for r in ['administrador', 'director', 'utp', 'coordinador', 'pedagogico', 'pedagógico'])
    if not is_admin:
        messages.error(request, "Acceso restringido. Se requieren permisos de Dirección o UTP para gestionar el plan de estudios.")
        return redirect('dashboard_usuario')

    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    
    plan_estudios = []
    total_asignaturas = 0
    total_con_docente = 0

    for cur in cursos:
        asignaturas = Asignatura.objects.filter(curso=cur, activo=True).select_related('docente').order_by('nombre')
        asig_count = asignaturas.count()
        asig_con_doc = asignaturas.filter(docente__isnull=False).count()
        total_asignaturas += asig_count
        total_con_docente += asig_con_doc

        plan_estudios.append({
            'curso': cur,
            'asignaturas': asignaturas,
            'cantidad': asig_count,
            'con_docente': asig_con_doc,
            'sin_docente': asig_count - asig_con_doc,
        })

    profesores_miembros = MiembroColegio.objects.filter(colegio=colegio, activo=True).select_related('usuario', 'rol').order_by('usuario__first_name')
    docentes_disponibles = [m.usuario for m in profesores_miembros]

    total_sin_docente = total_asignaturas - total_con_docente
    cobertura = round((total_con_docente / total_asignaturas * 100)) if total_asignaturas > 0 else 100

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'plan_estudios': plan_estudios,
        'docentes_disponibles': docentes_disponibles,
        'profesores_miembros': profesores_miembros,
        'total_asignaturas': total_asignaturas,
        'total_con_docente': total_con_docente,
        'total_sin_docente': total_sin_docente,
        'cobertura': cobertura,
        'is_admin': is_admin,
    }
    return render(request, 'colegios/listar_asignaturas.html', context)


@login_required
def asignar_docente_asignatura_view(request, asignatura_id):
    colegio = obtener_colegio_usuario(request.user)
    asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': 'No tienes permisos'}, status=403)
        messages.error(request, "No tienes permisos para asignar docentes.")
        return redirect('listar_asignaturas')

    if request.method == 'POST':
        docente_id = request.POST.get('docente_id', '').strip()
        horas = request.POST.get('horas_semanales', '').strip()

        if horas and horas.isdigit():
            asignatura.horas_semanales = max(1, int(horas))

        if not docente_id or docente_id == 'none' or docente_id == '0':
            asignatura.docente = None
            asignatura.save()
            msg = f"Se removió el docente de la asignatura '{asignatura.nombre}' ({asignatura.curso.nombre})."
        else:
            docente = get_object_or_404(User, id=docente_id)
            asignatura.docente = docente
            asignatura.save()
            doc_nombre = docente.get_full_name() or docente.username
            msg = f"Se asignó a {doc_nombre} para '{asignatura.nombre}' ({asignatura.curso.nombre})."

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': msg})

        messages.success(request, msg)

    return redirect('listar_asignaturas')


@login_required
def crear_asignatura_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes permisos.")
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para crear asignaturas.")
        return redirect('listar_asignaturas')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        cursos_ids = request.POST.getlist('cursos')
        docente_id = request.POST.get('docente')
        horas_semanales = request.POST.get('horas_semanales', '4').strip()

        if not nombre or not any(cursos_ids):
            messages.error(request, "El nombre de la asignatura y al menos un curso son requeridos.")
        else:
            docente = None
            if docente_id and docente_id != 'none':
                docente = get_object_or_404(User, id=docente_id)
            
            horas = int(horas_semanales) if horas_semanales.isdigit() else 4
            created_count = 0
            for curso_id in cursos_ids:
                curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)
                asig, created = Asignatura.objects.get_or_create(
                    colegio=colegio,
                    curso=curso,
                    nombre=nombre,
                    defaults={
                        'docente': docente,
                        'horas_semanales': horas,
                        'activo': True
                    }
                )
                if created:
                    created_count += 1
                elif not asig.activo:
                    asig.activo = True
                    asig.docente = docente
                    asig.horas_semanales = horas
                    asig.save()
                    created_count += 1
            
            if created_count > 0:
                messages.success(request, f"Asignatura '{nombre}' creada con éxito en {created_count} curso(s).")
            else:
                messages.info(request, f"La asignatura '{nombre}' ya existía en los cursos seleccionados.")
            return redirect('listar_asignaturas')

    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    profesores_miembros = MiembroColegio.objects.filter(colegio=colegio, activo=True).select_related('usuario', 'rol').order_by('usuario__first_name')
    profesores = [m.usuario for m in profesores_miembros]

    context = {
        'colegio': colegio,
        'cursos': cursos,
        'profesores': profesores,
        'titulo_pagina': 'Crear Nueva Asignatura',
        'boton_texto': 'Crear Asignatura',
        'es_creacion': True,
    }
    return render(request, 'colegios/crear_asignatura.html', context)

@login_required
def editar_asignatura_view(request, asignatura_id):
    colegio = obtener_colegio_usuario(request.user)
    asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para editar asignaturas.")
        return redirect('listar_asignaturas')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        curso_id = request.POST.get('curso')
        docente_id = request.POST.get('docente')
        horas_semanales = request.POST.get('horas_semanales', '4').strip()
        activo = request.POST.get('activo') == 'true'

        if not nombre or not curso_id:
            messages.error(request, "El nombre de la asignatura y el curso son requeridos.")
        else:
            curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)
            docente = None
            if docente_id and docente_id != 'none':
                docente = get_object_or_404(User, id=docente_id)
            
            asignatura.nombre = nombre
            asignatura.curso = curso
            asignatura.docente = docente
            asignatura.horas_semanales = int(horas_semanales) if horas_semanales.isdigit() else 4
            asignatura.activo = activo
            asignatura.save()
            
            messages.success(request, f"Asignatura '{nombre}' modificada con éxito.")
            return redirect('listar_asignaturas')

    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    profesores_miembros = MiembroColegio.objects.filter(colegio=colegio, activo=True).select_related('usuario', 'rol').order_by('usuario__first_name')
    profesores = [m.usuario for m in profesores_miembros]

    context = {
        'colegio': colegio,
        'asignatura': asignatura,
        'cursos': cursos,
        'profesores': profesores,
        'titulo_pagina': 'Editar Asignatura',
        'boton_texto': 'Guardar Cambios',
        'es_creacion': False,
    }
    return render(request, 'colegios/crear_asignatura.html', context)

@login_required
def eliminar_asignatura_view(request, asignatura_id):
    colegio = obtener_colegio_usuario(request.user)
    asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para eliminar asignaturas.")
        return redirect('listar_asignaturas')

    nombre = asignatura.nombre
    asignatura.delete()
    messages.success(request, f"La asignatura '{nombre}' ha sido eliminada correctamente.")
    return redirect('listar_asignaturas')

@login_required
def precargar_asignaturas_base_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes permisos.")
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para realizar esta acción.")
        return redirect('listar_asignaturas')

    if request.method == 'POST':
        asignaturas_seleccionadas = request.POST.getlist('asignaturas_base')
        cursos_seleccionados = request.POST.getlist('cursos_destino')

        if not asignaturas_seleccionadas or not cursos_seleccionados:
            messages.error(request, "Debes seleccionar al menos una asignatura y al menos un curso.")
            return redirect('listar_asignaturas')

        created_count = 0
        for curso_id in cursos_seleccionados:
            curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)
            for nombre_asig in asignaturas_seleccionadas:
                asig, created = Asignatura.objects.get_or_create(
                    colegio=colegio,
                    curso=curso,
                    nombre=nombre_asig,
                    defaults={'docente': None, 'horas_semanales': 4, 'activo': True}
                )
                if created:
                    created_count += 1
                elif not asig.activo:
                    asig.activo = True
                    asig.save()
                    created_count += 1

        if created_count > 0:
            messages.success(request, f"Se crearon exitosamente {created_count} asignaturas en los cursos seleccionados.")

        else:
            messages.info(request, "Todas las asignaturas seleccionadas ya existían en los cursos correspondientes.")

    return redirect('listar_asignaturas')

@login_required
def listar_cursos_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes un establecimiento asociado.")
        return redirect('dashboard_usuario')
        
    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    rol_str = (miembro.rol.nombre.lower() if miembro and miembro.rol else '')
    is_admin = request.user.is_superuser or request.user.colegios_administrados.filter(id=colegio.id).exists() or any(r in rol_str for r in ['administrador', 'director', 'utp', 'coordinador', 'pedagogico', 'pedagógico'])
    if not is_admin:
        messages.error(request, "Acceso denegado. Se requieren permisos de Dirección o UTP.")
        return redirect('dashboard_usuario')

    cursos_db = CursoColegio.objects.filter(colegio=colegio).order_by('nombre')
    cursos_con_secciones = []
    for c in cursos_db:
        secciones = SeccionCurso.objects.filter(curso=c).order_by('nombre')
        cursos_con_secciones.append({
            'curso': c,
            'secciones': secciones,
            'cantidad_secciones': secciones.count()
        })

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    paginator = Paginator(cursos_con_secciones, 10)
    page_number = request.GET.get('page', 1)
    cursos_page = paginator.get_page(page_number)

    profesores_miembros = MiembroColegio.objects.filter(colegio=colegio, activo=True).select_related('usuario', 'rol').order_by('usuario__first_name', 'usuario__username')
    profesores = [m.usuario for m in profesores_miembros]

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'cursos_con_secciones': cursos_page,
        'profesores': profesores,
        'hoy': timezone.now(),
        'is_admin': is_admin
    }

    return render(request, 'colegios/listar_cursos.html', context)



@login_required
def crear_curso_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_cursos')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            messages.error(request, "El nombre del curso es requerido.")
        else:
            exists = CursoColegio.objects.filter(colegio=colegio, nombre=nombre).exists()
            if exists:
                curso = CursoColegio.objects.filter(colegio=colegio, nombre=nombre).first()
                if not curso.activo:
                    curso.activo = True
                    curso.save()
                    messages.success(request, f"El curso '{nombre}' ha sido reactivado.")
                else:
                    messages.error(request, f"El curso '{nombre}' ya existe.")
            else:
                CursoColegio.objects.create(
                    colegio=colegio,
                    nombre=nombre,
                    activo=True
                )
                messages.success(request, f"Curso '{nombre}' creado con éxito.")
    return redirect('listar_cursos')

@login_required
def editar_curso_view(request, curso_id):
    colegio = obtener_colegio_usuario(request.user)
    curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_cursos')

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            messages.error(request, "El nombre no puede estar vacío.")
        else:
            curso.nombre = nombre
            curso.save()
            messages.success(request, f"Curso actualizado a '{nombre}' correctamente.")
    return redirect('listar_cursos')

@login_required
def baja_curso_view(request, curso_id):
    colegio = obtener_colegio_usuario(request.user)
    curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_cursos')

    curso.activo = not curso.activo
    curso.save()
    
    SeccionCurso.objects.filter(curso=curso).update(activo=curso.activo)
    
    estado = "activado" if curso.activo else "desactivado (baja lógica)"
    messages.success(request, f"El curso '{curso.nombre}' ha sido {estado} correctamente.")
    return redirect('listar_cursos')

@login_required
def crear_seccion_view(request, curso_id):
    colegio = obtener_colegio_usuario(request.user)
    curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_cursos')

    if request.method == 'POST':
        secciones_nuevas = request.POST.getlist('secciones_nuevas')
        if not secciones_nuevas:
            messages.error(request, "Debes seleccionar al menos una sección.")
        else:
            created_count = 0
            ya_existen = []
            for valor in secciones_nuevas:
                valor = valor.strip()
                # El valor viene como "Sección A" - extraemos la letra (último carácter)
                letra = valor[-1].upper() if valor else ''
                if not letra or len(letra) != 1:
                    continue
                nombre_seccion = f"{curso.nombre} {letra}"
                secc, created = SeccionCurso.objects.get_or_create(
                    curso=curso,
                    letra=letra,
                    defaults={'nombre': nombre_seccion, 'activo': True}
                )
                if created:
                    created_count += 1
                elif not secc.activo:
                    secc.activo = True
                    secc.nombre = nombre_seccion
                    secc.save()
                    created_count += 1
                else:
                    ya_existen.append(letra)
            if created_count > 0:
                messages.success(request, f"Se agregaron {created_count} secciones correctamente a {curso.nombre}.")
            if ya_existen:
                messages.warning(request, f"Las secciones {', '.join(ya_existen)} ya existen en {curso.nombre}.")
    return redirect('listar_cursos')


@login_required
def baja_seccion_view(request, seccion_id):
    colegio = obtener_colegio_usuario(request.user)
    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro and miembro.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_cursos')

    from colegios.models import Estudiante
    if Estudiante.objects.filter(seccion=seccion).exists():
        messages.error(request, f"No se puede eliminar la sección '{seccion.nombre}' porque contiene alumnos matriculados.")
    else:
        nombre = seccion.nombre
        seccion.delete()
        messages.success(request, f"La sección '{nombre}' ha sido eliminada correctamente.")
    return redirect('listar_cursos')

def asegurar_roles_base_colegio(colegio, usuario=None):

    if not RolColegio.objects.filter(colegio=colegio).exists():
        roles_base = RolColegio.objects.filter(es_base=True, colegio=None)
        if not roles_base.exists():
            nombres_base = [
                ('Administrador', 'Acceso total a la administración y gestión institucional.'),
                ('Director', 'Dirección académica, reportes y políticas de evaluación.'),
                ('Profesor', 'Gestión de clases, calificaciones, asistencia y observaciones.'),
                ('Inspector', 'Control de asistencia general y convivencia escolar.'),
                ('Secretario', 'Administración de matrículas y documentación de estudiantes.'),
                ('Contabilidad', 'Gestión de cobros, facturación y finanzas.'),
                ('Apoderado', 'Visualización de notas, asistencia y comunicados de pupilos.')
            ]
            for nombre, desc in nombres_base:
                RolColegio.objects.get_or_create(
                    colegio=None,
                    nombre=nombre,
                    defaults={'descripcion': desc, 'es_base': True, 'activo': True}
                )
            roles_base = RolColegio.objects.filter(es_base=True, colegio=None)

        for rb in roles_base:
            nuevo_rol = RolColegio.objects.create(
                colegio=colegio,
                nombre=rb.nombre,
                descripcion=rb.descripcion,
                es_base=False,
                activo=True
            )
            for rp in rb.permisos.all():
                RolPermiso.objects.create(
                    rol=nuevo_rol,
                    modulo=rp.modulo,
                    puede_ver=rp.puede_ver,
                    puede_crear=rp.puede_crear,
                    puede_editar=rp.puede_editar,
                    puede_eliminar=rp.puede_eliminar,
                    puede_exportar=rp.puede_exportar,
                    puede_aprobar=rp.puede_aprobar,
                    puede_enviar_mensajes=rp.puede_enviar_mensajes,
                    puede_administrar=rp.puede_administrar
                )
            if usuario and rb.nombre == 'Administrador':
                MiembroColegio.objects.get_or_create(
                    usuario=usuario,
                    colegio=colegio,
                    defaults={'rol': nuevo_rol, 'activo': True}
                )


@login_required
def listar_personal_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes un establecimiento asociado.")
        return redirect('dashboard_usuario')
        
    asegurar_roles_base_colegio(colegio, request.user)

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "Acceso restringido. Se requieren permisos de Director o Administrador para gestionar el personal y los roles.")
        return redirect('dashboard_usuario')

    personal_qs = MiembroColegio.objects.filter(colegio=colegio).select_related('usuario', 'rol').prefetch_related('permisos_individuales').order_by('usuario__first_name')
    roles = RolColegio.objects.filter(colegio=colegio).prefetch_related('permisos__modulo').order_by('nombre')
    
    modulos_colegio = ColegioModulo.objects.filter(colegio=colegio, activo=True).select_related('modulo').order_by('modulo__nombre')
    
    # Cursos con sus asignaturas para el modal de asignación rápida
    cursos_db = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    cursos_con_asignaturas = []
    for c in cursos_db:
        asigs = Asignatura.objects.filter(curso=c, activo=True).select_related('docente').order_by('nombre')
        if asigs.exists():
            cursos_con_asignaturas.append({
                'curso': c,
                'asignaturas': asigs
            })

    # Matriz de permisos de roles serializada para JS interactivo
    roles_permisos_dict = {}
    for r in roles:
        r_perms = {}
        for rp in r.permisos.all():
            r_perms[str(rp.modulo.id)] = {
                'view': rp.puede_ver,
                'create': rp.puede_crear,
                'edit': rp.puede_editar,
                'delete': rp.puede_eliminar,
                'export': rp.puede_exportar,
            }
        roles_permisos_dict[str(r.id)] = {
            'nombre': r.nombre,
            'descripcion': r.descripcion,
            'es_base': r.es_base,
            'permisos': r_perms
        }

    # Matriz de permisos individuales por personal serializada para JS
    from colegios.models import MiembroPermiso
    personal_permisos_dict = {}
    asignaturas_por_docente_dict = {}
    
    for m in personal_qs:
        m_perms = {}
        for mp in m.permisos_individuales.all():
            m_perms[str(mp.modulo.id)] = {
                'view': mp.puede_ver,
                'create': mp.puede_crear,
                'edit': mp.puede_editar,
                'delete': mp.puede_eliminar,
                'export': mp.puede_exportar,
            }
        personal_permisos_dict[str(m.id)] = {
            'usuario': m.usuario.get_full_name() or m.usuario.username,
            'rol_id': m.rol.id if m.rol else None,
            'rol_nombre': m.rol.nombre if m.rol else 'Sin Rol',
            'permisos_individuales': m_perms
        }
        # Asignaturas asignadas a este usuario
        asignaturas_por_docente_dict[str(m.id)] = list(
            Asignatura.objects.filter(colegio=colegio, docente=m.usuario, activo=True).values_list('id', flat=True)
        )

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    paginator = Paginator(personal_qs, 10)
    page_number = request.GET.get('page', 1)
    personal_page = paginator.get_page(page_number)

    context = {
        'colegio': colegio,
        'miembro': miembro_solicitante,
        'periodo': periodo,
        'personal': personal_page,
        'roles': roles,
        'modulos_colegio': modulos_colegio,
        'cursos_con_asignaturas': cursos_con_asignaturas,
        'roles_permisos_json': json.dumps(roles_permisos_dict),
        'personal_permisos_json': json.dumps(personal_permisos_dict),
        'asignaturas_por_docente_json': json.dumps(asignaturas_por_docente_dict),
        'total_personal': personal_qs.count(),
        'total_roles': roles.count(),
        'hoy': timezone.now(),
        'is_admin': is_admin
    }
    return render(request, 'colegios/listar_personal.html', context)


@login_required
def guardar_rol_permisos_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "Acceso denegado. Se requieren permisos de Administrador.")
        return redirect('listar_personal')

    if request.method == 'POST':
        rol_id = request.POST.get('rol_id', '').strip()
        nombre_rol = request.POST.get('nombre_rol', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if not nombre_rol:
            messages.error(request, "El nombre del rol es obligatorio.")
            return redirect('listar_personal')

        from planes.models import Modulo

        if rol_id and rol_id.isdigit():
            rol = get_object_or_404(RolColegio, id=int(rol_id), colegio=colegio)
            rol.nombre = nombre_rol
            rol.descripcion = descripcion
            rol.save()
            mensaje = f"Rol '{nombre_rol}' y sus permisos actualizados correctamente."
        else:
            rol_existente = RolColegio.objects.filter(colegio=colegio, nombre__iexact=nombre_rol).first()
            if rol_existente:
                messages.warning(request, f"Ya existe un rol con el nombre '{nombre_rol}'.")
                return redirect('listar_personal')

            rol = RolColegio.objects.create(
                colegio=colegio,
                nombre=nombre_rol,
                descripcion=descripcion,
                es_base=False,
                activo=True
            )
            mensaje = f"¡Nuevo rol '{nombre_rol}' creado exitosamente con sus permisos configurados!"

        # Procesar permisos para cada módulo activo del colegio
        modulos_colegio = ColegioModulo.objects.filter(colegio=colegio, activo=True).select_related('modulo')
        for mc in modulos_colegio:
            mod_id = mc.modulo.id
            puede_ver = request.POST.get(f'perm_view_{mod_id}') == 'on'
            puede_crear = request.POST.get(f'perm_create_{mod_id}') == 'on'
            puede_editar = request.POST.get(f'perm_edit_{mod_id}') == 'on'
            puede_eliminar = request.POST.get(f'perm_delete_{mod_id}') == 'on'
            puede_exportar = request.POST.get(f'perm_export_{mod_id}') == 'on'

            rp, _ = RolPermiso.objects.get_or_create(rol=rol, modulo=mc.modulo)
            rp.puede_ver = puede_ver
            rp.puede_crear = puede_crear
            rp.puede_editar = puede_editar
            rp.puede_eliminar = puede_eliminar
            rp.puede_exportar = puede_exportar
            rp.save()

        messages.success(request, mensaje)

    return redirect('listar_personal')


@login_required
def eliminar_rol_personalizado_view(request, rol_id):
    colegio = obtener_colegio_usuario(request.user)
    rol = get_object_or_404(RolColegio, id=rol_id, colegio=colegio)

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos para eliminar roles.")
        return redirect('listar_personal')

    if rol.nombre in ['Administrador', 'Director', 'Profesor']:
        messages.error(request, f"El rol '{rol.nombre}' es un rol del sistema y no puede ser eliminado.")
        return redirect('listar_personal')

    # Verificar si hay miembros asignados a este rol
    if MiembroColegio.objects.filter(colegio=colegio, rol=rol, activo=True).exists():
        messages.error(request, f"No puedes eliminar el rol '{rol.nombre}' porque actualmente tiene funcionarios asignados. Reasigna su rol antes de continuar.")
        return redirect('listar_personal')

    nombre = rol.nombre
    rol.delete()
    messages.success(request, f"El rol '{nombre}' ha sido eliminado exitosamente.")
    return redirect('listar_personal')


@login_required
def guardar_permisos_individuales_personal_view(request, miembro_id):
    colegio = obtener_colegio_usuario(request.user)
    miembro = get_object_or_404(MiembroColegio, id=miembro_id, colegio=colegio)

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "Acceso denegado. Se requieren permisos de Administrador.")
        return redirect('listar_personal')

    if request.method == 'POST':
        from colegios.models import MiembroPermiso

        limpiar_individuales = request.POST.get('limpiar_individuales') == 'true'
        if limpiar_individuales:
            MiembroPermiso.objects.filter(miembro=miembro).delete()
            messages.success(request, f"Se restauraron los permisos individuales de {miembro.usuario.username} a los valores predeterminados de su rol.")
            return redirect('listar_personal')

        modulos_colegio = ColegioModulo.objects.filter(colegio=colegio, activo=True).select_related('modulo')
        for mc in modulos_colegio:
            mod_id = mc.modulo.id
            tiene_override = request.POST.get(f'override_{mod_id}') == 'on'
            
            if tiene_override:
                puede_ver = request.POST.get(f'ind_view_{mod_id}') == 'on'
                puede_crear = request.POST.get(f'ind_create_{mod_id}') == 'on'
                puede_editar = request.POST.get(f'ind_edit_{mod_id}') == 'on'
                puede_eliminar = request.POST.get(f'ind_delete_{mod_id}') == 'on'
                puede_exportar = request.POST.get(f'ind_export_{mod_id}') == 'on'

                mp, _ = MiembroPermiso.objects.get_or_create(miembro=miembro, modulo=mc.modulo)
                mp.puede_ver = puede_ver
                mp.puede_crear = puede_crear
                mp.puede_editar = puede_editar
                mp.puede_eliminar = puede_eliminar
                mp.puede_exportar = puede_exportar
                mp.save()
            else:
                # Si no tiene override marcado, eliminar el registro para que herede del rol base
                MiembroPermiso.objects.filter(miembro=miembro, modulo=mc.modulo).delete()

        doc_nombre = miembro.usuario.get_full_name() or miembro.usuario.username
        messages.success(request, f"Permisos individuales de {doc_nombre} guardados exitosamente.")

    return redirect('listar_personal')


@login_required
def crear_rol_personalizado_view(request):
    return guardar_rol_permisos_view(request)




@login_required
def asignar_asignaturas_docente_view(request, miembro_id):
    colegio = obtener_colegio_usuario(request.user)
    miembro = get_object_or_404(MiembroColegio, id=miembro_id, colegio=colegio)

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "Acceso denegado. Se requieren permisos de administrador.")
        return redirect('listar_personal')

    if request.method == 'POST':
        asignaturas_ids = request.POST.getlist('asignaturas')
        # Limpiar asignaciones previas de este docente en este colegio
        Asignatura.objects.filter(colegio=colegio, docente=miembro.usuario).update(docente=None)
        # Asignar seleccionadas
        if asignaturas_ids:
            Asignatura.objects.filter(colegio=colegio, id__in=asignaturas_ids).update(docente=miembro.usuario)

        docente_nombre = miembro.usuario.get_full_name() or miembro.usuario.username
        messages.success(request, f"Se actualizaron exitosamente las asignaturas asignadas a {docente_nombre}.")

    return redirect('listar_personal')



@login_required
def editar_personal_view(request, miembro_id):
    colegio = obtener_colegio_usuario(request.user)
    miembro = get_object_or_404(MiembroColegio, id=miembro_id, colegio=colegio)

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_personal')

    if request.method == 'POST':
        rol_id = request.POST.get('rol')
        if rol_id:
            nuevo_rol = get_object_or_404(RolColegio, id=rol_id, colegio=colegio)
            miembro.rol = nuevo_rol
            miembro.save()
            messages.success(request, f"El rol de {miembro.usuario.username} ha sido cambiado a {nuevo_rol.nombre}.")
    return redirect('listar_personal')

@login_required
def baja_personal_view(request, miembro_id):
    colegio = obtener_colegio_usuario(request.user)
    miembro = get_object_or_404(MiembroColegio, id=miembro_id, colegio=colegio)

    if miembro.usuario == request.user:
        messages.error(request, "No puedes cambiar tu propio estado de acceso.")
        return redirect('listar_personal')

    miembro_solicitante = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = request.user.colegios_administrados.filter(id=colegio.id).exists() or (miembro_solicitante and miembro_solicitante.rol.nombre in ['Administrador', 'Director'])
    if not is_admin:
        messages.error(request, "No tienes permisos.")
        return redirect('listar_personal')

    miembro.activo = not miembro.activo
    miembro.save()
    
    estado = "reactivado" if miembro.activo else "desactivado (baja de acceso)"
    messages.success(request, f"El acceso para {miembro.usuario.username} ha sido {estado}.")
    return redirect('listar_personal')


# ── CENTRO DE REPORTES Y ANALÍTICA ESCOLAR ───────────────────────────────────

@login_required
def centro_reportes_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    estudiantes_qs = Estudiante.objects.filter(colegio=colegio, activo=True)
    
    total_estudiantes = estudiantes_qs.count()
    total_secciones = secciones.count()
    total_pie = estudiantes_qs.filter(es_pie=True).count()

    # Métricas Globales de Calificaciones
    from calificaciones.models import Nota
    notas_qs = Nota.objects.filter(evaluacion__colegio=colegio)
    from django.db.models import Avg, Count, Q
    promedio_global_val = notas_qs.aggregate(Avg('valor'))['valor__avg']
    promedio_global = round(float(promedio_global_val), 1) if promedio_global_val else 0.0
    alumnos_riesgo_notas = notas_qs.filter(valor__lt=4.0).values('estudiante').distinct().count()

    # Métricas Globales de Asistencia
    from asistencia.models import DetalleAsistencia, RegistroAsistencia
    detalles_asist = DetalleAsistencia.objects.filter(registro__seccion__curso__colegio=colegio)
    total_asist = detalles_asist.count()
    if total_asist > 0:
        presentes = detalles_asist.filter(estado__in=['presente', 'tarde', 'justificado']).count()
        tasa_asistencia_global = round((presentes / total_asist) * 100, 1)
    else:
        tasa_asistencia_global = 0.0

    from asistencia.utils import calcular_alumnos_en_riesgo
    alumnos_riesgo_asistencia = len(calcular_alumnos_en_riesgo(colegio))

    meses_choices = [
        (3, 'Marzo'), (4, 'Abril'), (5, 'Mayo'), (6, 'Junio'),
        (7, 'Julio'), (8, 'Agosto'), (9, 'Septiembre'), (10, 'Octubre'),
        (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'reportes',
        'secciones': secciones,
        'secciones_count': total_secciones,
        'estudiantes_count': total_estudiantes,
        'pie_count': total_pie,
        'promedio_global': promedio_global,
        'tasa_asistencia_global': tasa_asistencia_global,
        'alumnos_riesgo_asistencia': alumnos_riesgo_asistencia,
        'alumnos_riesgo_notas': alumnos_riesgo_notas,
        'meses_choices': meses_choices,
        'mes_actual': timezone.now().month,
        'anio_actual': periodo.anio_academico if periodo else timezone.now().year,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/reportes_hub.html', context)


@login_required
def generar_certificado_alumno_regular_view(request, estudiante_id=None):
    """Genera el certificado oficial de alumno regular con membrete institucional imprimible en PDF."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    if not estudiante_id:
        estudiante_id = request.GET.get('estudiante_id')

    if not estudiante_id:
        estudiantes = Estudiante.objects.filter(colegio=colegio, activo=True).select_related('seccion', 'seccion__curso').order_by('nombre_completo')
        return render(request, 'colegios/reportes/seleccionar_estudiante_certificado.html', {
            'colegio': colegio,
            'estudiantes': estudiantes,
            'active_page': 'reportes'
        })

    estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    anio_escolar = periodo.anio_academico if periodo else timezone.now().year

    MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    hoy = timezone.now()
    fecha_emision_texto = f"{hoy.day} de {MESES[hoy.month - 1]} de {hoy.year}"

    import hashlib
    hash_folio = hashlib.sha256(f"{estudiante.id}-{colegio.id}-{anio_escolar}-{hoy.strftime('%Y%m%d')}".encode('utf-8')).hexdigest()[:10].upper()
    codigo_verificacion = f"EDK-{colegio.id:03d}-{estudiante.id:04d}-{hash_folio}"

    nombre_director = "Director(a) de Establecimiento"
    if colegio.administrador and hasattr(colegio.administrador, 'perfil') and colegio.administrador.perfil.nombre_completo:
        nombre_director = colegio.administrador.perfil.nombre_completo
    elif colegio.administrador:
        nombre_director = colegio.administrador.get_full_name() or colegio.administrador.username

    context = {
        'colegio': colegio,
        'estudiante': estudiante,
        'periodo': periodo,
        'anio_escolar': anio_escolar,
        'fecha_emision_texto': fecha_emision_texto,
        'codigo_verificacion': codigo_verificacion,
        'nombre_director': nombre_director,
        'hoy': hoy,
    }
    return render(request, 'colegios/reportes/certificado_alumno_regular.html', context)


@login_required
def reporte_consolidado_notas_seccion_view(request):
    """Muestra la matriz completa de notas de una sección con ranking y estadísticas."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    seccion_id = request.GET.get('seccion')
    
    seccion = None
    if seccion_id and seccion_id.isdigit():
        seccion = secciones.filter(id=int(seccion_id)).first()
    if not seccion:
        seccion = secciones.first()

    if not seccion:
        messages.info(request, "No hay cursos configurados en el colegio.")
        return redirect('centro_reportes')

    asignaturas = Asignatura.objects.filter(curso=seccion.curso, activo=True).order_by('nombre')
    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    from calificaciones.models import Nota
    from django.db.models import Avg

    matriz = []
    promedios_asignatura = {}
    for a in asignaturas:
        promedios_asignatura[a.id] = []

    todos_promedios_alumnos = []

    for est in estudiantes:
        fila_notas = {}
        notas_alumno = []
        rojas_cnt = 0

        for asig in asignaturas:
            n_qs = Nota.objects.filter(evaluacion__seccion=seccion, evaluacion__asignatura=asig, estudiante=est)
            avg_asig = n_qs.aggregate(Avg('valor'))['valor__avg']
            if avg_asig is not None:
                val = round(float(avg_asig), 1)
                fila_notas[asig.id] = val
                notas_alumno.append(val)
                promedios_asignatura[asig.id].append(val)
                if val < 4.0:
                    rojas_cnt += 1
            else:
                fila_notas[asig.id] = None

        if notas_alumno:
            prom_est = round(sum(notas_alumno) / len(notas_alumno), 1)
            todos_promedios_alumnos.append(prom_est)
        else:
            prom_est = None

        matriz.append({
            'estudiante': est,
            'notas': fila_notas,
            'promedio': prom_est,
            'rojas_count': rojas_cnt,
            'estado': 'Aprobando' if (prom_est and prom_est >= 4.0 and rojas_cnt <= 2) else ('En Riesgo' if prom_est else 'Sin Notas')
        })

    # Calcular ranking de notas
    matriz_ordenada = sorted([m for m in matriz if m['promedio'] is not None], key=lambda x: x['promedio'], reverse=True)
    ranking_map = {m['estudiante'].id: idx + 1 for idx, m in enumerate(matriz_ordenada)}
    for m in matriz:
        m['ranking'] = ranking_map.get(m['estudiante'].id, '-')

    # Resumen por asignatura
    resumen_asignaturas = []
    for asig in asignaturas:
        vals = promedios_asignatura[asig.id]
        if vals:
            avg_col = round(sum(vals) / len(vals), 1)
            aprob = sum(1 for v in vals if v >= 4.0)
            pct_aprob = round((aprob / len(vals)) * 100, 1)
        else:
            avg_col = '-'
            pct_aprob = '-'
        resumen_asignaturas.append({
            'asignatura': asig,
            'promedio': avg_col,
            'pct_aprobacion': pct_aprob
        })

    promedio_seccion_global = round(sum(todos_promedios_alumnos) / len(todos_promedios_alumnos), 1) if todos_promedios_alumnos else '-'
    aprobados_total = sum(1 for m in matriz if m['promedio'] and m['promedio'] >= 4.0)
    tasa_aprobacion_seccion = round((aprobados_total / len(matriz) * 100), 1) if matriz else 0

    context = {
        'colegio': colegio,
        'seccion': seccion,
        'secciones': secciones,
        'asignaturas': asignaturas,
        'matriz': matriz,
        'resumen_asignaturas': resumen_asignaturas,
        'promedio_seccion_global': promedio_seccion_global,
        'tasa_aprobacion_seccion': tasa_aprobacion_seccion,
        'total_estudiantes': len(matriz),
        'active_page': 'reportes',
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/reportes/reporte_matriz_notas.html', context)


@login_required
def exportar_consolidado_notas_excel_view(request, seccion_id):
    """Exporta la matriz de calificaciones de una sección a una planilla Excel con estilos profesionales."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Avg
    from calificaciones.models import Nota

    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    asignaturas = Asignatura.objects.filter(curso=seccion.curso, activo=True).order_by('nombre')
    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz_Calificaciones"
    ws.views.sheetView[0].showGridLines = True

    FILL_HEADER = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FILL_ROJO = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    FONT_ROJO = Font(name="Segoe UI", size=9, bold=True, color="DC2626")
    FONT_NORMAL = Font(name="Segoe UI", size=9, color="0F172A")
    BORDER_THIN = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    ws.merge_cells('A1:G1')
    ws['A1'] = f"{colegio.nombre} - INFORME MATRIZ DE CALIFICACIONES: {seccion.nombre}"
    ws['A1'].font = Font(name="Segoe UI", size=13, bold=True, color="7C5CFC")
    ws['A1'].alignment = Alignment(vertical="center")

    headers = ["#", "RUT", "Nombre Completo"] + [a.nombre[:15] for a in asignaturas] + ["Promedio General", "Notas Rojas", "Estado"]
    ws.row_dimensions[3].height = 28

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx == 3 else (14 if col_idx > 3 else 12)

    current_row = 4
    for idx, est in enumerate(estudiantes, start=1):
        ws.row_dimensions[current_row].height = 20
        c1 = ws.cell(row=current_row, column=1, value=idx)
        c2 = ws.cell(row=current_row, column=2, value=est.rut or '-')
        c3 = ws.cell(row=current_row, column=3, value=est.nombre_completo)
        for c in [c1, c2, c3]:
            c.font = FONT_NORMAL
            c.border = BORDER_THIN

        notas_alumno = []
        rojas_cnt = 0
        for a_idx, asig in enumerate(asignaturas, start=4):
            n_qs = Nota.objects.filter(evaluacion__seccion=seccion, evaluacion__asignatura=asig, estudiante=est)
            avg_val = n_qs.aggregate(Avg('valor'))['valor__avg']
            cell = ws.cell(row=current_row, column=a_idx)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if avg_val is not None:
                val = round(float(avg_val), 1)
                cell.value = val
                notas_alumno.append(val)
                if val < 4.0:
                    cell.fill = FILL_ROJO
                    cell.font = FONT_ROJO
                    rojas_cnt += 1
                else:
                    cell.font = FONT_NORMAL
            else:
                cell.value = '-'
                cell.font = Font(name="Segoe UI", size=9, color="94A3B8")

        prom_col_idx = 4 + len(asignaturas)
        prom_cell = ws.cell(row=current_row, column=prom_col_idx)
        prom_cell.border = BORDER_THIN
        prom_cell.alignment = Alignment(horizontal="center", vertical="center")
        if notas_alumno:
            prom_val = round(sum(notas_alumno) / len(notas_alumno), 1)
            prom_cell.value = prom_val
            prom_cell.font = FONT_HEADER if prom_val >= 4.0 else FONT_ROJO
        else:
            prom_cell.value = '-'
            prom_cell.font = FONT_NORMAL

        rojas_cell = ws.cell(row=current_row, column=prom_col_idx + 1, value=rojas_cnt)
        rojas_cell.border = BORDER_THIN
        rojas_cell.alignment = Alignment(horizontal="center", vertical="center")
        rojas_cell.font = FONT_ROJO if rojas_cnt > 0 else FONT_NORMAL

        estado_val = "Aprobando" if (notas_alumno and (sum(notas_alumno)/len(notas_alumno)) >= 4.0 and rojas_cnt <= 2) else ("En Riesgo" if notas_alumno else "Sin Notas")
        estado_cell = ws.cell(row=current_row, column=prom_col_idx + 2, value=estado_val)
        estado_cell.border = BORDER_THIN
        estado_cell.alignment = Alignment(horizontal="center", vertical="center")
        estado_cell.font = FONT_NORMAL

        current_row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    slug_sec = seccion.nombre.lower().replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="Matriz_Calificaciones_{slug_sec}.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_mensual_asistencia_seccion_view(request):
    """Muestra la planilla mensual de asistencia consolidada por sección con alertas <85%."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    seccion_id = request.GET.get('seccion')
    mes_str = request.GET.get('mes')

    seccion = None
    if seccion_id and seccion_id.isdigit():
        seccion = secciones.filter(id=int(seccion_id)).first()
    if not seccion:
        seccion = secciones.first()

    if not seccion:
        messages.info(request, "No hay secciones configuradas.")
        return redirect('centro_reportes')

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    anio = periodo.anio_academico if periodo else timezone.now().year

    try:
        mes = int(mes_str) if mes_str else timezone.now().month
    except ValueError:
        mes = timezone.now().month

    from asistencia.models import RegistroAsistencia, DetalleAsistencia

    registros_mes = RegistroAsistencia.objects.filter(
        seccion=seccion,
        fecha__year=anio,
        fecha__month=mes
    ).order_by('fecha')

    dias_registrados = list(registros_mes.values_list('fecha', flat=True).distinct())
    total_dias_habiles = len(dias_registrados)

    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')
    tabla_alumnos = []
    alertas_desercion_cnt = 0
    todos_pct = []

    for est in estudiantes:
        detalles = DetalleAsistencia.objects.filter(registro__in=registros_mes, estudiante=est)
        presentes = detalles.filter(estado__in=['presente', 'tarde', 'justificado']).count()
        ausentes = detalles.filter(estado='ausente').count()
        tardes = detalles.filter(estado='tarde').count()

        if total_dias_habiles > 0:
            pct_asist = round((presentes / total_dias_habiles) * 100, 1)
            todos_pct.append(pct_asist)
        else:
            pct_asist = 100.0

        es_critico = (pct_asist < 85.0 and total_dias_habiles > 0)
        if es_critico:
            alertas_desercion_cnt += 1

        tabla_alumnos.append({
            'estudiante': est,
            'presentes': presentes,
            'ausentes': ausentes,
            'tardes': tardes,
            'porcentaje': pct_asist,
            'es_critico': es_critico,
        })

    tasa_seccion_mes = round(sum(todos_pct) / len(todos_pct), 1) if todos_pct else 100.0

    meses_choices = [
        (3, 'Marzo'), (4, 'Abril'), (5, 'Mayo'), (6, 'Junio'),
        (7, 'Julio'), (8, 'Agosto'), (9, 'Septiembre'), (10, 'Octubre'),
        (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'colegio': colegio,
        'seccion': seccion,
        'secciones': secciones,
        'mes': mes,
        'anio': anio,
        'meses_choices': meses_choices,
        'total_dias_habiles': total_dias_habiles,
        'tabla_alumnos': tabla_alumnos,
        'tasa_seccion_mes': tasa_seccion_mes,
        'alertas_desercion_cnt': alertas_desercion_cnt,
        'active_page': 'reportes',
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/reportes/reporte_asistencia_mensual.html', context)


@login_required
def exportar_mensual_asistencia_excel_view(request, seccion_id):
    """Exporta el reporte consolidado de asistencia mensual de una sección a Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    mes_str = request.GET.get('mes')
    try:
        mes = int(mes_str) if mes_str else timezone.now().month
    except ValueError:
        mes = timezone.now().month

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    anio = periodo.anio_academico if periodo else timezone.now().year

    from asistencia.models import RegistroAsistencia, DetalleAsistencia

    registros_mes = RegistroAsistencia.objects.filter(
        seccion=seccion,
        fecha__year=anio,
        fecha__month=mes
    ).order_by('fecha')

    dias_habiles = len(set(registros_mes.values_list('fecha', flat=True)))
    estudiantes = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencia_Mes_{mes}"
    ws.views.sheetView[0].showGridLines = True

    FILL_HEADER = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    FILL_ROJO = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    FONT_ROJO = Font(name="Segoe UI", size=9, bold=True, color="DC2626")
    FONT_NORMAL = Font(name="Segoe UI", size=9, color="0F172A")
    BORDER_THIN = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'), top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    ws.merge_cells('A1:F1')
    ws['A1'] = f"{colegio.nombre} - REPORTE MENSUAL DE ASISTENCIA (SIGE): {seccion.nombre} - Mes {mes}/{anio}"
    ws['A1'].font = Font(name="Segoe UI", size=12, bold=True, color="10B981")

    headers = ["#", "RUT", "Nombre Completo", "Días Hábiles", "Días Presente", "Inasistencias", "% Asistencia Mensual", "Alerta (<85%)"]
    ws.row_dimensions[3].height = 26

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = 26 if col_idx == 3 else 16

    current_row = 4
    for idx, est in enumerate(estudiantes, start=1):
        detalles = DetalleAsistencia.objects.filter(registro__in=registros_mes, estudiante=est)
        presentes = detalles.filter(estado__in=['presente', 'tarde', 'justificado']).count()
        ausentes = detalles.filter(estado='ausente').count()
        pct = round((presentes / dias_habiles * 100), 1) if dias_habiles > 0 else 100.0

        ws.row_dimensions[current_row].height = 20
        c1 = ws.cell(row=current_row, column=1, value=idx)
        c2 = ws.cell(row=current_row, column=2, value=est.rut or '-')
        c3 = ws.cell(row=current_row, column=3, value=est.nombre_completo)
        c4 = ws.cell(row=current_row, column=4, value=dias_habiles)
        c5 = ws.cell(row=current_row, column=5, value=presentes)
        c6 = ws.cell(row=current_row, column=6, value=ausentes)
        c7 = ws.cell(row=current_row, column=7, value=f"{pct}%")
        c8 = ws.cell(row=current_row, column=8, value="CRÍTICO (<85%)" if (pct < 85.0 and dias_habiles > 0) else "NORMAL")

        for c in [c1, c2, c3, c4, c5, c6, c7, c8]:
            c.border = BORDER_THIN
            c.font = FONT_NORMAL
            if c != c3:
                c.alignment = Alignment(horizontal="center", vertical="center")

        if pct < 85.0 and dias_habiles > 0:
            c7.fill = FILL_ROJO
            c7.font = FONT_ROJO
            c8.fill = FILL_ROJO
            c8.font = FONT_ROJO

        current_row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    slug_sec = seccion.nombre.lower().replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="Asistencia_Mensual_{slug_sec}_Mes_{mes}.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_convivencia_consolidado_view(request):
    """Consolidado de anotaciones, méritos y deméritos por sección y alertas de convivencia."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    seccion_id = request.GET.get('seccion')

    from colegios.models import AnotacionEstudiante
    anotaciones_qs = AnotacionEstudiante.objects.filter(colegio=colegio).select_related('estudiante', 'estudiante__seccion', 'docente', 'asignatura')

    if seccion_id and seccion_id.isdigit():
        anotaciones_qs = anotaciones_qs.filter(estudiante__seccion_id=int(seccion_id))

    total_positivas = anotaciones_qs.filter(tipo='positiva').count()
    total_negativas = anotaciones_qs.filter(tipo='negativa').count()
    total_citaciones = anotaciones_qs.filter(tipo='citacion').count()
    total_neutras = anotaciones_qs.filter(tipo='neutra').count()

    context = {
        'colegio': colegio,
        'secciones': secciones,
        'seccion_seleccionada': int(seccion_id) if (seccion_id and seccion_id.isdigit()) else None,
        'anotaciones': anotaciones_qs.order_by('-fecha', '-id')[:50],
        'total_positivas': total_positivas,
        'total_negativas': total_negativas,
        'total_citaciones': total_citaciones,
        'total_neutras': total_neutras,
        'active_page': 'reportes',
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/reportes/reporte_convivencia.html', context)


@login_required
def reporte_resumen_ejecutivo_institucional_view(request):
    """Ficha ejecutiva 360° con métricas institucionales para directivos y MINEDUC."""
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    estudiantes_qs = Estudiante.objects.filter(colegio=colegio, activo=True)
    secciones_qs = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True)

    # Métricas
    total_matricula = estudiantes_qs.count()
    hombres_cnt = estudiantes_qs.filter(genero='masculino').count()
    mujeres_cnt = estudiantes_qs.filter(genero='femenino').count()
    pie_cnt = estudiantes_qs.filter(es_pie=True).count()

    basica_cnt = estudiantes_qs.filter(seccion__curso__nivel='basica').count()
    media_cnt = estudiantes_qs.filter(seccion__curso__nivel='media').count()

    from calificaciones.models import Nota
    from django.db.models import Avg
    promedio_global = Nota.objects.filter(evaluacion__colegio=colegio).aggregate(Avg('valor'))['valor__avg']
    promedio_global_fmt = round(float(promedio_global), 1) if promedio_global else 0.0

    from colegios.models import TallerExtracurricular, CuentaFinanciera
    from solicitudes.models import MiembroColegio
    talleres_cnt = TallerExtracurricular.objects.filter(colegio=colegio, activo=True).count()
    personal_cnt = MiembroColegio.objects.filter(colegio=colegio, activo=True).count()

    context = {
        'colegio': colegio,
        'periodo': periodo,
        'total_matricula': total_matricula,
        'hombres_cnt': hombres_cnt,
        'mujeres_cnt': mujeres_cnt,
        'pie_cnt': pie_cnt,
        'basica_cnt': basica_cnt,
        'media_cnt': media_cnt,
        'promedio_global': promedio_global_fmt,
        'talleres_cnt': talleres_cnt,
        'personal_cnt': personal_cnt,
        'secciones_cnt': secciones_qs.count(),
        'active_page': 'reportes',
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/reportes/reporte_resumen_ejecutivo.html', context)


@login_required
def configuracion_politicas_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )
    if not is_admin:
        messages.error(request, "Acceso denegado. Se requieren permisos de Administrador.")
        return redirect('dashboard_usuario')

    periodo, _ = ConfiguracionAcademica.objects.get_or_create(
        colegio=colegio,
        defaults={
            'anio_academico': timezone.now().year,
            'fecha_inicio': timezone.now().date(),
            'fecha_termino': timezone.now().date(),
        }
    )

    if request.method == 'POST':
        periodo.modalidad_asistencia = request.POST.get('modalidad_asistencia', 'asignatura')
        try:
            periodo.porcentaje_asistencia_minima = int(request.POST.get('porcentaje_asistencia_minima', 85))
        except ValueError:
            periodo.porcentaje_asistencia_minima = 85

        periodo.tipo_calificacion = request.POST.get('tipo_calificacion', 'numerica')
        try:
            periodo.nota_minima_aprobacion = float(request.POST.get('nota_minima_aprobacion', 4.0))
        except ValueError:
            periodo.nota_minima_aprobacion = 4.0

        try:
            periodo.porcentaje_exigencia = int(request.POST.get('porcentaje_exigencia', 60))
        except ValueError:
            periodo.porcentaje_exigencia = 60

        periodo.regla_redondeo = request.POST.get('regla_redondeo', 'un_decimal')
        periodo.tipo_calculo_promedio = request.POST.get('tipo_calculo_promedio', 'ponderado')
        periodo.visibilidad_notas_apoderados = request.POST.get('visibilidad_notas_apoderados', 'inmediata')
        periodo.notificar_ausencias = (request.POST.get('notificar_ausencias') == 'on')
        periodo.notificar_notas_rojas = (request.POST.get('notificar_notas_rojas') == 'on')

        periodo.save()
        messages.success(request, "¡Configuración de Políticas Académicas actualizada exitosamente!")
        return redirect('configuracion_politicas')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/configuracion_politicas.html', context)


@login_required
def hoja_vida_estudiante_view(request, estudiante_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)
    seccion = estudiante.seccion
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    if request.method == 'POST':
        tipo = request.POST.get('tipo', 'neutra')
        gravedad = request.POST.get('gravedad', 'leve')
        titulo = request.POST.get('titulo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        asignatura_id = request.POST.get('asignatura_id')

        if not titulo or not descripcion:
            messages.error(request, "El título y la descripción son obligatorios.")
            return redirect('hoja_vida_estudiante', estudiante_id=estudiante.id)

        asig_obj = None
        if asignatura_id:
            asig_obj = Asignatura.objects.filter(id=asignatura_id, colegio=colegio).first()

        from colegios.models import AnotacionEstudiante
        AnotacionEstudiante.objects.create(
            colegio=colegio,
            estudiante=estudiante,
            asignatura=asig_obj,
            docente=request.user,
            tipo=tipo,
            gravedad=gravedad,
            titulo=titulo,
            descripcion=descripcion,
            fecha=timezone.now().date()
        )
        messages.success(request, f"¡Anotación agregada exitosamente a la hoja de vida de {estudiante.nombre_completo}!")
        return redirect('hoja_vida_estudiante', estudiante_id=estudiante.id)

    from colegios.models import AnotacionEstudiante
    anotaciones = AnotacionEstudiante.objects.filter(estudiante=estudiante).select_related('docente', 'asignatura')

    positivas_cnt = anotaciones.filter(tipo='positiva').count()
    negativas_cnt = anotaciones.filter(tipo='negativa').count()
    citaciones_cnt = anotaciones.filter(tipo='citacion').count()
    neutras_cnt = anotaciones.filter(tipo='neutra').count()

    if negativas_cnt == 0:
        semaforo_estado = 'excelente'
        semaforo_texto = 'Excelente Conducta'
        semaforo_color = 'success'
    elif negativas_cnt <= 2:
        semaforo_estado = 'atencion'
        semaforo_texto = 'Atención Requerida'
        semaforo_color = 'warning'
    else:
        semaforo_estado = 'alerta'
        semaforo_texto = 'Alerta Convivencia Escolar'
        semaforo_color = 'danger'

    asignaturas_curso = []
    estudiantes_seccion = []
    if seccion:
        asignaturas_curso = Asignatura.objects.filter(curso=seccion.curso, activo=True).order_by('nombre')
        estudiantes_seccion = Estudiante.objects.filter(seccion=seccion, activo=True).order_by('nombre_completo')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'convivencia',
        'estudiante': estudiante,
        'seccion': seccion,
        'estudiantes_seccion': estudiantes_seccion,
        'anotaciones': anotaciones,
        'positivas_cnt': positivas_cnt,
        'negativas_cnt': negativas_cnt,
        'citaciones_cnt': citaciones_cnt,
        'neutras_cnt': neutras_cnt,
        'semaforo_estado': semaforo_estado,
        'semaforo_texto': semaforo_texto,
        'semaforo_color': semaforo_color,
        'asignaturas_curso': asignaturas_curso,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/hoja_vida_estudiante.html', context)



@login_required
def eliminar_anotacion_view(request, anotacion_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    from colegios.models import AnotacionEstudiante
    anotacion = get_object_or_404(AnotacionEstudiante, id=anotacion_id, colegio=colegio)
    estudiante_id = anotacion.estudiante.id

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    if not is_admin and anotacion.docente != request.user:
        messages.error(request, "No tienes permiso para eliminar esta anotación.")
        return redirect('hoja_vida_estudiante', estudiante_id=estudiante_id)

    anotacion.delete()
    messages.success(request, "Anotación eliminada correctamente.")
    return redirect('hoja_vida_estudiante', estudiante_id=estudiante_id)


@login_required
def convivencia_hub_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    secciones = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__nombre', 'nombre')

    seccion_id = request.GET.get('seccion')
    busqueda = request.GET.get('q', '').strip()

    estudiantes_qs = Estudiante.objects.filter(colegio=colegio, activo=True).select_related('seccion', 'seccion__curso').order_by('seccion__curso__nombre', 'nombre_completo')

    if seccion_id and seccion_id.isdigit():
        estudiantes_qs = estudiantes_qs.filter(seccion_id=int(seccion_id))

    if busqueda:
        from django.db.models import Q
        estudiantes_qs = estudiantes_qs.filter(
            Q(nombre_completo__icontains=busqueda) | Q(rut__icontains=busqueda)
        )

    from colegios.models import AnotacionEstudiante
    estudiantes_data = []
    total_alertas_count = 0

    for est in estudiantes_qs:
        anotaciones = AnotacionEstudiante.objects.filter(estudiante=est)
        pos_cnt = anotaciones.filter(tipo='positiva').count()
        neg_cnt = anotaciones.filter(tipo='negativa').count()
        cit_cnt = anotaciones.filter(tipo='citacion').count()

        if neg_cnt == 0:
            semaforo_color = 'success'
            semaforo_text = 'Excelente'
        elif neg_cnt <= 2:
            semaforo_color = 'warning'
            semaforo_text = 'Atención'
        else:
            semaforo_color = 'danger'
            semaforo_text = 'Alerta'
            total_alertas_count += 1

        estudiantes_data.append({
            'estudiante': est,
            'pos_cnt': pos_cnt,
            'neg_cnt': neg_cnt,
            'cit_cnt': cit_cnt,
            'semaforo_color': semaforo_color,
            'semaforo_text': semaforo_text,
        })

    paginator = Paginator(estudiantes_data, 10)
    page_number = request.GET.get('page', 1)
    estudiantes_data_page = paginator.get_page(page_number)

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'convivencia',
        'secciones': secciones,
        'seccion_seleccionada': int(seccion_id) if (seccion_id and seccion_id.isdigit()) else None,
        'busqueda': busqueda,
        'estudiantes_data': estudiantes_data_page,
        'total_monitoreados': len(estudiantes_data),
        'total_alertas_count': total_alertas_count,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/convivencia_hub.html', context)




@login_required
@login_required
def calendario_escolar_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )
    is_profesor = (miembro and miembro.rol and miembro.rol.nombre == 'Profesor')

    from colegios.models import EventoAgenda
    from datetime import datetime, time, timedelta

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear_bloque_clase':
            # 1. Crear Bloque de Clase Semanal Recurrente
            asignatura_id = request.POST.get('asignatura_id')
            curso_id = request.POST.get('curso_id')
            docente_id = request.POST.get('docente_id')
            dia_semana = request.POST.get('dia_semana')
            hora_inicio_str = request.POST.get('hora_inicio', '08:00')
            hora_fin_str = request.POST.get('hora_fin', '09:30')
            lugar = request.POST.get('lugar', '').strip()

            asig_obj = Asignatura.objects.filter(id=asignatura_id, colegio=colegio).first() if (asignatura_id and asignatura_id.isdigit()) else None
            curso_obj = CursoColegio.objects.filter(id=curso_id, colegio=colegio).first() if (curso_id and curso_id.isdigit()) else None
            docente_obj = User.objects.filter(id=docente_id).first() if (docente_id and docente_id.isdigit()) else request.user

            if asig_obj and dia_semana is not None and dia_semana.isdigit():
                dia_num = int(dia_semana)
                
                # Base date: primer lunes de referencia
                base_monday = timezone.now().date() - timedelta(days=timezone.now().date().weekday())
                target_date = base_monday + timedelta(days=dia_num)

                try:
                    h_ini = datetime.strptime(hora_inicio_str, '%H:%M').time()
                    h_fin = datetime.strptime(hora_fin_str, '%H:%M').time()
                    dt_ini = timezone.make_aware(datetime.combine(target_date, h_ini))
                    dt_fin = timezone.make_aware(datetime.combine(target_date, h_fin))
                except ValueError:
                    dt_ini = timezone.now()
                    dt_fin = timezone.now() + timedelta(hours=1, minutes=30)

                nombre_curso = curso_obj.nombre if curso_obj else (asig_obj.curso.nombre if asig_obj.curso else '')
                titulo_clase = f"{asig_obj.nombre}" + (f" ({nombre_curso})" if nombre_curso else "")

                EventoAgenda.objects.create(
                    colegio=colegio,
                    creado_por=request.user,
                    asignado_a=docente_obj,
                    es_recurrente=True,
                    dia_semana=dia_num,
                    titulo=titulo_clase,
                    tipo='clase',
                    fecha_inicio=dt_ini,
                    fecha_fin=dt_fin,
                    lugar=lugar if lugar else (f"Sala {nombre_curso}" if nombre_curso else "Sala de Clases"),
                    curso=curso_obj or asig_obj.curso,
                    asignatura=asig_obj,
                    descripcion=f"Horario fijo semanal de {asig_obj.nombre}."
                )
                messages.success(request, f"¡Bloque de clase para '{asig_obj.nombre}' ({docente_obj.get_full_name() or docente_obj.username}) programado exitosamente!")
                return redirect(f"{reverse('calendario_escolar')}?tab=horario&docente_id={docente_obj.id}")

        else:
            # 2. Crear Reunión o Evento Extraordinario
            titulo = request.POST.get('titulo', '').strip()
            tipo = request.POST.get('tipo', 'reunion')
            fecha_str = request.POST.get('fecha_inicio')
            hora_str = request.POST.get('hora_inicio', '08:00')
            hora_fin_str = request.POST.get('hora_fin')
            lugar = request.POST.get('lugar', '').strip()
            curso_id = request.POST.get('curso_id')
            asignado_a_id = request.POST.get('asignado_a_id')
            es_para_todos = (request.POST.get('es_para_todos') in ['on', '1', 'true'])
            descripcion = request.POST.get('descripcion', '').strip()

            if titulo and fecha_str:
                try:
                    dt_str = f"{fecha_str} {hora_str}"
                    fecha_inicio = timezone.make_aware(datetime.strptime(dt_str, '%Y-%m-%d %H:%M'))
                    if hora_fin_str:
                        dt_fin_str = f"{fecha_str} {hora_fin_str}"
                        fecha_fin = timezone.make_aware(datetime.strptime(dt_fin_str, '%Y-%m-%d %H:%M'))
                    else:
                        fecha_fin = fecha_inicio + timedelta(hours=1)
                except ValueError:
                    fecha_inicio = timezone.now()
                    fecha_fin = timezone.now() + timedelta(hours=1)

                curso_obj = CursoColegio.objects.filter(id=curso_id, colegio=colegio).first() if (curso_id and curso_id.isdigit()) else None
                asignado_obj = User.objects.filter(id=asignado_a_id).first() if (asignado_a_id and asignado_a_id.isdigit()) else None

                EventoAgenda.objects.create(
                    colegio=colegio,
                    creado_por=request.user,
                    asignado_a=asignado_obj,
                    es_para_todos=es_para_todos,
                    es_recurrente=False,
                    titulo=titulo,
                    tipo=tipo,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    lugar=lugar if lugar else None,
                    curso=curso_obj,
                    descripcion=descripcion if descripcion else None
                )
                messages.success(request, f"¡Reunión/Evento '{titulo}' agendado exitosamente!")
                return redirect(f"{reverse('calendario_escolar')}?tab=reuniones")

    # Parámetros y Filtros de Vista
    tab_activa = request.GET.get('tab', 'horario')
    docente_id_filtro = request.GET.get('docente_id')
    curso_id_filtro = request.GET.get('curso_id')

    # Si es profesor y no eligió filtro, pre-cargar su propio horario
    if is_profesor and not is_admin and not docente_id_filtro:
        docente_id_filtro = str(request.user.id)

    personal = MiembroColegio.objects.filter(colegio=colegio, activo=True).select_related('usuario', 'rol').order_by('usuario__first_name', 'usuario__last_name')
    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('orden', 'nombre')
    asignaturas = Asignatura.objects.filter(colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'nombre')

    # ── 1. MATRIZ DE HORARIO SEMANAL DE CLASES ────────────────────────────────
    clases_qs = EventoAgenda.objects.filter(colegio=colegio, tipo='clase', es_recurrente=True).select_related('asignatura', 'curso', 'asignado_a')

    docente_seleccionado = None
    if docente_id_filtro and docente_id_filtro.isdigit():
        clases_qs = clases_qs.filter(asignado_a_id=int(docente_id_filtro))
        docente_seleccionado = User.objects.filter(id=int(docente_id_filtro)).first()

    curso_seleccionado = None
    if curso_id_filtro and curso_id_filtro.isdigit():
        clases_qs = clases_qs.filter(curso_id=int(curso_id_filtro))
        curso_seleccionado = CursoColegio.objects.filter(id=int(curso_id_filtro)).first()

    # Bloques estándar diarios
    bloques_horarios = [
        {'id': 1, 'nombre': 'Bloque 1', 'hora_inicio': '08:00', 'hora_fin': '09:30', 'tipo': 'clase'},
        {'id': 2, 'nombre': 'Bloque 2', 'hora_inicio': '09:45', 'hora_fin': '11:15', 'tipo': 'clase'},
        {'id': 3, 'nombre': 'Bloque 3', 'hora_inicio': '11:30', 'hora_fin': '13:00', 'tipo': 'clase'},
        {'id': 0, 'nombre': 'Almuerzo / Colación', 'hora_inicio': '13:00', 'hora_fin': '14:00', 'tipo': 'recreo'},
        {'id': 4, 'nombre': 'Bloque 4', 'hora_inicio': '14:00', 'hora_fin': '15:30', 'tipo': 'clase'},
        {'id': 5, 'nombre': 'Bloque 5', 'hora_inicio': '15:45', 'hora_fin': '17:15', 'tipo': 'clase'},
    ]

    dias_semana_nombres = [
        (0, 'Lunes'), (1, 'Martes'), (2, 'Miércoles'), (3, 'Jueves'), (4, 'Viernes')
    ]

    # Construir grilla {bloque_idx: {dia_idx: [eventos]}}
    grilla_horario = []
    for blk in bloques_horarios:
        fila_dias = []
        for dia_num, dia_nom in dias_semana_nombres:
            # Buscar clases que coincidan con el día y hora aproximada
            if blk['tipo'] == 'clase':
                eventos_celda = [
                    ev for ev in clases_qs
                    if ev.dia_semana == dia_num and ev.fecha_inicio and ev.fecha_inicio.strftime('%H:%M') <= blk['hora_inicio'] <= (ev.fecha_fin.strftime('%H:%M') if ev.fecha_fin else '23:59')
                ]
            else:
                eventos_celda = []
            fila_dias.append({
                'dia_num': dia_num,
                'dia_nom': dia_nom,
                'eventos': eventos_celda
            })
        grilla_horario.append({
            'bloque': blk,
            'dias': fila_dias
        })

    # ── 2. AGENDA DE REUNIONES Y EVENTOS EXTRAORDINARIOS ─────────────────────
    hoy_date = timezone.now().date()
    reuniones_qs = EventoAgenda.objects.filter(colegio=colegio).exclude(tipo='clase').select_related('curso', 'asignado_a', 'creado_por').order_by('fecha_inicio')
    
    proximas_reuniones = reuniones_qs.filter(fecha_inicio__date__gte=hoy_date)[:20]
    reuniones_hoy = reuniones_qs.filter(fecha_inicio__date=hoy_date)

    # Métricas Globales de Conteo
    total_clases = EventoAgenda.objects.filter(colegio=colegio, tipo='clase').count()
    total_reuniones = reuniones_qs.filter(tipo__in=['reunion', 'reunion_apoderados', 'consejo_profesores', 'entrevista']).count()
    total_evaluaciones = reuniones_qs.filter(tipo='evaluacion').count()
    total_actividades = reuniones_qs.filter(tipo__in=['actividad', 'feriado']).count()

    abs_ical_url = request.build_absolute_uri(reverse('exportar_ical_agenda')) + f"?colegio_id={colegio.id}"
    webcal_url = abs_ical_url.replace('https://', 'webcal://').replace('http://', 'webcal://')
    google_cal_feed_url = f"https://calendar.google.com/calendar/r?cid={webcal_url}"

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'is_profesor': is_profesor,
        'active_page': 'calendario',
        'tab_activa': tab_activa,
        'docente_id_filtro': int(docente_id_filtro) if (docente_id_filtro and docente_id_filtro.isdigit()) else None,
        'docente_seleccionado': docente_seleccionado,
        'curso_id_filtro': int(curso_id_filtro) if (curso_id_filtro and curso_id_filtro.isdigit()) else None,
        'curso_seleccionado': curso_seleccionado,
        'personal': personal,
        'cursos': cursos,
        'asignaturas': asignaturas,
        'dias_semana_nombres': dias_semana_nombres,
        'bloques_horarios': bloques_horarios,
        'grilla_horario': grilla_horario,
        'total_clases_programadas': clases_qs.count(),
        'proximas_reuniones': proximas_reuniones,
        'reuniones_hoy': reuniones_hoy,
        'total_clases': total_clases,
        'total_reuniones': total_reuniones,
        'total_evaluaciones': total_evaluaciones,
        'total_actividades': total_actividades,
        'hoy': timezone.now(),
        'abs_ical_url': abs_ical_url,
        'webcal_url': webcal_url,
        'google_cal_feed_url': google_cal_feed_url,
    }
    return render(request, 'colegios/calendario_escolar.html', context)


@login_required
def api_eventos_calendario_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return JsonResponse([], safe=False)

    docente_id = request.GET.get('docente_id')
    filtro_tipo = request.GET.get('tipo', '') # 'reuniones', 'clases', 'todos'
    from colegios.models import EventoAgenda
    from django.db.models import Q

    qs = EventoAgenda.objects.filter(colegio=colegio)

    if filtro_tipo == 'reuniones':
        qs = qs.exclude(tipo='clase')
    elif filtro_tipo == 'clases':
        qs = qs.filter(tipo='clase')

    if docente_id and docente_id.isdigit():
        target_uid = int(docente_id)
        qs = qs.filter(Q(asignado_a_id=target_uid) | Q(creado_por_id=target_uid) | Q(es_para_todos=True))

    events_list = []
    color_map = {
        'clase': '#7C5CFC',                 # Morado
        'reunion_apoderados': '#D97706',    # Ámbar / Naranja
        'consejo_profesores': '#2563EB',    # Azul
        'entrevista': '#0891B2',            # Cian
        'evaluacion': '#E11D48',            # Rojo carmesí
        'actividad': '#059669',             # Verde
        'feriado': '#64748B',               # Gris pizarra
        'reunion': '#D97706',               # Ámbar
    }

    for ev in qs:
        item = {
            'id': ev.id,
            'title': ev.titulo,
            'backgroundColor': color_map.get(ev.tipo, '#7C5CFC'),
            'borderColor': color_map.get(ev.tipo, '#7C5CFC'),
            'extendedProps': {
                'tipo': ev.get_tipo_display(),
                'tipo_raw': ev.tipo,
                'lugar': ev.lugar or 'Sin definir',
                'curso': ev.curso.nombre if ev.curso else 'General',
                'docente': ev.asignado_a.get_full_name() if ev.asignado_a else ('Todos' if ev.es_para_todos else (ev.creado_por.get_full_name() if ev.creado_por else 'Institución')),
                'descripcion': ev.descripcion or '',
                'es_recurrente': ev.es_recurrente,
            }
        }
        if ev.es_recurrente and ev.dia_semana is not None:
            fc_dow = (ev.dia_semana + 1) % 7
            item['daysOfWeek'] = [fc_dow]
            item['startTime'] = ev.fecha_inicio.strftime('%H:%M:%S')
            if ev.fecha_fin:
                item['endTime'] = ev.fecha_fin.strftime('%H:%M:%S')
        else:
            item['start'] = ev.fecha_inicio.isoformat()
            if ev.fecha_fin:
                item['end'] = ev.fecha_fin.isoformat()

        events_list.append(item)

    return JsonResponse(events_list, safe=False)



def exportar_ical_agenda_view(request):
    from django.http import HttpResponse
    from colegios.models import Colegio, EventoAgenda

    colegio = None
    if request.user.is_authenticated:
        colegio = obtener_colegio_usuario(request.user)

    colegio_id = request.GET.get('colegio_id')
    if not colegio and colegio_id and colegio_id.isdigit():
        colegio = Colegio.objects.filter(id=colegio_id).first()

    if not colegio:
        colegio = Colegio.objects.first()

    if not colegio:
        return HttpResponse("Colegio no encontrado.", status=404)

    eventos = EventoAgenda.objects.filter(colegio=colegio).order_by('fecha_inicio')

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Eduteka//Agenda Escolar//ES",
        "CALSCALE:GREGORIAN",
        f"X-WR-CALNAME:Agenda {colegio.nombre}",
    ]

    for ev in eventos:
        dtstart = ev.fecha_inicio.strftime('%Y%m%dT%H%M%SZ')
        dtstamp = ev.fecha_creacion.strftime('%Y%m%dT%H%M%SZ')
        summary = ev.titulo.replace('\n', ' ')
        desc = (ev.descripcion or '').replace('\n', ' ')
        location = (ev.lugar or '').replace('\n', ' ')

        lines.extend([
            "BEGIN:VEVENT",
            f"UID:eduteka-event-{ev.id}@{colegio.id}",
            f"DTSTAMP:{dtstamp}",
            f"DTSTART:{dtstart}",
            f"SUMMARY:{summary}",
            f"DESCRIPTION:{desc}",
            f"LOCATION:{location}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    ics_content = "\r\n".join(lines)

    response = HttpResponse(ics_content, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="agenda_{colegio.id}.ics"'
    return response



@login_required
def eliminar_evento_agenda_view(request, evento_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    from colegios.models import EventoAgenda
    evento = get_object_or_404(EventoAgenda, id=evento_id, colegio=colegio)

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    if not is_admin and evento.creado_por != request.user:
        messages.error(request, "No tienes permiso para eliminar este evento.")
        return redirect('calendario_escolar')

    evento.delete()
    messages.success(request, "Evento eliminado de la agenda.")
    return redirect('calendario_escolar')


@require_POST
def actualizar_modulo_colegio(request):
    try:
        data = json.loads(request.body)
        colegio_id = data.get('colegio_id')
        modulo = data.get('modulo')
        estado = data.get('estado')

        if colegio_id:
            colegio = Colegio.objects.get(id=colegio_id)
            config, created = ConfiguracionModulos.objects.get_or_create(colegio=colegio)
            if hasattr(config, modulo):
                setattr(config, modulo, estado)
                config.save()
        else:
            from planes.models import Modulo as ModuloModel
            mod_obj = ModuloModel.objects.filter(nombre__icontains=modulo.replace('_', ' ')).first()
            if mod_obj:
                mod_obj.activo = estado
                mod_obj.save()

        return JsonResponse({'status': 'ok', 'message': f'Módulo {modulo} actualizado'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@login_required
def ver_detalle_colegio(request, pk):
    colegio = get_object_or_404(Colegio, pk=pk)
    suscripcion = getattr(colegio, 'suscripcion', None)
    config_modulos, _ = ConfiguracionModulos.objects.get_or_create(colegio=colegio)
    context = {
        'colegio': colegio,
        'suscripcion': suscripcion,
        'config_modulos': config_modulos,
    }
    return render(request, 'colegios/detalle_colegio.html', context)


@login_required
def editar_colegio(request, pk):
    colegio = get_object_or_404(Colegio, pk=pk)

    if request.method == 'POST':
        colegio.nombre = request.POST.get('nombre', colegio.nombre).strip()
        colegio.nombre_corto = request.POST.get('nombre_corto', colegio.nombre_corto)
        colegio.correo_institucional = request.POST.get('correo_institucional', colegio.correo_institucional).strip()
        colegio.telefono = request.POST.get('telefono', colegio.telefono).strip()
        colegio.direccion = request.POST.get('direccion', colegio.direccion)
        colegio.ciudad_comuna = request.POST.get('ciudad_comuna', colegio.ciudad_comuna).strip()
        colegio.tipo_institucion = request.POST.get('tipo_institucion', colegio.tipo_institucion)
        colegio.cantidad_alumnos = request.POST.get('cantidad_alumnos', colegio.cantidad_alumnos)
        colegio.estado = request.POST.get('estado', colegio.estado)

        if 'logo' in request.FILES:
            colegio.logo = request.FILES['logo']

        colegio.save()
        messages.success(request, f"Datos de {colegio.nombre} actualizados con éxito.")
        return redirect('dashboard_superadmin_colegios')

    return render(request, 'colegios/editar_colegio.html', {'colegio': colegio})


@login_required
def editar_suscripcion_colegio(request, colegio_id):
    from planes.models import Plan
    colegio = get_object_or_404(Colegio, id=colegio_id)
    suscripcion = getattr(colegio, 'suscripcion', None)
    planes = Plan.objects.filter(activo=True)

    if request.method == 'POST':
        plan_id = request.POST.get('plan_id')
        tipo_facturacion = request.POST.get('tipo_facturacion', 'mensual')
        estado = request.POST.get('estado', 'activa')

        if plan_id:
            plan_obj = get_object_or_404(Plan, id=plan_id)
            monto = plan_obj.precio_anual if tipo_facturacion == 'anual' else plan_obj.precio_mensual

            Suscripcion.objects.update_or_create(
                colegio=colegio,
                defaults={
                    'plan': plan_obj,
                    'tipo_facturacion': tipo_facturacion,
                    'monto': monto,
                    'estado': estado,
                }
            )
            messages.success(request, f"Plan y suscripción de {colegio.nombre} actualizados con éxito.")
            return redirect('ver_detalle_colegio', pk=colegio.id)
        else:
            messages.error(request, "Debe seleccionar un plan de suscripción.")

    context = {
        'colegio': colegio,
        'suscripcion': suscripcion,
        'planes': planes,
    }
    return render(request, 'colegios/editar_suscripcion.html', context)


@login_required
def asignar_profesor_jefe_view(request, seccion_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes un colegio asociado.")
        return redirect('dashboard_usuario')

    seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
    if request.method == 'POST':
        profesor_jefe_id = request.POST.get('profesor_jefe_id')
        if profesor_jefe_id and profesor_jefe_id != 'none':
            docente_obj = User.objects.filter(id=profesor_jefe_id).first()
            seccion.profesor_jefe = docente_obj
            seccion.save()
            nombre_doc = docente_obj.get_full_name() or docente_obj.username
            messages.success(request, f"¡{nombre_doc} asignado/a como Profesor(a) Jefe de {seccion.curso.nombre} - {seccion.nombre}!")
        else:
            seccion.profesor_jefe = None
            seccion.save()
            messages.info(request, f"Se ha desasignado la jefatura de {seccion.curso.nombre} - {seccion.nombre}.")
    return redirect('listar_cursos')


@login_required
def mis_cursos_docente_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    # 1. Jefaturas de curso donde es Profesor Jefe
    jefaturas = SeccionCurso.objects.filter(
        curso__colegio=colegio,
        profesor_jefe=request.user,
        activo=True
    ).select_related('curso').prefetch_related('estudiantes', 'curso__asignaturas')

    jefaturas_data = []
    for sec in jefaturas:
        alumnos_count = sec.estudiantes.filter(activo=True).count()
        asig_count = sec.curso.asignaturas.filter(activo=True).count()
        jefaturas_data.append({
            'seccion': sec,
            'alumnos_count': alumnos_count,
            'asignaturas_count': asig_count
        })

    # 2. Asignaturas dictadas por el docente
    asignaturas_dictadas = Asignatura.objects.filter(
        colegio=colegio,
        docente=request.user,
        activo=True
    ).select_related('curso').prefetch_related('curso__secciones')

    asignaturas_data = []
    for asig in asignaturas_dictadas:
        secciones_asig = asig.curso.secciones.filter(activo=True)
        total_alumnos = Estudiante.objects.filter(seccion__in=secciones_asig, activo=True).count()
        asignaturas_data.append({
            'asignatura': asig,
            'secciones': secciones_asig,
            'total_alumnos': total_alumnos,
        })

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'mis_cursos',
        'jefaturas': jefaturas_data,
        'jefaturas_count': len(jefaturas_data),
        'asignaturas': asignaturas_data,
        'asignaturas_count': len(asignaturas_data),
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/mis_cursos.html', context)


@login_required
def estadisticas_colegio_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    if not is_admin:
        messages.error(request, "El módulo de estadísticas institucionales está reservado para el Equipo Directivo y Administradores.")
        return redirect('dashboard_usuario')

    hoy = timezone.now().date()
    estudiantes_qs = Estudiante.objects.filter(colegio=colegio, activo=True).select_related('seccion', 'seccion__curso')
    total_estudiantes = estudiantes_qs.count()

    # 1. GÉNERO
    hombres_count = estudiantes_qs.filter(genero='masculino').count()
    mujeres_count = estudiantes_qs.filter(genero='femenino').count()
    otros_count = estudiantes_qs.filter(genero='otro').count()
    no_informa_count = total_estudiantes - (hombres_count + mujeres_count + otros_count)

    pct_hombres = round((hombres_count / total_estudiantes * 100), 1) if total_estudiantes else 0
    pct_mujeres = round((mujeres_count / total_estudiantes * 100), 1) if total_estudiantes else 0
    pct_otros = round(((otros_count + no_informa_count) / total_estudiantes * 100), 1) if total_estudiantes else 0

    # 2. RANGOS ETARIOS (EDADES)
    edad_menor_6 = 0
    edad_6_9 = 0
    edad_10_13 = 0
    edad_14_17 = 0
    edad_18_mas = 0
    edades_lista = []

    for est in estudiantes_qs:
        if est.fecha_nacimiento:
            fn = est.fecha_nacimiento
            age = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))
            edades_lista.append(age)
            if age < 6:
                edad_menor_6 += 1
            elif 6 <= age <= 9:
                edad_6_9 += 1
            elif 10 <= age <= 13:
                edad_10_13 += 1
            elif 14 <= age <= 17:
                edad_14_17 += 1
            else:
                edad_18_mas += 1
        else:
            # Estimación por nivel
            edad_10_13 += 1
            edades_lista.append(12)

    edad_promedio = round(sum(edades_lista) / len(edades_lista), 1) if edades_lista else 0

    # 3. PROGRAMA PIE & INCLUSIÓN
    pie_count = estudiantes_qs.filter(es_pie=True).count()
    regular_count = total_estudiantes - pie_count
    pie_neep = estudiantes_qs.filter(es_pie=True, tipo_pie='neep').count()
    pie_neet = estudiantes_qs.filter(es_pie=True, tipo_pie='neet').count()
    pie_otros = pie_count - (pie_neep + pie_neet)
    pct_pie = round((pie_count / total_estudiantes * 100), 1) if total_estudiantes else 0

    # 4. CALIFICACIONES & RANGOS DE NOTAS
    from calificaciones.models import Nota, Evaluacion
    notas_qs = Nota.objects.filter(evaluacion__colegio=colegio)
    total_notas = notas_qs.count()

    notas_sobresaliente = notas_qs.filter(valor__gte=6.0).count() # 6.0 - 7.0
    notas_bueno = notas_qs.filter(valor__gte=5.0, valor__lt=6.0).count() # 5.0 - 5.9
    notas_suficiente = notas_qs.filter(valor__gte=4.0, valor__lt=5.0).count() # 4.0 - 4.9
    notas_insuficiente = notas_qs.filter(valor__lt=4.0).count() # < 4.0

    from django.db.models import Avg
    promedio_global_val = notas_qs.aggregate(Avg('valor'))['valor__avg']
    promedio_global = round(float(promedio_global_val), 1) if promedio_global_val else 0.0

    aprobadas = notas_sobresaliente + notas_bueno + notas_suficiente
    tasa_aprobacion_global = round((aprobadas / total_notas * 100), 1) if total_notas else 0

    # Promedio por Nivel
    cursos_qs = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nivel', 'nombre')
    rendimiento_cursos = []
    for c in cursos_qs:
        n_curso = Nota.objects.filter(evaluacion__colegio=colegio, evaluacion__seccion__curso=c)
        avg_c = n_curso.aggregate(Avg('valor'))['valor__avg']
        rendimiento_cursos.append({
            'nombre': c.nombre,
            'nivel': c.get_nivel_display(),
            'promedio': round(float(avg_c), 1) if avg_c else '-',
            'total_notas': n_curso.count(),
            'aprobadas_pct': round((n_curso.filter(valor__gte=4.0).count() / n_curso.count() * 100), 1) if n_curso.exists() else 0
        })

    # 5. ASISTENCIA INSTITUCIONAL
    from asistencia.models import DetalleAsistencia
    detalles_asistencia_qs = DetalleAsistencia.objects.filter(registro__seccion__curso__colegio=colegio)
    total_asistencias = detalles_asistencia_qs.count()

    presentes = detalles_asistencia_qs.filter(estado='presente').count()
    ausentes = detalles_asistencia_qs.filter(estado='ausente').count()
    atrasados = detalles_asistencia_qs.filter(estado='tarde').count()
    justificados = detalles_asistencia_qs.filter(estado='justificado').count()

    asistencia_positiva = presentes + atrasados + justificados
    pct_asistencia_global = round((asistencia_positiva / total_asistencias * 100), 1) if total_asistencias else 0

    # 6. CONVIVENCIA ESCOLAR
    from colegios.models import AnotacionEstudiante
    anotaciones_qs = AnotacionEstudiante.objects.filter(estudiante__colegio=colegio)
    total_anotaciones = anotaciones_qs.count()

    anotaciones_pos = anotaciones_qs.filter(tipo='positiva').count()
    anotaciones_neg = anotaciones_qs.filter(tipo='negativa').count()
    anotaciones_dest = anotaciones_qs.filter(tipo='destacada').count()

    # 7. DISTRIBUCIÓN POR NIVELES
    dist_niveles = {}
    for est in estudiantes_qs:
        if est.seccion and est.seccion.curso:
            niv = est.seccion.curso.get_nivel_display()
            dist_niveles[niv] = dist_niveles.get(niv, 0) + 1

    niveles_labels = list(dist_niveles.keys())
    niveles_counts = list(dist_niveles.values())

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'estadisticas',
        'hoy': timezone.now(),
        # Totales
        'total_estudiantes': total_estudiantes,
        'total_notas': total_notas,
        'total_asistencias': total_asistencias,
        'total_anotaciones': total_anotaciones,
        'promedio_global': promedio_global,
        'tasa_aprobacion_global': tasa_aprobacion_global,
        'pct_asistencia_global': pct_asistencia_global,
        'edad_promedio': edad_promedio,
        # Género
        'hombres_count': hombres_count,
        'mujeres_count': mujeres_count,
        'otros_count': otros_count + no_informa_count,
        'pct_hombres': pct_hombres,
        'pct_mujeres': pct_mujeres,
        'pct_otros': pct_otros,
        # Rangos Etarios
        'edad_menor_6': edad_menor_6,
        'edad_6_9': edad_6_9,
        'edad_10_13': edad_10_13,
        'edad_14_17': edad_14_17,
        'edad_18_mas': edad_18_mas,
        # Notas
        'notas_sobresaliente': notas_sobresaliente,
        'notas_bueno': notas_bueno,
        'notas_suficiente': notas_suficiente,
        'notas_insuficiente': notas_insuficiente,
        'rendimiento_cursos': rendimiento_cursos,
        # PIE
        'pie_count': pie_count,
        'regular_count': regular_count,
        'pie_neep': pie_neep,
        'pie_neet': pie_neet,
        'pie_otros': pie_otros,
        'pct_pie': pct_pie,
        # Asistencia
        'presentes': presentes,
        'ausentes': ausentes,
        'atrasados': atrasados,
        'justificados': justificados,
        # Convivencia
        'anotaciones_pos': anotaciones_pos,
        'anotaciones_neg': anotaciones_neg,
        'anotaciones_dest': anotaciones_dest,
        # Niveles
        'niveles_labels': niveles_labels,
        'niveles_counts': niveles_counts,
    }
    return render(request, 'colegios/estadisticas.html', context)


@login_required
def exportar_estadisticas_excel_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "Colegio no encontrado.")
        return redirect('dashboard_usuario')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    if not is_admin:
        messages.error(request, "Acceso no autorizado.")
        return redirect('dashboard_usuario')

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Quitar hoja por defecto

    # Estilos Globales
    header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # -------------------------------------------------------------
    # HOJA 1: RESUMEN EJECUTIVO
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Resumen Ejecutivo")
    ws1.views.sheetView[0].showGridLines = True

    ws1['A1'] = f"EDUTEKA - INFORME ESTADÍSTICO INSTITUCIONAL"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Establecimiento: {colegio.nombre} | Fecha de Emisión: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = sub_font

    kpis = [
        ("Métrica Institucional", "Valor Global", "Descripción / Estado"),
        ("Matrícula Total Activa", Estudiante.objects.filter(colegio=colegio, activo=True).count(), "Estudiantes matriculados"),
        ("Alumnos Programa PIE", Estudiante.objects.filter(colegio=colegio, activo=True, es_pie=True).count(), "Necesidades Educativas Especiales"),
        ("Cursos Habilitados", CursoColegio.objects.filter(colegio=colegio, activo=True).count(), "Niveles en funcionamiento"),
        ("Secciones Habilitadas", SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).count(), "Aulas activas"),
        ("Asignaturas en Malla", Asignatura.objects.filter(colegio=colegio, activo=True).count(), "Cátedras impartidas"),
        ("Personal y Funcionarios", MiembroColegio.objects.filter(colegio=colegio, activo=True).count(), "Docentes y administrativos"),
    ]

    from calificaciones.models import Nota
    notas_qs = Nota.objects.filter(evaluacion__colegio=colegio)
    from django.db.models import Avg
    avg_n = notas_qs.aggregate(Avg('valor'))['valor__avg']
    prom_gral = round(float(avg_n), 1) if avg_n else 0.0
    kpis.append(("Promedio General de Notas", prom_gral, "Escala 1.0 a 7.0"))
    
    tasa_aprob = round((notas_qs.filter(valor__gte=4.0).count() / notas_qs.count() * 100), 1) if notas_qs.exists() else 0
    kpis.append(("Tasa Global de Aprobación", f"{tasa_aprob}%", "Calificaciones >= 4.0"))

    from asistencia.models import DetalleAsistencia
    detalles_asist = DetalleAsistencia.objects.filter(registro__seccion__curso__colegio=colegio)
    pct_asist = round((detalles_asist.filter(estado__in=['presente', 'tarde', 'justificado']).count() / detalles_asist.count() * 100), 1) if detalles_asist.exists() else 0
    kpis.append(("Asistencia Global del Colegio", f"{pct_asist}%", "Presentismo escolar"))

    for row_idx, row_data in enumerate(kpis, start=4):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            if row_idx == 4:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
            else:
                cell.font = bold_font if col_idx == 1 else regular_font
                cell.alignment = Alignment(horizontal="center" if col_idx == 2 else "left", vertical="center")

    # -------------------------------------------------------------
    # HOJA 2: DEMOGRAFÍA & ESTUDIANTES
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Demografía y Alumnos")
    ws2.views.sheetView[0].showGridLines = True

    ws2['A1'] = "NÓMINA Y DEMOGRAFÍA DE ESTUDIANTES"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Colegio: {colegio.nombre}"
    ws2['A2'].font = sub_font

    headers_demo = ["#", "Estudiante", "RUT", "Curso", "Sección", "Género", "Fecha Nac.", "Edad Aprox.", "PIE", "Tipo NEE"]
    for col_idx, h in enumerate(headers_demo, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    hoy = timezone.now().date()
    estudiantes_list = Estudiante.objects.filter(colegio=colegio, activo=True).select_related('seccion', 'seccion__curso').order_by('seccion__curso__nombre', 'nombre_completo')
    
    for row_idx, est in enumerate(estudiantes_list, start=5):
        age_str = "-"
        if est.fecha_nacimiento:
            fn = est.fecha_nacimiento
            age_str = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))

        row_data = [
            row_idx - 4,
            est.nombre_completo,
            est.rut or "-",
            est.seccion.curso.nombre if est.seccion and est.seccion.curso else "-",
            est.seccion.nombre if est.seccion else "-",
            est.get_genero_display() if est.genero else "No Informa",
            est.fecha_nacimiento.strftime('%d/%m/%Y') if est.fecha_nacimiento else "-",
            age_str,
            "SÍ" if est.es_pie else "NO",
            est.get_tipo_pie_display() if est.es_pie and est.tipo_pie else "-"
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx not in [2] else "left", vertical="center")

    # -------------------------------------------------------------
    # HOJA 3: RENDIMIENTO Y CALIFICACIONES
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Rendimiento Académico")
    ws3.views.sheetView[0].showGridLines = True

    ws3['A1'] = "ESTADÍSTICAS DE RENDIMIENTO ACADÉMICO"
    ws3['A1'].font = title_font
    ws3['A2'] = f"Colegio: {colegio.nombre}"
    ws3['A2'].font = sub_font

    headers_acad = ["Curso / Nivel", "Total Notas", "Sobresaliente (6.0-7.0)", "Bueno (5.0-5.9)", "Suficiente (4.0-4.9)", "Insuficiente (<4.0)", "Promedio", "% Aprobación"]
    for col_idx, h in enumerate(headers_acad, start=1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    cursos_all = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nivel', 'nombre')
    for row_idx, c in enumerate(cursos_all, start=5):
        n_c = Nota.objects.filter(evaluacion__colegio=colegio, evaluacion__seccion__curso=c)
        tot_c = n_c.count()
        sobr_c = n_c.filter(valor__gte=6.0).count()
        buen_c = n_c.filter(valor__gte=5.0, valor__lt=6.0).count()
        suf_c = n_c.filter(valor__gte=4.0, valor__lt=5.0).count()
        insuf_c = n_c.filter(valor__lt=4.0).count()
        avg_c = n_c.aggregate(Avg('valor'))['valor__avg']
        prom_c_str = round(float(avg_c), 1) if avg_c else "-"
        pct_aprob_c = f"{round((tot_c - insuf_c) / tot_c * 100, 1)}%" if tot_c > 0 else "-"

        row_data = [c.nombre, tot_c, sobr_c, buen_c, suf_c, insuf_c, prom_c_str, pct_aprob_c]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

    # Autoajuste de ancho de columnas en todas las hojas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len and len(val_str) < 50:
                        max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"estadisticas_{colegio.nombre.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==============================================================================
# LÓGICA Y VISTAS DEL MÓDULO DE FINANZAS, CAJA CHICA Y FACTURAS
# ==============================================================================

def inicializar_datos_finanzas(colegio):
    from .models import CuentaFinanciera, CategoriaFinanciera, ProyectoEscolar
    from decimal import Decimal

    if not CuentaFinanciera.objects.filter(colegio=colegio).exists():
        # Bolsas de Subvención
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Bolsa Subvención Escolar Preferencial (SEP)",
            tipo="subvencion_sep",
            fondo_asociado="sep",
            banco="Banco Estado",
            numero_cuenta="SEP-00129-3",
            saldo_inicial=8500000.0,
            saldo_actual=8500000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Bolsa Programa de Integración Escolar (PIE)",
            tipo="subvencion_pie",
            fondo_asociado="pie",
            banco="Banco Estado",
            numero_cuenta="PIE-00418-7",
            saldo_inicial=5200000.0,
            saldo_actual=5200000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Bolsa Subvención General / Operacional",
            tipo="subvencion_general",
            fondo_asociado="subvencion_general",
            banco="Banco Santander",
            numero_cuenta="0-000-00-12345-6",
            saldo_inicial=12000000.0,
            saldo_actual=12000000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Fondo de Mantenimiento e Infraestructura",
            tipo="fondo_mantenimiento",
            fondo_asociado="mantenimiento",
            banco="Banco Estado",
            numero_cuenta="MNT-99214-1",
            saldo_inicial=3800000.0,
            saldo_actual=3800000.0
        )

        # Cajas Chicas Especializadas
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Caja Chica Útiles Escolares & Fungibles",
            tipo="caja_chica_utiles",
            fondo_asociado="sep",
            saldo_inicial=350000.0,
            saldo_actual=350000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Caja Chica Convivencia & Inspectoría",
            tipo="caja_chica_convivencia",
            fondo_asociado="subvencion_general",
            saldo_inicial=200000.0,
            saldo_actual=200000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Caja Chica Ciencias & Laboratorios",
            tipo="caja_chica_ciencias",
            fondo_asociado="sep",
            saldo_inicial=250000.0,
            saldo_actual=250000.0
        )
        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre="Caja Chica Dirección / Rectoría",
            tipo="caja_chica_rectoria",
            fondo_asociado="fondos_propios",
            saldo_inicial=300000.0,
            saldo_actual=300000.0
        )

    categorias_base = [
        ("Subvención Escolar / Mineduc", "ingreso", "bi-bank", "#10B981"),
        ("Aporte Fondo SEP", "ingreso", "bi-award-fill", "#7C5CFC"),
        ("Aporte Fondo PIE", "ingreso", "bi-person-hearts", "#3B82F6"),
        ("Aportes Centro de Padres & Donaciones", "ingreso", "bi-people-fill", "#8B5CF6"),
        ("Otros Ingresos", "ingreso", "bi-plus-circle", "#64748B"),
        ("Útiles Escolares & Material Fungible", "egreso", "bi-journal-check", "#F59E0B"),
        ("Recursos Pedagógicos & Aula", "egreso", "bi-pencil-square", "#7C5CFC"),
        ("Material Didáctico Sensorial (PIE)", "egreso", "bi-puzzle-fill", "#3B82F6"),
        ("Mantención e Infraestructura", "egreso", "bi-tools", "#6366F1"),
        ("Equipamiento Tecnológico", "egreso", "bi-laptop", "#10B981"),
        ("Servicios Básicos (Luz / Agua / Gas)", "egreso", "bi-lightning-charge", "#EF4444"),
        ("Caja Chica Gastos Menores", "egreso", "bi-wallet2", "#06B6D4"),
        ("Eventos y Actividades Formativas", "egreso", "bi-trophy", "#10B981"),
    ]

    for nombre, tipo, icono, color in categorias_base:
        CategoriaFinanciera.objects.get_or_create(
            colegio=colegio,
            nombre=nombre,
            tipo=tipo,
            defaults={'icono': icono, 'color': color, 'activo': True}
        )

    # Proyectos Base si no existen
    if not ProyectoEscolar.objects.filter(colegio=colegio).exists():
        ProyectoEscolar.objects.create(
            colegio=colegio,
            codigo="PRY-SEP-01",
            nombre="Campaña Útiles Escolares Alumnos Prioritarios 2026",
            descripcion="Adquisición de cuadernos, resmas, lápices, mochilas y estuches para el 100% de estudiantes prioritarios según registro social.",
            tipo_fondo="sep",
            categoria_supereduc="utiles_escolares",
            presupuesto_asignado=Decimal('1850000.0'),
            estado="en_ejecucion"
        )
        ProyectoEscolar.objects.create(
            colegio=colegio,
            codigo="PRY-PIE-02",
            nombre="Implementación Sala Sensorial y Recursos Fonoaudiológicos",
            descripcion="Equipamiento con materiales didácticos sensoriales, software de fonoaudiología y mobiliario ergonómico para aula de recursos PIE.",
            tipo_fondo="pie",
            categoria_supereduc="pedagogico",
            presupuesto_asignado=Decimal('2400000.0'),
            estado="en_ejecucion"
        )
        ProyectoEscolar.objects.create(
            colegio=colegio,
            codigo="PRY-MNT-03",
            nombre="Mantención Eléctrica y Pintura Aulas Pabellón B",
            descripcion="Renovación de iluminación LED, pintura lavable de salas y reparación de canaletas perimetrales.",
            tipo_fondo="mantenimiento",
            categoria_supereduc="infraestructura",
            presupuesto_asignado=Decimal('3100000.0'),
            estado="en_ejecucion"
        )


@login_required
def finanzas_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import RolPermiso, MiembroPermiso
    tiene_permiso = is_admin
    if not tiene_permiso and miembro and miembro.rol:
        tiene_permiso = RolPermiso.objects.filter(rol=miembro.rol, modulo__nombre__iexact='Finanzas', puede_ver=True).exists()
    if not tiene_permiso and miembro:
        tiene_permiso = MiembroPermiso.objects.filter(miembro=miembro, modulo__nombre__iexact='Finanzas', puede_ver=True).exists()

    if not tiene_permiso:
        messages.error(request, "Acceso restringido. Se requieren permisos de Administración para ver las finanzas del colegio.")
        return redirect('dashboard_usuario')

    # Inicializar cuentas y proyectos si es primera vez
    inicializar_datos_finanzas(colegio)

    from .models import CuentaFinanciera, CategoriaFinanciera, MovimientoFinanciero, FacturaGasto, ProyectoEscolar
    from django.db.models import Sum, Q
    from decimal import Decimal

    # Cuentas activas
    cuentas = CuentaFinanciera.objects.filter(colegio=colegio, activo=True).order_by('tipo', 'nombre')
    cajas_chicas = cuentas.filter(tipo__startswith='caja_chica')
    bolsas_subvencion = cuentas.exclude(tipo__startswith='caja_chica')
    
    # Proyectos activos
    proyectos = ProyectoEscolar.objects.filter(colegio=colegio, activo=True).order_by('-fecha_inicio', '-id')

    categorias_ingreso = CategoriaFinanciera.objects.filter(colegio=colegio, tipo='ingreso', activo=True).order_by('nombre')
    categorias_egreso = CategoriaFinanciera.objects.filter(colegio=colegio, tipo='egreso', activo=True).order_by('nombre')
    todas_categorias = CategoriaFinanciera.objects.filter(colegio=colegio, activo=True).order_by('tipo', 'nombre')

    # Filtros
    cuenta_filtro = request.GET.get('cuenta')
    proyecto_filtro = request.GET.get('proyecto')
    fondo_filtro = request.GET.get('fondo')
    tipo_filtro = request.GET.get('tipo')
    categoria_filtro = request.GET.get('categoria')
    busqueda = request.GET.get('q', '').strip()
    tab_activa = request.GET.get('tab', 'fondos')

    movimientos_qs = MovimientoFinanciero.objects.filter(colegio=colegio).select_related('cuenta', 'proyecto', 'categoria', 'registrado_por').order_by('-fecha', '-id')

    if cuenta_filtro and cuenta_filtro.isdigit():
        movimientos_qs = movimientos_qs.filter(cuenta_id=int(cuenta_filtro))
    if proyecto_filtro and proyecto_filtro.isdigit():
        movimientos_qs = movimientos_qs.filter(proyecto_id=int(proyecto_filtro))
    if fondo_filtro:
        movimientos_qs = movimientos_qs.filter(Q(tipo_fondo=fondo_filtro) | Q(cuenta__fondo_asociado=fondo_filtro))
    if tipo_filtro in ['ingreso', 'egreso']:
        movimientos_qs = movimientos_qs.filter(tipo=tipo_filtro)
    if categoria_filtro and categoria_filtro.isdigit():
        movimientos_qs = movimientos_qs.filter(categoria_id=int(categoria_filtro))
    if busqueda:
        movimientos_qs = movimientos_qs.filter(
            Q(concepto__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(numero_comprobante__icontains=busqueda) |
            Q(proyecto__nombre__icontains=busqueda)
        )

    # Facturas
    facturas_qs = FacturaGasto.objects.filter(colegio=colegio).select_related('proyecto', 'movimiento_asociado', 'registrado_por').order_by('-fecha_emision', '-id')

    # Cálculos por Bolsas de Subvención
    saldo_sep = Decimal(cuentas.filter(Q(fondo_asociado='sep') | Q(tipo='subvencion_sep')).aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)
    saldo_pie = Decimal(cuentas.filter(Q(fondo_asociado='pie') | Q(tipo='subvencion_pie')).aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)
    saldo_general = Decimal(cuentas.filter(Q(fondo_asociado='subvencion_general') | Q(tipo='subvencion_general')).aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)
    saldo_mantenimiento = Decimal(cuentas.filter(Q(fondo_asociado='mantenimiento') | Q(tipo='fondo_mantenimiento')).aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)
    saldo_propios = Decimal(cuentas.filter(fondo_asociado='fondos_propios').aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)

    # Totales Globales
    saldo_total_disponible = Decimal(cuentas.aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)
    saldo_total_cajas_chicas = Decimal(cajas_chicas.aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)

    # Totales Proyectos
    total_presupuesto_proyectos = Decimal(proyectos.aggregate(Sum('presupuesto_asignado'))['presupuesto_asignado__sum'] or 0)
    total_gastado_proyectos = sum([p.total_gastado for p in proyectos])
    total_saldo_proyectos = total_presupuesto_proyectos - total_gastado_proyectos

    hoy = timezone.now().date()
    mes_actual = hoy.month
    anio_actual = hoy.year

    ingresos_mes = Decimal(MovimientoFinanciero.objects.filter(
        colegio=colegio, tipo='ingreso', estado='completado', fecha__year=anio_actual, fecha__month=mes_actual
    ).aggregate(Sum('monto'))['monto__sum'] or 0)

    egresos_mes = Decimal(MovimientoFinanciero.objects.filter(
        colegio=colegio, tipo='egreso', estado='completado', fecha__year=anio_actual, fecha__month=mes_actual
    ).aggregate(Sum('monto'))['monto__sum'] or 0)

    balance_mes = ingresos_mes - egresos_mes

    facturas_pendientes_monto = Decimal(facturas_qs.filter(estado_pago='pendiente').aggregate(Sum('monto_total'))['monto_total__sum'] or 0)
    facturas_pendientes_count = facturas_qs.filter(estado_pago='pendiente').count()

    # Datos para Gráfico de Flujo de Caja (Últimos 6 meses)
    import calendar
    flujo_labels = []
    flujo_ingresos = []
    flujo_egresos = []

    for i in range(5, -1, -1):
        m = mes_actual - i
        y = anio_actual
        if m <= 0:
            m += 12
            y -= 1
        nombre_mes = calendar.month_abbr[m].capitalize()
        flujo_labels.append(f"{nombre_mes} {y}")

        ing_m = MovimientoFinanciero.objects.filter(
            colegio=colegio, tipo='ingreso', estado='completado', fecha__year=y, fecha__month=m
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0')
        egr_m = MovimientoFinanciero.objects.filter(
            colegio=colegio, tipo='egreso', estado='completado', fecha__year=y, fecha__month=m
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0')

        flujo_ingresos.append(float(ing_m))
        flujo_egresos.append(float(egr_m))

    # Datos para Gráfico de Gastos por Categoría Supereduc
    gastos_por_cat = MovimientoFinanciero.objects.filter(
        colegio=colegio, tipo='egreso', estado='completado', fecha__year=anio_actual
    ).values('clasificacion_supereduc').annotate(total=Sum('monto')).order_by('-total')

    # Diccionario de nombres legibles Supereduc
    supereduc_dict = dict(ProyectoEscolar.CATEGORIA_SUPEREDUC)
    cat_labels = [supereduc_dict.get(item['clasificacion_supereduc'], item['clasificacion_supereduc']) for item in gastos_por_cat[:6]]
    cat_valores = [float(item['total']) for item in gastos_por_cat[:6]]

    # Paginación de movimientos
    paginator = Paginator(movimientos_qs, 15)
    page_number = request.GET.get('page', 1)
    movimientos_page = paginator.get_page(page_number)

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'finanzas',
        'hoy': timezone.now(),
        # Cuentas, Cajas & Proyectos
        'cuentas': cuentas,
        'cajas_chicas': cajas_chicas,
        'bolsas_subvencion': bolsas_subvencion,
        'proyectos': proyectos,
        'categorias_ingreso': categorias_ingreso,
        'categorias_egreso': categorias_egreso,
        'todas_categorias': todas_categorias,
        # Listados
        'movimientos': movimientos_page,
        'facturas': facturas_qs[:30],
        'tab_activa': tab_activa,
        # Filtros seleccionados
        'cuenta_filtro': int(cuenta_filtro) if cuenta_filtro and cuenta_filtro.isdigit() else None,
        'proyecto_filtro': int(proyecto_filtro) if proyecto_filtro and proyecto_filtro.isdigit() else None,
        'fondo_filtro': fondo_filtro,
        'tipo_filtro': tipo_filtro,
        'categoria_filtro': int(categoria_filtro) if categoria_filtro and categoria_filtro.isdigit() else None,
        'busqueda': busqueda,
        # KPIs de Bolsas y Subvenciones
        'saldo_sep': saldo_sep,
        'saldo_pie': saldo_pie,
        'saldo_general': saldo_general,
        'saldo_mantenimiento': saldo_mantenimiento,
        'saldo_propios': saldo_propios,
        'saldo_total_disponible': saldo_total_disponible,
        'saldo_total_cajas_chicas': saldo_total_cajas_chicas,
        # KPIs Proyectos
        'total_presupuesto_proyectos': total_presupuesto_proyectos,
        'total_gastado_proyectos': total_gastado_proyectos,
        'total_saldo_proyectos': total_saldo_proyectos,
        # KPIs Mes
        'ingresos_mes': ingresos_mes,
        'egresos_mes': egresos_mes,
        'balance_mes': balance_mes,
        'facturas_pendientes_monto': facturas_pendientes_monto,
        'facturas_pendientes_count': facturas_pendientes_count,
        # Gráficos
        'flujo_labels': flujo_labels,
        'flujo_ingresos': flujo_ingresos,
        'flujo_egresos': flujo_egresos,
        'cat_labels': cat_labels,
        'cat_valores': cat_valores,
        # Opciones select
        'tipos_fondo_opciones': ProyectoEscolar.TIPO_FONDO,
        'categorias_supereduc_opciones': ProyectoEscolar.CATEGORIA_SUPEREDUC,
    }
    return render(request, 'colegios/finanzas_dashboard.html', context)


@login_required
def crear_proyecto_escolar_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import ProyectoEscolar
        from decimal import Decimal

        codigo = request.POST.get('codigo', '').strip().upper()
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        tipo_fondo = request.POST.get('tipo_fondo', 'sep')
        categoria_supereduc = request.POST.get('categoria_supereduc', 'pedagogico')
        presupuesto_str = request.POST.get('presupuesto_asignado', '0').replace('.', '').replace(',', '.').replace('$', '').strip()
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_termino_str = request.POST.get('fecha_termino')

        try:
            presupuesto_asignado = Decimal(presupuesto_str)
        except Exception:
            presupuesto_asignado = Decimal('0.0')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else timezone.now().date()
        except ValueError:
            fecha_inicio = timezone.now().date()

        fecha_termino = None
        if fecha_termino_str:
            try:
                fecha_termino = datetime.strptime(fecha_termino_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        if not codigo:
            codigo = f"PRY-{tipo_fondo.upper()[:3]}-{timezone.now().strftime('%d%H')}"

        proyecto = ProyectoEscolar.objects.create(
            colegio=colegio,
            codigo=codigo,
            nombre=nombre if nombre else "Nuevo Proyecto Escolar",
            descripcion=descripcion,
            tipo_fondo=tipo_fondo,
            categoria_supereduc=categoria_supereduc,
            presupuesto_asignado=presupuesto_asignado,
            fecha_inicio=fecha_inicio,
            fecha_termino=fecha_termino,
            responsable=request.user,
            estado='en_ejecucion',
            activo=True
        )

        messages.success(request, f"¡Proyecto '{proyecto.nombre}' ({proyecto.codigo}) creado con presupuesto de ${presupuesto_asignado:,.0f}!")

    return redirect('/colegios/finanzas/?tab=proyectos')


@login_required
def editar_proyecto_escolar_view(request, proyecto_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ProyectoEscolar
    from decimal import Decimal

    proyecto = get_object_or_404(ProyectoEscolar, id=proyecto_id, colegio=colegio)

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        tipo_fondo = request.POST.get('tipo_fondo', proyecto.tipo_fondo)
        categoria_supereduc = request.POST.get('categoria_supereduc', proyecto.categoria_supereduc)
        presupuesto_str = request.POST.get('presupuesto_asignado', '0').replace('.', '').replace(',', '.').replace('$', '').strip()
        estado = request.POST.get('estado', proyecto.estado)

        try:
            proyecto.presupuesto_asignado = Decimal(presupuesto_str)
        except Exception:
            pass

        if nombre:
            proyecto.nombre = nombre
        proyecto.descripcion = descripcion
        proyecto.tipo_fondo = tipo_fondo
        proyecto.categoria_supereduc = categoria_supereduc
        proyecto.estado = estado
        proyecto.save()

        messages.success(request, f"Proyecto '{proyecto.codigo}' actualizado correctamente.")

    return redirect('/colegios/finanzas/?tab=proyectos')


@login_required
def eliminar_proyecto_escolar_view(request, proyecto_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ProyectoEscolar
    proyecto = get_object_or_404(ProyectoEscolar, id=proyecto_id, colegio=colegio)
    proyecto.activo = False
    proyecto.save()
    messages.info(request, f"Proyecto '{proyecto.nombre}' archivado.")
    return redirect('/colegios/finanzas/?tab=proyectos')


@login_required
def crear_movimiento_financiero_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "Colegio no encontrado.")
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import CuentaFinanciera, CategoriaFinanciera, MovimientoFinanciero, ProyectoEscolar
        from decimal import Decimal

        tipo = request.POST.get('tipo', 'egreso')
        cuenta_id = request.POST.get('cuenta_id')
        proyecto_id = request.POST.get('proyecto_id')
        categoria_id = request.POST.get('categoria_id')
        tipo_fondo = request.POST.get('tipo_fondo')
        clasificacion_supereduc = request.POST.get('clasificacion_supereduc', 'pedagogico')
        
        monto_str = request.POST.get('monto', '0').replace('.', '').replace(',', '.').replace('$', '').strip()
        concepto = request.POST.get('concepto', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        fecha_str = request.POST.get('fecha')
        metodo_pago = request.POST.get('metodo_pago', 'efectivo')
        numero_comprobante = request.POST.get('numero_comprobante', '').strip()
        comprobante_file = request.FILES.get('comprobante_adjunto')

        cuenta = get_object_or_404(CuentaFinanciera, id=cuenta_id, colegio=colegio)
        categoria = CategoriaFinanciera.objects.filter(id=categoria_id, colegio=colegio).first() if categoria_id else None
        proyecto = ProyectoEscolar.objects.filter(id=proyecto_id, colegio=colegio).first() if proyecto_id else None

        if not tipo_fondo:
            if proyecto:
                tipo_fondo = proyecto.tipo_fondo
            elif cuenta.fondo_asociado:
                tipo_fondo = cuenta.fondo_asociado
            else:
                tipo_fondo = 'subvencion_general'

        try:
            monto = Decimal(monto_str)
        except Exception:
            messages.error(request, "El monto ingresado no es válido.")
            return redirect('finanzas_dashboard')

        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
        except ValueError:
            fecha_obj = timezone.now().date()

        movimiento = MovimientoFinanciero.objects.create(
            colegio=colegio,
            cuenta=cuenta,
            proyecto=proyecto,
            categoria=categoria,
            tipo=tipo,
            tipo_fondo=tipo_fondo,
            clasificacion_supereduc=clasificacion_supereduc,
            monto=monto,
            concepto=concepto if concepto else f"Movimiento de {tipo.capitalize()}",
            descripcion=descripcion,
            fecha=fecha_obj,
            metodo_pago=metodo_pago,
            numero_comprobante=numero_comprobante,
            comprobante_adjunto=comprobante_file,
            registrado_por=request.user,
            estado='completado'
        )

        # Actualizar saldo de la cuenta
        if tipo == 'ingreso':
            cuenta.saldo_actual += monto
        else:
            cuenta.saldo_actual -= monto
        cuenta.save()

        signo_msg = 'Ingreso registrado (+${:,.0f})' if tipo == 'ingreso' else 'Egreso registrado (-${:,.0f})'
        messages.success(request, f"¡{signo_msg.format(monto)} en {cuenta.nombre}!")

    return redirect('/colegios/finanzas/?tab=movimientos')


@login_required
def eliminar_movimiento_financiero_view(request, movimiento_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import MovimientoFinanciero
    mov = get_object_or_404(MovimientoFinanciero, id=movimiento_id, colegio=colegio)
    
    # Revertir saldo de la cuenta
    cuenta = mov.cuenta
    if mov.tipo == 'ingreso':
        cuenta.saldo_actual -= mov.monto
    else:
        cuenta.saldo_actual += mov.monto
    cuenta.save()

    mov.delete()
    messages.info(request, "Movimiento eliminado y saldo de la cuenta ajustado.")
    return redirect('/colegios/finanzas/?tab=movimientos')


@login_required
def crear_factura_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import FacturaGasto, ProyectoEscolar
        from decimal import Decimal

        tipo_doc = request.POST.get('tipo_documento', 'factura_afecta')
        proyecto_id = request.POST.get('proyecto_id')
        tipo_fondo = request.POST.get('tipo_fondo', 'subvencion_general')
        clasificacion_supereduc = request.POST.get('clasificacion_supereduc', 'pedagogico')
        
        folio = request.POST.get('folio', '').strip()
        proveedor_nombre = request.POST.get('proveedor_nombre', '').strip()
        proveedor_rut = request.POST.get('proveedor_rut', '').strip()
        fecha_emision_str = request.POST.get('fecha_emision')
        fecha_venc_str = request.POST.get('fecha_vencimiento')
        monto_total_str = request.POST.get('monto_total', '0').replace('.', '').replace(',', '.').replace('$', '').strip()
        archivo = request.FILES.get('archivo_factura')
        observaciones = request.POST.get('observaciones', '').strip()

        proyecto = ProyectoEscolar.objects.filter(id=proyecto_id, colegio=colegio).first() if proyecto_id else None
        if proyecto and not tipo_fondo:
            tipo_fondo = proyecto.tipo_fondo

        try:
            monto_total = Decimal(monto_total_str)
        except Exception:
            messages.error(request, "Monto inválido.")
            return redirect('finanzas_dashboard')

        # Calcular neto e IVA estimado si es factura afecta
        if tipo_doc == 'factura_afecta':
            monto_neto = round(monto_total / Decimal('1.19'), 2)
            monto_iva = monto_total - monto_neto
        else:
            monto_neto = monto_total
            monto_iva = Decimal('0.0')

        try:
            fecha_emision = datetime.strptime(fecha_emision_str, '%Y-%m-%d').date() if fecha_emision_str else timezone.now().date()
        except ValueError:
            fecha_emision = timezone.now().date()

        fecha_venc = None
        if fecha_venc_str:
            try:
                fecha_venc = datetime.strptime(fecha_venc_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        FacturaGasto.objects.create(
            colegio=colegio,
            proyecto=proyecto,
            tipo_documento=tipo_doc,
            tipo_fondo=tipo_fondo,
            clasificacion_supereduc=clasificacion_supereduc,
            folio=folio if folio else f"DOC-{timezone.now().strftime('%H%M%S')}",
            proveedor_nombre=proveedor_nombre if proveedor_nombre else "Proveedor Varios",
            proveedor_rut=proveedor_rut,
            fecha_emision=fecha_emision,
            fecha_vencimiento=fecha_venc,
            monto_neto=monto_neto,
            monto_iva=monto_iva,
            monto_total=monto_total,
            estado_pago='pendiente',
            archivo_factura=archivo,
            observaciones=observaciones,
            registrado_por=request.user
        )

        messages.success(request, f"¡Documento #{folio} ({proveedor_nombre}) registrado exitosamente!")

    return redirect('/colegios/finanzas/?tab=facturas')


@login_required
def pagar_factura_view(request, factura_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import FacturaGasto, CuentaFinanciera, CategoriaFinanciera, MovimientoFinanciero
    factura = get_object_or_404(FacturaGasto, id=factura_id, colegio=colegio)

    if request.method == 'POST':
        cuenta_id = request.POST.get('cuenta_pago_id')
        metodo_pago = request.POST.get('metodo_pago', 'transferencia')

        cuenta = get_object_or_404(CuentaFinanciera, id=cuenta_id, colegio=colegio)
        categoria_egreso = CategoriaFinanciera.objects.filter(colegio=colegio, tipo='egreso').first()

        # Crear movimiento de egreso
        mov = MovimientoFinanciero.objects.create(
            colegio=colegio,
            cuenta=cuenta,
            proyecto=factura.proyecto,
            categoria=categoria_egreso,
            tipo='egreso',
            tipo_fondo=factura.tipo_fondo,
            clasificacion_supereduc=factura.clasificacion_supereduc,
            monto=factura.monto_total,
            concepto=f"Pago {factura.get_tipo_documento_display()} #{factura.folio} - {factura.proveedor_nombre}",
            fecha=timezone.now().date(),
            metodo_pago=metodo_pago,
            numero_comprobante=factura.folio,
            registrado_por=request.user,
            estado='completado'
        )

        # Descontar de cuenta
        cuenta.saldo_actual -= factura.monto_total
        cuenta.save()

        factura.estado_pago = 'pagado'
        factura.movimiento_asociado = mov
        factura.save()

        messages.success(request, f"¡Factura #{factura.folio} pagada con éxito con cargo a {cuenta.nombre}!")

    return redirect('/colegios/finanzas/?tab=facturas')


@login_required
def eliminar_factura_view(request, factura_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import FacturaGasto
    factura = get_object_or_404(FacturaGasto, id=factura_id, colegio=colegio)
    factura.delete()
    messages.info(request, "Documento eliminado.")
    return redirect('/colegios/finanzas/?tab=facturas')


@login_required
def crear_cuenta_financiera_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import CuentaFinanciera
        from decimal import Decimal

        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', 'caja_chica')
        fondo_asociado = request.POST.get('fondo_asociado', 'fondos_propios')
        banco = request.POST.get('banco', '').strip()
        numero_cuenta = request.POST.get('numero_cuenta', '').strip()
        saldo_inicial_str = request.POST.get('saldo_inicial', '0').replace('.', '').replace(',', '.').replace('$', '').strip()

        try:
            saldo_inicial = Decimal(saldo_inicial_str)
        except Exception:
            saldo_inicial = Decimal('0.0')

        CuentaFinanciera.objects.create(
            colegio=colegio,
            nombre=nombre if nombre else "Nueva Cuenta",
            tipo=tipo,
            fondo_asociado=fondo_asociado,
            banco=banco,
            numero_cuenta=numero_cuenta,
            saldo_inicial=saldo_inicial,
            saldo_actual=saldo_inicial,
            activo=True
        )

        messages.success(request, f"¡Cuenta / Caja '{nombre}' creada con éxito!")

    return redirect('/colegios/finanzas/?tab=fondos')


@login_required
def crear_categoria_financiera_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import CategoriaFinanciera
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', 'egreso')
        icono = request.POST.get('icono', 'bi-tag')
        color = request.POST.get('color', '#7C5CFC')

        if nombre:
            CategoriaFinanciera.objects.get_or_create(
                colegio=colegio,
                nombre=nombre,
                tipo=tipo,
                defaults={'icono': icono, 'color': color, 'activo': True}
            )
            messages.success(request, f"Categoría '{nombre}' creada.")

    return redirect('finanzas_dashboard')


# ==============================================================================
# INFORMES DE RENDICIÓN DE CUENTAS & REPORTES SUPEREDUC
# ==============================================================================

@login_required
def rendicion_subvencion_imprimible_view(request, tipo_fondo):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import MovimientoFinanciero, FacturaGasto, ProyectoEscolar, CuentaFinanciera
    from django.db.models import Sum
    from decimal import Decimal

    nombres_fondo = dict(ProyectoEscolar.TIPO_FONDO)
    titulo_fondo = nombres_fondo.get(tipo_fondo, f"Fondo {tipo_fondo.upper()}")

    # Egresos asociados al fondo
    egresos = MovimientoFinanciero.objects.filter(
        colegio=colegio, tipo='egreso', tipo_fondo=tipo_fondo
    ).select_related('proyecto', 'cuenta', 'categoria', 'registrado_por').order_by('fecha', 'id')

    # Facturas asociadas al fondo
    facturas = FacturaGasto.objects.filter(
        colegio=colegio, tipo_fondo=tipo_fondo
    ).select_related('proyecto', 'registrado_por').order_by('fecha_emision', 'id')

    proyectos = ProyectoEscolar.objects.filter(colegio=colegio, tipo_fondo=tipo_fondo, activo=True)
    cuentas = CuentaFinanciera.objects.filter(colegio=colegio, fondo_asociado=tipo_fondo, activo=True)

    total_presupuesto = Decimal(proyectos.aggregate(Sum('presupuesto_asignado'))['presupuesto_asignado__sum'] or 0)
    total_ejecutado = Decimal(egresos.aggregate(Sum('monto'))['monto__sum'] or 0)
    saldo_disponible = Decimal(cuentas.aggregate(Sum('saldo_actual'))['saldo_actual__sum'] or 0)

    # Desglose por Partida Supereduc
    desglose_partidas = {}
    for eg in egresos:
        partida = eg.get_clasificacion_supereduc_display() if hasattr(eg, 'get_clasificacion_supereduc_display') else eg.clasificacion_supereduc
        desglose_partidas[partida] = desglose_partidas.get(partida, Decimal('0.0')) + eg.monto

    context = {
        'colegio': colegio,
        'tipo_fondo': tipo_fondo,
        'titulo_fondo': titulo_fondo,
        'hoy': timezone.now(),
        'egresos': egresos,
        'facturas': facturas,
        'proyectos': proyectos,
        'cuentas': cuentas,
        'total_presupuesto': total_presupuesto,
        'total_ejecutado': total_ejecutado,
        'saldo_disponible': saldo_disponible,
        'desglose_partidas': desglose_partidas,
    }
    return render(request, 'colegios/finanzas/rendicion_subvencion_imprimible.html', context)


@login_required
def acta_rendicion_cajachica_imprimible_view(request, cuenta_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import CuentaFinanciera, MovimientoFinanciero
    from django.db.models import Sum
    from decimal import Decimal

    cuenta = get_object_or_404(CuentaFinanciera, id=cuenta_id, colegio=colegio)
    movimientos = MovimientoFinanciero.objects.filter(cuenta=cuenta).select_related('proyecto', 'categoria', 'registrado_por').order_by('fecha', 'id')

    total_ingresos = Decimal(movimientos.filter(tipo='ingreso').aggregate(Sum('monto'))['monto__sum'] or 0)
    total_egresos = Decimal(movimientos.filter(tipo='egreso').aggregate(Sum('monto'))['monto__sum'] or 0)

    context = {
        'colegio': colegio,
        'cuenta': cuenta,
        'hoy': timezone.now(),
        'movimientos': movimientos,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'saldo_actual': cuenta.saldo_actual,
    }
    return render(request, 'colegios/finanzas/acta_rendicion_cajachica.html', context)


@login_required
def exportar_finanzas_excel_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    from .models import CuentaFinanciera, MovimientoFinanciero, FacturaGasto, ProyectoEscolar

    # HOJA 1: RENDICIÓN SUPEREDUC & SUBVENCIONES
    ws1 = wb.create_sheet(title="Rendición Supereduc")
    ws1.views.sheetView[0].showGridLines = True
    ws1['A1'] = f"INFORME DE RENDICIÓN DE CUENTAS ESCOLAR (SUPEREDUC)"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Colegio: {colegio.nombre} | Ciudad: {colegio.ciudad_comuna} | Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = sub_font

    headers_super = ["Fecha Gasto", "Subvención / Fondo", "Proyecto PME", "Partida Supereduc", "N° Doc / Boleta", "Concepto & Detalle", "Monto Rendido ($)", "Cuenta / Caja"]
    for col_idx, h in enumerate(headers_super, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    egresos_super = MovimientoFinanciero.objects.filter(colegio=colegio, tipo='egreso').select_related('cuenta', 'proyecto', 'categoria').order_by('tipo_fondo', '-fecha')
    for row_idx, m in enumerate(egresos_super, start=5):
        row_data = [
            m.fecha.strftime('%d/%m/%Y'),
            m.get_tipo_fondo_display() if hasattr(m, 'get_tipo_fondo_display') else m.tipo_fondo,
            m.proyecto.nombre if m.proyecto else "Gastos Generales / Operación",
            m.get_clasificacion_supereduc_display() if hasattr(m, 'get_clasificacion_supereduc_display') else m.clasificacion_supereduc,
            m.numero_comprobante or "-",
            m.concepto,
            f"${m.monto:,.0f}",
            m.cuenta.nombre
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 5] else "left", vertical="center")

    # HOJA 2: PROYECTOS ESCOLARES
    ws2 = wb.create_sheet(title="Proyectos Escolares")
    ws2.views.sheetView[0].showGridLines = True
    ws2['A1'] = "CONTROL PRESUPUESTARIO DE PROYECTOS ESCOLARES & CENTROS DE COSTO"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Colegio: {colegio.nombre}"
    ws2['A2'].font = sub_font

    headers_pry = ["Código", "Nombre Proyecto", "Fondo / Subvención", "Partida", "Presupuesto ($)", "Gastado Real ($)", "Saldo Disponible ($)", "% Avance", "Estado"]
    for col_idx, h in enumerate(headers_pry, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    proyectos_list = ProyectoEscolar.objects.filter(colegio=colegio, activo=True)
    for row_idx, p in enumerate(proyectos_list, start=5):
        row_data = [
            p.codigo,
            p.nombre,
            p.get_tipo_fondo_display(),
            p.get_categoria_supereduc_display(),
            f"${p.presupuesto_asignado:,.0f}",
            f"${p.total_gastado:,.0f}",
            f"${p.saldo_disponible:,.0f}",
            f"{p.porcentaje_ejecucion}%",
            p.get_estado_display()
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 8, 9] else "left", vertical="center")

    # HOJA 3: LIBRO DE MOVIMIENTOS
    ws3 = wb.create_sheet(title="Libro Ingresos y Egresos")
    ws3.views.sheetView[0].showGridLines = True
    ws3['A1'] = "LIBRO COMPLETO DE INGRESOS Y EGRESOS"
    ws3['A1'].font = title_font

    headers_mov = ["Fecha", "Tipo", "Cuenta / Caja", "Fondo", "Proyecto", "Concepto", "Comprobante", "Monto ($)", "Registrado Por"]
    for col_idx, h in enumerate(headers_mov, start=1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    movimientos = MovimientoFinanciero.objects.filter(colegio=colegio).select_related('cuenta', 'proyecto', 'registrado_por').order_by('-fecha', '-id')
    for row_idx, m in enumerate(movimientos, start=5):
        signo = "+" if m.tipo == 'ingreso' else "-"
        row_data = [
            m.fecha.strftime('%d/%m/%Y'),
            m.get_tipo_display(),
            m.cuenta.nombre,
            m.get_tipo_fondo_display() if hasattr(m, 'get_tipo_fondo_display') else m.tipo_fondo,
            m.proyecto.codigo if m.proyecto else "-",
            m.concepto,
            m.numero_comprobante or "-",
            f"{signo}${m.monto:,.0f}",
            m.registrado_por.get_full_name() if m.registrado_por else "-"
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5, 7] else "left", vertical="center")

    # Autoajuste de columnas en todas las hojas
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value and cell.row >= 4:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Rendicion_Finanzas_{colegio.nombre.replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response

    facturas = FacturaGasto.objects.filter(colegio=colegio).order_by('-fecha_emision', '-id')
    for row_idx, f in enumerate(facturas, start=5):
        row_data = [
            f.get_tipo_documento_display(),
            f.folio,
            f.proveedor_nombre,
            f.proveedor_rut or "-",
            f.fecha_emision.strftime('%d/%m/%Y'),
            f.fecha_vencimiento.strftime('%d/%m/%Y') if f.fecha_vencimiento else "-",
            f"${f.monto_neto:,.0f}",
            f"${f.monto_iva:,.0f}",
            f"${f.monto_total:,.0f}",
            f.get_estado_pago_display()
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [2, 4, 5, 6, 10] else "left", vertical="center")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len and len(val_str) < 50:
                        max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"finanzas_{colegio.nombre.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==============================================================================
# MÓDULO DE INVENTARIO ESCOLAR, STOCK Y DIRECTORIO DE PROVEEDORES
# ==============================================================================

def inicializar_datos_inventario(colegio):
    from .models import ProveedorColegio, CategoriaInventario, ItemInventario

    # Proveedores iniciales
    p1, _ = ProveedorColegio.objects.get_or_create(
        colegio=colegio,
        nombre="Librería & Papelería Central SpA",
        defaults={
            'rut': '76.432.890-1',
            'categoria_insumos': 'Papelería, Útiles y Cartelería',
            'contacto_nombre': 'Rodrigo Valenzuela',
            'telefono': '+56991234567',
            'email': 'contacto@libreriacentral.cl',
            'direccion': 'Av. Providencia 1240, Santiago',
            'notas': 'Descuento institucional 10% en compras sobre $100.000.'
        }
    )

    p2, _ = ProveedorColegio.objects.get_or_create(
        colegio=colegio,
        nombre="TecnoEscuela & Soluciones TI",
        defaults={
            'rut': '77.890.123-4',
            'categoria_insumos': 'Tecnología, Impresión y Proyectores',
            'contacto_nombre': 'Camila Arancibia',
            'telefono': '+56987654321',
            'email': 'ventas@tecnoescuela.cl',
            'direccion': 'Calle Las Industrias 890, Santiago',
            'notas': 'Garantía técnica 12 meses en hardware y recargas de tóner.'
        }
    )

    p3, _ = ProveedorColegio.objects.get_or_create(
        colegio=colegio,
        nombre="Distribuidora Limpimax Institucional",
        defaults={
            'rut': '78.111.222-3',
            'categoria_insumos': 'Aseo, Higiene y Desinfección',
            'contacto_nombre': 'Mauricio Soto',
            'telefono': '+56955544433',
            'email': 'pedidos@limpimax.cl',
            'direccion': 'San Pablo 4500, Quinta Normal',
            'notas': 'Despacho gratuito los días martes y jueves.'
        }
    )

    # Categorías
    cat_pap, _ = CategoriaInventario.objects.get_or_create(colegio=colegio, nombre="Útiles & Papelería", defaults={'icono': 'bi-pencil-square', 'color': '#7C5CFC'})
    cat_tec, _ = CategoriaInventario.objects.get_or_create(colegio=colegio, nombre="Tecnología & Impresión", defaults={'icono': 'bi-laptop', 'color': '#3B82F6'})
    cat_ase, _ = CategoriaInventario.objects.get_or_create(colegio=colegio, nombre="Aseo & Limpieza", defaults={'icono': 'bi-droplet-half', 'color': '#10B981'})
    cat_mob, _ = CategoriaInventario.objects.get_or_create(colegio=colegio, nombre="Mobiliario & Infraestructura", defaults={'icono': 'bi-building', 'color': '#F59E0B'})

    # Items iniciales con alertas si no existen
    if not ItemInventario.objects.filter(colegio=colegio).exists():
        ItemInventario.objects.create(
            colegio=colegio,
            nombre="Resmas de Papel Carta 75g (Chamex / Report)",
            sku="PAP-CRT-01",
            categoria=cat_pap,
            proveedor_principal=p1,
            tipo="consumible",
            stock_actual=3,
            stock_minimo=12,
            unidad_medida="resmas",
            ubicacion="Bodega Principal - Estante A1",
            costo_unitario=4200.0,
            descripcion="Papel blanco multiuso para guías escolares y evaluaciones impresas."
        )

        ItemInventario.objects.create(
            colegio=colegio,
            nombre="Plumones de Pizarra Recargables (Caja 12 un. Surtidos)",
            sku="PLU-PIZ-12",
            categoria=cat_pap,
            proveedor_principal=p1,
            tipo="consumible",
            stock_actual=4,
            stock_minimo=10,
            unidad_medida="cajas",
            ubicacion="Sala de Profesores / Armario Insumos",
            costo_unitario=8900.0,
            descripcion="Plumones punta redonda para salas de clases (Azul, Negro, Rojo)."
        )

        ItemInventario.objects.create(
            colegio=colegio,
            nombre="Tóner Impresora HP LaserJet Pro 4003",
            sku="TON-HP-4003",
            categoria=cat_tec,
            proveedor_principal=p2,
            tipo="consumible",
            stock_actual=1,
            stock_minimo=3,
            unidad_medida="unidades",
            ubicacion="Oficina de Dirección",
            costo_unitario=45000.0,
            descripcion="Cartucho de tóner de alto rendimiento para fotocopiadora y secretaría."
        )

        ItemInventario.objects.create(
            colegio=colegio,
            nombre="Alcohol Gel Desinfectante 70% (Bidón 5L)",
            sku="ASE-ALC-5L",
            categoria=cat_ase,
            proveedor_principal=p3,
            tipo="consumible",
            stock_actual=8,
            stock_minimo=4,
            unidad_medida="bidones",
            ubicacion="Bodega de Auxiliares",
            costo_unitario=12500.0,
            descripcion="Para recarga de dispensadores en salas y patios."
        )

        ItemInventario.objects.create(
            colegio=colegio,
            nombre="Impresora Multifuncional Epson EcoTank L3250",
            sku="IMP-EPS-L32",
            categoria=cat_tec,
            proveedor_principal=p2,
            tipo="activo",
            stock_actual=3,
            stock_minimo=1,
            unidad_medida="unidades",
            ubicacion="Secretaría y Sala PIE",
            costo_unitario=189000.0,
            descripcion="Impresoras con sistema continuo para uso administrativo."
        )


@login_required
def inventario_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import RolPermiso, MiembroPermiso
    tiene_permiso = is_admin
    if not tiene_permiso and miembro and miembro.rol:
        tiene_permiso = RolPermiso.objects.filter(rol=miembro.rol, modulo__nombre__in=['Inventario', 'Bodega'], puede_ver=True).exists()
    if not tiene_permiso and miembro:
        tiene_permiso = MiembroPermiso.objects.filter(miembro=miembro, modulo__nombre__in=['Inventario', 'Bodega'], puede_ver=True).exists()

    if not tiene_permiso:
        messages.error(request, "Acceso restringido. Se requieren permisos de Inventario o Administración.")
        return redirect('dashboard_usuario')

    inicializar_datos_inventario(colegio)

    from .models import ItemInventario, CategoriaInventario, ProveedorColegio, MovimientoStock
    from django.db.models import Q, F, Sum, ExpressionWrapper, DecimalField

    categorias = CategoriaInventario.objects.filter(colegio=colegio, activo=True)
    proveedores = ProveedorColegio.objects.filter(colegio=colegio, activo=True)

    # Filtros
    cat_filtro = request.GET.get('categoria')
    prov_filtro = request.GET.get('proveedor')
    tipo_filtro = request.GET.get('tipo')
    estado_filtro = request.GET.get('estado')
    busqueda = request.GET.get('q', '').strip()

    items_qs = ItemInventario.objects.filter(colegio=colegio, activo=True).select_related('categoria', 'proveedor_principal')

    if cat_filtro and cat_filtro.isdigit():
        items_qs = items_qs.filter(categoria_id=int(cat_filtro))
    if prov_filtro and prov_filtro.isdigit():
        items_qs = items_qs.filter(proveedor_principal_id=int(prov_filtro))
    if tipo_filtro in ['consumible', 'activo']:
        items_qs = items_qs.filter(tipo=tipo_filtro)
    if estado_filtro == 'alerta':
        items_qs = items_qs.filter(stock_actual__lte=F('stock_minimo'), stock_actual__gt=0)
    elif estado_filtro == 'agotado':
        items_qs = items_qs.filter(stock_actual__lte=0)
    elif estado_filtro == 'optimo':
        items_qs = items_qs.filter(stock_actual__gt=F('stock_minimo'))

    if busqueda:
        items_qs = items_qs.filter(
            Q(nombre__icontains=busqueda) |
            Q(sku__icontains=busqueda) |
            Q(ubicacion__icontains=busqueda) |
            Q(proveedor_principal__nombre__icontains=busqueda)
        )

    # KPIs
    todos_items = ItemInventario.objects.filter(colegio=colegio, activo=True)
    total_articulos = todos_items.count()
    total_en_alerta = todos_items.filter(stock_actual__lte=F('stock_minimo'), stock_actual__gt=0).count()
    total_agotados = todos_items.filter(stock_actual__lte=0).count()
    total_criticos = total_en_alerta + total_agotados

    # Valor total estimado del inventario
    valor_total_inv = sum([item.stock_actual * item.costo_unitario for item in todos_items if item.stock_actual > 0])

    # Movimientos recientes de bodega
    movimientos_recientes = MovimientoStock.objects.filter(item__colegio=colegio).select_related('item', 'registrado_por').order_by('-fecha')[:15]

    # Paginación
    paginator = Paginator(items_qs.order_by('stock_actual', 'nombre'), 15)
    page_number = request.GET.get('page', 1)
    items_page = paginator.get_page(page_number)

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'inventario',
        'hoy': timezone.now(),
        # Listados
        'items': items_page,
        'categorias': categorias,
        'proveedores': proveedores,
        'movimientos_recientes': movimientos_recientes,
        # Filtros
        'cat_filtro': int(cat_filtro) if cat_filtro and cat_filtro.isdigit() else None,
        'prov_filtro': int(prov_filtro) if prov_filtro and prov_filtro.isdigit() else None,
        'tipo_filtro': tipo_filtro,
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
        # KPIs
        'total_articulos': total_articulos,
        'total_en_alerta': total_en_alerta,
        'total_agotados': total_agotados,
        'total_criticos': total_criticos,
        'valor_total_inv': valor_total_inv,
    }
    return render(request, 'colegios/inventario_dashboard.html', context)


@login_required
def crear_item_inventario_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import ItemInventario, CategoriaInventario, ProveedorColegio, MovimientoStock
        from decimal import Decimal

        nombre = request.POST.get('nombre', '').strip()
        sku = request.POST.get('sku', '').strip()
        categoria_id = request.POST.get('categoria_id')
        proveedor_id = request.POST.get('proveedor_id')
        tipo = request.POST.get('tipo', 'consumible')
        stock_actual = int(request.POST.get('stock_actual', 0) or 0)
        stock_minimo = int(request.POST.get('stock_minimo', 5) or 5)
        unidad_medida = request.POST.get('unidad_medida', 'unidades').strip()
        ubicacion = request.POST.get('ubicacion', 'Bodega Principal').strip()
        costo_str = request.POST.get('costo_unitario', '0').replace('.', '').replace(',', '.').replace('$', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        try:
            costo_unitario = Decimal(costo_str)
        except Exception:
            costo_unitario = Decimal('0.0')

        categoria = CategoriaInventario.objects.filter(id=categoria_id, colegio=colegio).first() if categoria_id else None
        proveedor = ProveedorColegio.objects.filter(id=proveedor_id, colegio=colegio).first() if proveedor_id else None

        item = ItemInventario.objects.create(
            colegio=colegio,
            nombre=nombre,
            sku=sku,
            categoria=categoria,
            proveedor_principal=proveedor,
            tipo=tipo,
            stock_actual=stock_actual,
            stock_minimo=stock_minimo,
            unidad_medida=unidad_medida,
            ubicacion=ubicacion,
            costo_unitario=costo_unitario,
            descripcion=descripcion
        )

        if stock_actual > 0:
            MovimientoStock.objects.create(
                item=item,
                tipo='entrada',
                cantidad=stock_actual,
                stock_resultante=stock_actual,
                motivo='Stock Inicial al registrar artículo',
                registrado_por=request.user
            )

        messages.success(request, f"¡Artículo '{nombre}' agregado al inventario con éxito!")

    return redirect('inventario_dashboard')


@login_required
def ajustar_stock_view(request, item_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ItemInventario, MovimientoStock
    item = get_object_or_404(ItemInventario, id=item_id, colegio=colegio)

    if request.method == 'POST':
        tipo_mov = request.POST.get('tipo_movimiento', 'entrada') # entrada / salida / ajuste
        cantidad = int(request.POST.get('cantidad', 1) or 1)
        motivo = request.POST.get('motivo', '').strip()
        entregado_a = request.POST.get('entregado_a', '').strip()

        if tipo_mov == 'entrada':
            item.stock_actual += cantidad
            mov_desc = f"+{cantidad} {item.unidad_medida} (Entrada / Compra)"
        elif tipo_mov == 'salida':
            if cantidad > item.stock_actual:
                messages.error(request, f"No puedes retirar {cantidad} {item.unidad_medida}. Solo hay {item.stock_actual} en stock.")
                return redirect('inventario_dashboard')
            item.stock_actual -= cantidad
            mov_desc = f"-{cantidad} {item.unidad_medida} (Salida / Entrega a {entregado_a or 'Personal'})"
        else: # ajuste directo
            item.stock_actual = cantidad
            mov_desc = f"Ajuste manual a {cantidad} {item.unidad_medida}"

        item.save()

        MovimientoStock.objects.create(
            item=item,
            tipo=tipo_mov,
            cantidad=cantidad,
            stock_resultante=item.stock_actual,
            motivo=motivo if motivo else mov_desc,
            entregado_a=entregado_a,
            registrado_por=request.user
        )

        messages.success(request, f"¡Stock de '{item.nombre}' actualizado a {item.stock_actual} {item.unidad_medida}!")

    return redirect('inventario_dashboard')


@login_required
def eliminar_item_inventario_view(request, item_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ItemInventario
    item = get_object_or_404(ItemInventario, id=item_id, colegio=colegio)
    item.delete()
    messages.info(request, "Artículo eliminado del inventario.")
    return redirect('inventario_dashboard')


@login_required
def proveedores_directorio_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import RolPermiso, MiembroPermiso
    tiene_permiso = is_admin
    if not tiene_permiso and miembro and miembro.rol:
        tiene_permiso = RolPermiso.objects.filter(rol=miembro.rol, modulo__nombre__iexact='Proveedores', puede_ver=True).exists()
    if not tiene_permiso and miembro:
        tiene_permiso = MiembroPermiso.objects.filter(miembro=miembro, modulo__nombre__iexact='Proveedores', puede_ver=True).exists()

    if not tiene_permiso:
        messages.error(request, "Acceso restringido. Se requieren permisos de Proveedores o Administración.")
        return redirect('dashboard_usuario')

    inicializar_datos_inventario(colegio)

    from .models import ProveedorColegio
    busqueda = request.GET.get('q', '').strip()
    proveedores_qs = ProveedorColegio.objects.filter(colegio=colegio, activo=True).prefetch_related('articulos_suministrados')

    if busqueda:
        from django.db.models import Q
        proveedores_qs = proveedores_qs.filter(
            Q(nombre__icontains=busqueda) |
            Q(categoria_insumos__icontains=busqueda) |
            Q(contacto_nombre__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(telefono__icontains=busqueda)
        )

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'proveedores',
        'hoy': timezone.now(),
        'proveedores': proveedores_qs,
        'busqueda': busqueda,
    }
    return render(request, 'colegios/proveedores_directorio.html', context)


@login_required
def crear_proveedor_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import ProveedorColegio
        nombre = request.POST.get('nombre', '').strip()
        rut = request.POST.get('rut', '').strip()
        categoria_insumos = request.POST.get('categoria_insumos', 'Papelería y Útiles').strip()
        contacto_nombre = request.POST.get('contacto_nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        sitio_web = request.POST.get('sitio_web', '').strip()
        notas = request.POST.get('notas', '').strip()

        ProveedorColegio.objects.create(
            colegio=colegio,
            nombre=nombre,
            rut=rut,
            categoria_insumos=categoria_insumos,
            contacto_nombre=contacto_nombre,
            telefono=telefono,
            email=email,
            direccion=direccion,
            sitio_web=sitio_web,
            notas=notas,
            activo=True
        )

        messages.success(request, f"¡Proveedor '{nombre}' registrado exitosamente!")

    return redirect('proveedores_directorio')


@login_required
def editar_proveedor_view(request, proveedor_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ProveedorColegio
    prov = get_object_or_404(ProveedorColegio, id=proveedor_id, colegio=colegio)

    if request.method == 'POST':
        prov.nombre = request.POST.get('nombre', prov.nombre).strip()
        prov.rut = request.POST.get('rut', prov.rut).strip()
        prov.categoria_insumos = request.POST.get('categoria_insumos', prov.categoria_insumos).strip()
        prov.contacto_nombre = request.POST.get('contacto_nombre', prov.contacto_nombre).strip()
        prov.telefono = request.POST.get('telefono', prov.telefono).strip()
        prov.email = request.POST.get('email', prov.email).strip()
        prov.direccion = request.POST.get('direccion', prov.direccion).strip()
        prov.sitio_web = request.POST.get('sitio_web', prov.sitio_web).strip()
        prov.notas = request.POST.get('notas', prov.notas).strip()
        prov.save()

        messages.success(request, f"¡Proveedor '{prov.nombre}' actualizado!")

    return redirect('proveedores_directorio')


@login_required
def eliminar_proveedor_view(request, proveedor_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import ProveedorColegio
    prov = get_object_or_404(ProveedorColegio, id=proveedor_id, colegio=colegio)
    prov.activo = False
    prov.save()
    messages.info(request, f"Proveedor '{prov.nombre}' archivado.")
    return redirect('proveedores_directorio')


@login_required
def exportar_inventario_excel_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
    alert_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    from .models import ItemInventario, ProveedorColegio, MovimientoStock

    # HOJA 1: CATÁLOGO DE INVENTARIO
    ws1 = wb.create_sheet(title="Inventario General")
    ws1.views.sheetView[0].showGridLines = True
    ws1['A1'] = f"EDUTEKA - INVENTARIO GENERAL DEL COLEGIO"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Colegio: {colegio.nombre} | Emisión: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = sub_font

    headers_inv = ["Código / SKU", "Artículo", "Categoría", "Tipo", "Stock Actual", "Stock Mínimo", "Unidad", "Estado", "Ubicación", "Costo Unit. ($)", "Proveedor Principal"]
    for col_idx, h in enumerate(headers_inv, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    items = ItemInventario.objects.filter(colegio=colegio, activo=True).select_related('categoria', 'proveedor_principal').order_by('stock_actual', 'nombre')
    for row_idx, it in enumerate(items, start=5):
        estado_label = "Agotado (Crítico)" if it.stock_actual <= 0 else ("Bajo Stock (Alerta)" if it.stock_actual <= it.stock_minimo else "Óptimo")
        row_data = [
            it.sku or "-",
            it.nombre,
            it.categoria.nombre if it.categoria else "Sin Categoría",
            it.get_tipo_display(),
            it.stock_actual,
            it.stock_minimo,
            it.unidad_medida,
            estado_label,
            it.ubicacion,
            f"${it.costo_unitario:,.0f}",
            it.proveedor_principal.nombre if it.proveedor_principal else "-"
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 6, 7, 8] else "left", vertical="center")
            if it.en_alerta and col_idx == 8:
                cell.fill = alert_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="B91C1C")

    # HOJA 2: DIRECTORIO DE PROVEEDORES
    ws2 = wb.create_sheet(title="Proveedores")
    ws2.views.sheetView[0].showGridLines = True
    ws2['A1'] = "DIRECTORIO DE PROVEEDORES INSTITUCIONALES"
    ws2['A1'].font = title_font
    ws2['A2'] = f"Colegio: {colegio.nombre}"
    ws2['A2'].font = sub_font

    headers_prov = ["Razón Social / Proveedor", "RUT", "Rubro / Categoría", "Contacto Comercial", "Teléfono / WhatsApp", "Correo Electrónico", "Dirección", "Notas"]
    for col_idx, h in enumerate(headers_prov, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    proveedores = ProveedorColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    for row_idx, p in enumerate(proveedores, start=5):
        row_data = [
            p.nombre,
            p.rut or "-",
            p.categoria_insumos,
            p.contacto_nombre or "-",
            p.telefono or "-",
            p.email or "-",
            p.direccion or "-",
            p.notas or "-"
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_idx in [2, 5] else "left", vertical="center")

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len and len(val_str) < 50:
                        max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventario_{colegio.nombre.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ==============================================================================
# MÓDULO DE TALLERES EXTRACURRICULARES (ACLES) & ASISTENCIA MULTICURSO
# ==============================================================================

def inicializar_datos_talleres(colegio):
    from .models import TallerExtracurricular, InscripcionTaller, Estudiante, SesionAsistenciaTaller, DetalleAsistenciaTaller

    if not TallerExtracurricular.objects.filter(colegio=colegio).exists():
        admin_docente = colegio.administrador

        t1 = TallerExtracurricular.objects.create(
            colegio=colegio,
            nombre="Taller de Robótica & Programación Escolar",
            categoria="cientifico",
            docente_encargado=admin_docente,
            dias_horario="Martes y Jueves 15:30 - 17:00",
            lugar="Laboratorio de Informática / Sala STEM",
            cupo_maximo=20,
            descripcion="Desarrollo de proyectos con Arduino, sensores y lógica de programación en bloques para competencias escolares.",
            icono="bi-robot",
            color="#3B82F6"
        )

        t2 = TallerExtracurricular.objects.create(
            colegio=colegio,
            nombre="Selección de Fútbol Masculino & Femenino",
            categoria="deportivo",
            monitor_externo="Prof. Gabriel Retamal (DT Deportivo)",
            dias_horario="Lunes y Miércoles 16:00 - 17:45",
            lugar="Cancha Principal de Césped",
            cupo_maximo=30,
            descripcion="Entrenamiento formativo, táctico y participación en ligas y torneos intercomunales.",
            icono="bi-trophy-fill",
            color="#10B981"
        )

        t3 = TallerExtracurricular.objects.create(
            colegio=colegio,
            nombre="Taller de Teatro & Expresión Corporal",
            categoria="artistico",
            docente_encargado=admin_docente,
            dias_horario="Viernes 15:00 - 17:00",
            lugar="Auditorio / Sala de Artes",
            cupo_maximo=25,
            descripcion="Expresión escénica, improvisación, modulación de voz y montaje de la obra de fin de año.",
            icono="bi-palette-fill",
            color="#EC4899"
        )

        t4 = TallerExtracurricular.objects.create(
            colegio=colegio,
            nombre="Club de Ajedrez & Pensamiento Estratégico",
            categoria="academico",
            docente_encargado=admin_docente,
            dias_horario="Miércoles 15:30 - 17:00",
            lugar="Biblioteca Central",
            cupo_maximo=20,
            descripcion="Tácticas de apertura, medio juego, finales y torneos suizos internos.",
            icono="bi-grid-3x3-gap-fill",
            color="#F59E0B"
        )

        # Inscribir algunos estudiantes de muestra si existen
        estudiantes = list(Estudiante.objects.filter(colegio=colegio, activo=True)[:15])
        if estudiantes:
            for est in estudiantes[:8]:
                InscripcionTaller.objects.get_or_create(taller=t1, estudiante=est)
            for est in estudiantes[4:14]:
                InscripcionTaller.objects.get_or_create(taller=t2, estudiante=est)
            for est in estudiantes[2:10]:
                InscripcionTaller.objects.get_or_create(taller=t3, estudiante=est)


@login_required
def talleres_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import TallerExtracurricular, InscripcionTaller, SesionAsistenciaTaller, CursoColegio
    from django.db.models import Q, Count

    categoria_filtro = request.GET.get('categoria', '').strip()
    busqueda = request.GET.get('q', '').strip()

    talleres_qs = TallerExtracurricular.objects.filter(colegio=colegio, activo=True).annotate(
        num_inscritos=Count('inscripciones', filter=Q(inscripciones__activo=True)),
        num_sesiones=Count('sesiones_asistencia')
    ).select_related('docente_encargado')

    if categoria_filtro:
        talleres_qs = talleres_qs.filter(categoria=categoria_filtro)

    if busqueda:
        talleres_qs = talleres_qs.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(lugar__icontains=busqueda) |
            Q(monitor_externo__icontains=busqueda) |
            Q(docente_encargado__first_name__icontains=busqueda) |
            Q(docente_encargado__last_name__icontains=busqueda)
        )

    # KPIs
    todos_talleres = TallerExtracurricular.objects.filter(colegio=colegio, activo=True)
    total_talleres = todos_talleres.count()
    total_inscripciones = InscripcionTaller.objects.filter(taller__colegio=colegio, activo=True).count()
    total_sesiones = SesionAsistenciaTaller.objects.filter(taller__colegio=colegio).count()
    
    # Docentes/Personal disponibles para ser monitores
    docentes_disponibles = User.objects.filter(
        Q(membresias_colegio__colegio=colegio, membresias_colegio__activo=True) |
        Q(colegios_administrados=colegio)
    ).distinct()

    categorias_choices = TallerExtracurricular.CATEGORIA_CHOICES

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'talleres',
        'hoy': timezone.now(),
        'talleres': talleres_qs.order_by('nombre'),
        'total_talleres': total_talleres,
        'total_inscripciones': total_inscripciones,
        'total_sesiones': total_sesiones,
        'categoria_filtro': categoria_filtro,
        'busqueda': busqueda,
        'categorias_choices': categorias_choices,
        'docentes_disponibles': docentes_disponibles,
    }
    return render(request, 'colegios/talleres_dashboard.html', context)


@login_required
def crear_taller_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import TallerExtracurricular

        nombre = request.POST.get('nombre', '').strip()
        categoria = request.POST.get('categoria', 'deportivo')
        docente_id = request.POST.get('docente_encargado')
        monitor_externo = request.POST.get('monitor_externo', '').strip()
        dias_horario = request.POST.get('dias_horario', '').strip()
        lugar = request.POST.get('lugar', 'Gimnasio / Sala Multiuso').strip()
        cupo_str = request.POST.get('cupo_maximo', '30')
        descripcion = request.POST.get('descripcion', '').strip()
        icono = request.POST.get('icono', 'bi-star-fill')
        color = request.POST.get('color', '#7C5CFC')

        try:
            cupo_maximo = int(cupo_str) if cupo_str else 30
        except Exception:
            cupo_maximo = 30

        docente = User.objects.filter(id=docente_id).first() if docente_id else None

        taller = TallerExtracurricular.objects.create(
            colegio=colegio,
            nombre=nombre,
            categoria=categoria,
            docente_encargado=docente,
            monitor_externo=monitor_externo,
            dias_horario=dias_horario,
            lugar=lugar,
            cupo_maximo=cupo_maximo,
            descripcion=descripcion,
            icono=icono,
            color=color,
            activo=True
        )

        messages.success(request, f"¡Taller '{nombre}' creado exitosamente!")
        return redirect('detalle_taller', taller_id=taller.id)

    return redirect('talleres_dashboard')


@login_required
def detalle_taller_view(request, taller_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import TallerExtracurricular, InscripcionTaller, SesionAsistenciaTaller, DetalleAsistenciaTaller, Estudiante, CursoColegio

    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    inscripciones = InscripcionTaller.objects.filter(taller=taller, activo=True).select_related('estudiante', 'estudiante__seccion', 'estudiante__seccion__curso').order_by('estudiante__nombre_completo')
    sesiones = SesionAsistenciaTaller.objects.filter(taller=taller).prefetch_related('detalles').order_by('-fecha', '-id')

    # Estudiantes ya inscritos IDs
    inscritos_ids = inscripciones.values_list('estudiante_id', flat=True)

    # Estudiantes disponibles de todos los cursos para inscribir
    cursos = CursoColegio.objects.filter(colegio=colegio, activo=True).prefetch_related('secciones')
    
    # Filtro del selector de alumnos
    curso_filtro = request.GET.get('curso_filtro')
    estudiantes_disponibles_qs = Estudiante.objects.filter(colegio=colegio, activo=True).exclude(id__in=inscritos_ids).select_related('seccion', 'seccion__curso')

    if curso_filtro and curso_filtro.isdigit():
        estudiantes_disponibles_qs = estudiantes_disponibles_qs.filter(seccion__curso_id=int(curso_filtro))

    estudiantes_disponibles = estudiantes_disponibles_qs.order_by('seccion__curso__orden', 'nombre_completo')[:100]

    # Calcular estadísticas de asistencia de los inscritos
    total_sesiones_count = sesiones.count()
    inscripciones_con_stats = []
    for insc in inscripciones:
        if total_sesiones_count > 0:
            asistidas = DetalleAsistenciaTaller.objects.filter(
                sesion__taller=taller,
                estudiante=insc.estudiante,
                estado__in=['presente', 'tarde', 'justificado']
            ).count()
            porc = round((asistidas / total_sesiones_count) * 100, 1)
        else:
            porc = 100.0
        inscripciones_con_stats.append({
            'inscripcion': insc,
            'estudiante': insc.estudiante,
            'porcentaje_asistencia': porc
        })

    # Docentes disponibles para edición
    docentes_disponibles = User.objects.filter(
        Q(membresias_colegio__colegio=colegio, membresias_colegio__activo=True) |
        Q(colegios_administrados=colegio)
    ).distinct()


    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'talleres',
        'hoy': timezone.now(),
        'taller': taller,
        'inscripciones_con_stats': inscripciones_con_stats,
        'total_inscritos': len(inscripciones),
        'sesiones': sesiones,
        'total_sesiones': total_sesiones_count,
        'cursos': cursos,
        'curso_filtro': int(curso_filtro) if curso_filtro and curso_filtro.isdigit() else None,
        'estudiantes_disponibles': estudiantes_disponibles,
        'docentes_disponibles': docentes_disponibles,
        'categorias_choices': TallerExtracurricular.CATEGORIA_CHOICES,
    }
    return render(request, 'colegios/detalle_taller.html', context)


@login_required
def inscribir_estudiante_taller_view(request, taller_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import TallerExtracurricular, InscripcionTaller, Estudiante

    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)

    if request.method == 'POST':
        estudiantes_ids = request.POST.getlist('estudiantes_ids')
        estudiante_id = request.POST.get('estudiante_id')
        
        if estudiante_id and not estudiantes_ids:
            estudiantes_ids = [estudiante_id]

        agregados = 0
        for est_id in estudiantes_ids:
            if est_id and est_id.isdigit():
                est = Estudiante.objects.filter(id=int(est_id), colegio=colegio).first()
                if est:
                    _, created = InscripcionTaller.objects.get_or_create(
                        taller=taller,
                        estudiante=est,
                        defaults={'activo': True}
                    )
                    if created:
                        agregados += 1

        if agregados > 0:
            messages.success(request, f"¡Se inscribieron {agregados} alumno(s) en {taller.nombre}!")
        else:
            messages.info(request, "Los alumnos seleccionados ya estaban inscritos.")

    return redirect('detalle_taller', taller_id=taller.id)


@login_required
def desinscribir_estudiante_taller_view(request, taller_id, inscripcion_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import TallerExtracurricular, InscripcionTaller

    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    inscripcion = get_object_or_404(InscripcionTaller, id=inscripcion_id, taller=taller)
    
    nombre_est = inscripcion.estudiante.nombre_completo
    inscripcion.delete()
    messages.info(request, f"{nombre_est} fue retirado(a) del taller.")

    return redirect('detalle_taller', taller_id=taller.id)


@login_required
def tomar_asistencia_taller_view(request, taller_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import TallerExtracurricular, InscripcionTaller, SesionAsistenciaTaller, DetalleAsistenciaTaller, Estudiante

    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    inscripciones = InscripcionTaller.objects.filter(taller=taller, activo=True).select_related('estudiante', 'estudiante__seccion', 'estudiante__seccion__curso').order_by('estudiante__nombre_completo')

    if request.method == 'POST':
        fecha_str = request.POST.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        contenido = request.POST.get('contenido_actividad', '').strip()

        import datetime
        try:
            fecha_sesion = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except Exception:
            fecha_sesion = timezone.now().date()

        sesion, _ = SesionAsistenciaTaller.objects.get_or_create(
            taller=taller,
            fecha=fecha_sesion,
            defaults={
                'contenido_actividad': contenido,
                'registrado_por': request.user
            }
        )
        sesion.contenido_actividad = contenido
        sesion.registrado_por = request.user
        sesion.save()

        # Guardar estado de cada estudiante inscrito
        for insc in inscripciones:
            est = insc.estudiante
            estado = request.POST.get(f'estado_{est.id}', 'presente')
            obs = request.POST.get(f'obs_{est.id}', '').strip()

            DetalleAsistenciaTaller.objects.update_or_create(
                sesion=sesion,
                estudiante=est,
                defaults={
                    'estado': estado,
                    'observacion': obs
                }
            )

        messages.success(request, f"¡Asistencia de {taller.nombre} guardada con éxito para el {fecha_sesion.strftime('%d/%m/%Y')}!")
        return redirect('detalle_taller', taller_id=taller.id)

    # GET
    fecha_hoy = timezone.now().strftime('%Y-%m-%d')
    # Verificar si ya hay una sesión guardada hoy
    sesion_hoy = SesionAsistenciaTaller.objects.filter(taller=taller, fecha=timezone.now().date()).first()
    detalles_dict = {}
    if sesion_hoy:
        for d in sesion_hoy.detalles.all():
            detalles_dict[d.estudiante_id] = {'estado': d.estado, 'observacion': d.observacion}

    for insc in inscripciones:
        det = detalles_dict.get(insc.estudiante_id, {})
        insc.estudiante.estado_actual = det.get('estado', 'presente')
        insc.estudiante.obs_actual = det.get('observacion', '')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'talleres',
        'hoy': timezone.now(),
        'taller': taller,
        'inscripciones': inscripciones,
        'fecha_hoy': fecha_hoy,
        'sesion_hoy': sesion_hoy,
    }
    return render(request, 'colegios/tomar_asistencia_taller.html', context)



@login_required
def detalle_sesion_taller_view(request, taller_id, sesion_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import TallerExtracurricular, SesionAsistenciaTaller

    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    sesion = get_object_or_404(SesionAsistenciaTaller, id=sesion_id, taller=taller)
    detalles = sesion.detalles.select_related('estudiante', 'estudiante__seccion', 'estudiante__seccion__curso').order_by('estudiante__nombre_completo')

    context = {
        'colegio': colegio,
        'active_page': 'talleres',
        'taller': taller,
        'sesion': sesion,
        'detalles': detalles,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/detalle_sesion_taller.html', context)


@login_required
def eliminar_sesion_taller_view(request, taller_id, sesion_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import TallerExtracurricular, SesionAsistenciaTaller
    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    sesion = get_object_or_404(SesionAsistenciaTaller, id=sesion_id, taller=taller)
    sesion.delete()
    messages.info(request, "Sesión de asistencia eliminada.")
    return redirect('detalle_taller', taller_id=taller.id)


@login_required
def editar_taller_view(request, taller_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import TallerExtracurricular
    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)

    if request.method == 'POST':
        taller.nombre = request.POST.get('nombre', taller.nombre).strip()
        taller.categoria = request.POST.get('categoria', taller.categoria)
        docente_id = request.POST.get('docente_encargado')
        taller.docente_encargado = User.objects.filter(id=docente_id).first() if docente_id else None
        taller.monitor_externo = request.POST.get('monitor_externo', '').strip()
        taller.dias_horario = request.POST.get('dias_horario', taller.dias_horario).strip()
        taller.lugar = request.POST.get('lugar', taller.lugar).strip()
        cupo_str = request.POST.get('cupo_maximo')
        try:
            taller.cupo_maximo = int(cupo_str) if cupo_str else None
        except Exception:
            pass
        taller.descripcion = request.POST.get('descripcion', taller.descripcion).strip()
        taller.color = request.POST.get('color', taller.color)
        taller.save()

        messages.success(request, f"¡Taller '{taller.nombre}' actualizado!")

    return redirect('detalle_taller', taller_id=taller.id)


@login_required
def eliminar_taller_view(request, taller_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import TallerExtracurricular
    taller = get_object_or_404(TallerExtracurricular, id=taller_id, colegio=colegio)
    taller.activo = False
    taller.save()
    messages.info(request, f"Taller '{taller.nombre}' archivado.")
    return redirect('talleres_dashboard')


# ==============================================================================
# VISTAS DE EVALUACIÓN Y GESTIÓN SIMCE / DIA
# ==============================================================================

def inicializar_datos_simce_demo(colegio, user):
    """Inicializa datos históricos y ensayos de prueba para que el colegio tenga métricas reales al comenzar."""
    from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE, PuntajeHistoricoSIMCE, CursoColegio, Estudiante
    import random

    # 1. Histórico Oficial Agencia de Calidad (si no existen)
    if not PuntajeHistoricoSIMCE.objects.filter(colegio=colegio).exists():
        historicos = [
            # 2024
            {'anio': 2024, 'nivel_escolar': '4° Básico', 'asignatura': 'Lectura', 'puntaje_colegio': 272, 'puntaje_gse': 258, 'puntaje_nacional': 253, 'insuf': 15.0, 'elem': 40.0, 'adec': 45.0},
            {'anio': 2024, 'nivel_escolar': '4° Básico', 'asignatura': 'Matemática', 'puntaje_colegio': 265, 'puntaje_gse': 252, 'puntaje_nacional': 249, 'insuf': 22.0, 'elem': 45.0, 'adec': 33.0},
            {'anio': 2024, 'nivel_escolar': 'II° Medio', 'asignatura': 'Lectura', 'puntaje_colegio': 260, 'puntaje_gse': 250, 'puntaje_nacional': 248, 'insuf': 25.0, 'elem': 45.0, 'adec': 30.0},
            {'anio': 2024, 'nivel_escolar': 'II° Medio', 'asignatura': 'Matemática', 'puntaje_colegio': 258, 'puntaje_gse': 245, 'puntaje_nacional': 244, 'insuf': 30.0, 'elem': 42.0, 'adec': 28.0},
            # 2023
            {'anio': 2023, 'nivel_escolar': '4° Básico', 'asignatura': 'Lectura', 'puntaje_colegio': 264, 'puntaje_gse': 255, 'puntaje_nacional': 250, 'insuf': 20.0, 'elem': 45.0, 'adec': 35.0},
            {'anio': 2023, 'nivel_escolar': '4° Básico', 'asignatura': 'Matemática', 'puntaje_colegio': 254, 'puntaje_gse': 248, 'puntaje_nacional': 245, 'insuf': 28.0, 'elem': 44.0, 'adec': 28.0},
            # 2022
            {'anio': 2022, 'nivel_escolar': '4° Básico', 'asignatura': 'Lectura', 'puntaje_colegio': 258, 'puntaje_gse': 251, 'puntaje_nacional': 247, 'insuf': 26.0, 'elem': 46.0, 'adec': 28.0},
            {'anio': 2022, 'nivel_escolar': '4° Básico', 'asignatura': 'Matemática', 'puntaje_colegio': 248, 'puntaje_gse': 244, 'puntaje_nacional': 241, 'insuf': 34.0, 'elem': 42.0, 'adec': 24.0},
        ]
        for h in historicos:
            PuntajeHistoricoSIMCE.objects.create(
                colegio=colegio,
                anio=h['anio'],
                nivel_escolar=h['nivel_escolar'],
                asignatura=h['asignatura'],
                puntaje_colegio=h['puntaje_colegio'],
                puntaje_gse=h['puntaje_gse'],
                puntaje_nacional=h['puntaje_nacional'],
                nivel_insuficiente_pct=h['insuf'],
                nivel_elemental_pct=h['elem'],
                nivel_adecuado_pct=h['adec']
            )

    # 2. Ensayos Internos de ejemplo
    if not EnsayoSIMCE.objects.filter(colegio=colegio).exists():
        cursos = CursoColegio.objects.filter(colegio=colegio, activo=True)
        if cursos.exists():
            curso_sample = cursos.first()
            ensayos_seed = [
                {'titulo': f"1° Ensayo SIMCE Matemática - {curso_sample.nombre}", 'asig': 'matematica', 'preg': 35},
                {'titulo': f"1° Ensayo SIMCE Lectura - {curso_sample.nombre}", 'asig': 'lectura', 'preg': 30},
            ]
            estudiantes = Estudiante.objects.filter(seccion__curso=curso_sample, activo=True)

            for item in ensayos_seed:
                ens = EnsayoSIMCE.objects.create(
                    colegio=colegio,
                    titulo=item['titulo'],
                    asignatura=item['asig'],
                    curso=curso_sample,
                    fecha=timezone.now().date(),
                    total_preguntas=item['preg'],
                    creado_por=user,
                    descripcion="Simulacro institucional de diagnóstico para nivelación y refuerzo UTP."
                )
                for est in estudiantes:
                    correctas = random.randint(int(item['preg'] * 0.4), item['preg'])
                    res = ResultadoEnsayoSIMCE(
                        ensayo=ens,
                        estudiante=est,
                        respuestas_correctas=correctas
                    )
                    res.calcular_puntaje_y_nivel()
                    res.save()


@login_required
def simce_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE, PuntajeHistoricoSIMCE, CursoColegio

    # Filtros
    asignatura_filtro = request.GET.get('asignatura', '')
    curso_filtro = request.GET.get('curso', '')

    ensayos_qs = EnsayoSIMCE.objects.filter(colegio=colegio).select_related('curso', 'creado_por').prefetch_related('resultados')
    if asignatura_filtro:
        ensayos_qs = ensayos_qs.filter(asignatura=asignatura_filtro)
    if curso_filtro:
        ensayos_qs = ensayos_qs.filter(curso_id=curso_filtro)

    # Métricas Globales de Ensayos
    todos_ensayos = EnsayoSIMCE.objects.filter(colegio=colegio).prefetch_related('resultados')
    total_ensayos = todos_ensayos.count()
    
    todos_resultados = ResultadoEnsayoSIMCE.objects.filter(ensayo__colegio=colegio)
    total_evaluados = todos_resultados.count()
    
    if todos_resultados.exists():
        promedio_general_simce = round(sum(r.puntaje_simce for r in todos_resultados) / total_evaluados)
        promedio_logro_general = round(sum(r.porcentaje_logro for r in todos_resultados) / total_evaluados, 1)
        total_insuficientes = todos_resultados.filter(nivel_aprendizaje='insuficiente').count()
        total_elementales = todos_resultados.filter(nivel_aprendizaje='elemental').count()
        total_adecuados = todos_resultados.filter(nivel_aprendizaje='adecuado').count()
        pct_insuficiente = round((total_insuficientes / total_evaluados) * 100, 1)
        pct_elemental = round((total_elementales / total_evaluados) * 100, 1)
        pct_adecuado = round((total_adecuados / total_evaluados) * 100, 1)
    else:
        promedio_general_simce = 0
        promedio_logro_general = 0.0
        total_insuficientes = total_elementales = total_adecuados = 0
        pct_insuficiente = pct_elemental = pct_adecuado = 0.0

    # Histórico Oficial
    historicos = PuntajeHistoricoSIMCE.objects.filter(colegio=colegio).order_by('-anio', 'nivel_escolar', 'asignatura')
    
    # Preparar datos JSON para gráficos
    # 1. Gráfico Histórico Oficial
    hist_recientes = list(historicos[:8])
    hist_recientes.reverse()
    chart_labels = [f"{h.anio} {h.nivel_escolar} ({h.asignatura[:3]})" for h in hist_recientes]
    chart_colegio = [h.puntaje_colegio for h in hist_recientes]
    chart_gse = [h.puntaje_gse for h in hist_recientes]
    chart_nacional = [h.puntaje_nacional for h in hist_recientes]

    # Cursos disponibles para crear nuevo ensayo
    cursos_colegio = CursoColegio.objects.filter(colegio=colegio, activo=True).order_by('nivel', 'nombre')

    # Estudiantes en Nivel Insuficiente (Riesgo Pedagógico)
    alumnos_riesgo = ResultadoEnsayoSIMCE.objects.filter(
        ensayo__colegio=colegio,
        nivel_aprendizaje='insuficiente'
    ).select_related('estudiante', 'estudiante__seccion', 'ensayo').order_by('puntaje_simce')[:20]

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'simce',
        'hoy': timezone.now(),
        'ensayos': ensayos_qs,
        'total_ensayos': total_ensayos,
        'total_evaluados': total_evaluados,
        'promedio_general_simce': promedio_general_simce,
        'promedio_logro_general': promedio_logro_general,
        'pct_insuficiente': pct_insuficiente,
        'pct_elemental': pct_elemental,
        'pct_adecuado': pct_adecuado,
        'total_insuficientes': total_insuficientes,
        'total_elementales': total_elementales,
        'total_adecuados': total_adecuados,
        'historicos': historicos,
        'cursos_colegio': cursos_colegio,
        'asignaturas_choices': EnsayoSIMCE.ASIGNATURAS_SIMCE,
        'asignatura_filtro': asignatura_filtro,
        'curso_filtro': curso_filtro,
        'chart_labels': json.dumps(chart_labels),
        'chart_colegio': json.dumps(chart_colegio),
        'chart_gse': json.dumps(chart_gse),
        'chart_nacional': json.dumps(chart_nacional),
        'alumnos_riesgo': alumnos_riesgo,
    }
    return render(request, 'colegios/simce_dashboard.html', context)


@login_required
def crear_ensayo_simce_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE, CursoColegio, Estudiante

        titulo = request.POST.get('titulo', '').strip()
        asignatura = request.POST.get('asignatura', 'matematica')
        curso_id = request.POST.get('curso')
        fecha_str = request.POST.get('fecha')
        total_preguntas = int(request.POST.get('total_preguntas', 35))
        descripcion = request.POST.get('descripcion', '').strip()

        curso = get_object_or_404(CursoColegio, id=curso_id, colegio=colegio)
        
        try:
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
        except ValueError:
            fecha_obj = timezone.now().date()

        ensayo = EnsayoSIMCE.objects.create(
            colegio=colegio,
            titulo=titulo,
            asignatura=asignatura,
            curso=curso,
            fecha=fecha_obj,
            total_preguntas=total_preguntas,
            descripcion=descripcion,
            creado_por=request.user
        )

        # Generar automáticamente filas vacías de resultados para los estudiantes del curso
        estudiantes = Estudiante.objects.filter(seccion__curso=curso, activo=True).order_by('nombre_completo')
        for est in estudiantes:
            ResultadoEnsayoSIMCE.objects.get_or_create(
                ensayo=ensayo,
                estudiante=est,
                defaults={
                    'respuestas_correctas': 0,
                    'porcentaje_logro': 0.0,
                    'puntaje_simce': 150,
                    'nivel_aprendizaje': 'insuficiente'
                }
            )

        messages.success(request, f"¡Ensayo '{titulo}' creado con éxito! Ya puedes registrar los resultados.")
        return redirect('detalle_ensayo_simce', ensayo_id=ensayo.id)

    return redirect('simce_dashboard')


@login_required
def detalle_ensayo_simce_view(request, ensayo_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    periodo = ConfiguracionAcademica.objects.filter(colegio=colegio).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE, Estudiante

    ensayo = get_object_or_404(EnsayoSIMCE, id=ensayo_id, colegio=colegio)
    
    # Asegurar que todos los alumnos del curso tengan fila en el ensayo
    alumnos_curso = Estudiante.objects.filter(seccion__curso=ensayo.curso, activo=True)
    for al in alumnos_curso:
        ResultadoEnsayoSIMCE.objects.get_or_create(
            ensayo=ensayo,
            estudiante=al,
            defaults={'respuestas_correctas': 0, 'puntaje_simce': 150, 'nivel_aprendizaje': 'insuficiente'}
        )

    resultados = ResultadoEnsayoSIMCE.objects.filter(ensayo=ensayo).select_related('estudiante', 'estudiante__seccion').order_by('estudiante__nombre_completo')
    
    context = {
        'colegio': colegio,
        'miembro': miembro,
        'periodo': periodo,
        'is_admin': is_admin,
        'active_page': 'simce',
        'hoy': timezone.now(),
        'ensayo': ensayo,
        'resultados': resultados,
        'desglose': ensayo.desglose_niveles,
    }
    return render(request, 'colegios/detalle_ensayo_simce.html', context)


@login_required
def guardar_resultados_ensayo_view(request, ensayo_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE

    ensayo = get_object_or_404(EnsayoSIMCE, id=ensayo_id, colegio=colegio)

    if request.method == 'POST':
        resultados = ResultadoEnsayoSIMCE.objects.filter(ensayo=ensayo).select_related('estudiante')
        for res in resultados:
            correctas_str = request.POST.get(f'correctas_{res.id}', '0')
            obs = request.POST.get(f'obs_{res.id}', '').strip()

            try:
                correctas = int(correctas_str)
            except ValueError:
                correctas = 0

            # Validar que no supere el total de preguntas
            if correctas > ensayo.total_preguntas:
                correctas = ensayo.total_preguntas
            if correctas < 0:
                correctas = 0

            res.respuestas_correctas = correctas
            res.observaciones = obs
            res.calcular_puntaje_y_nivel()
            res.save()

        messages.success(request, f"¡Resultados del ensayo '{ensayo.titulo}' guardados y recalculados exitosamente!")
        return redirect('detalle_ensayo_simce', ensayo_id=ensayo.id)

    return redirect('detalle_ensayo_simce', ensayo_id=ensayo.id)


@login_required
def eliminar_ensayo_simce_view(request, ensayo_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import EnsayoSIMCE
    ensayo = get_object_or_404(EnsayoSIMCE, id=ensayo_id, colegio=colegio)
    titulo = ensayo.titulo
    ensayo.delete()
    messages.info(request, f"Ensayo '{titulo}' eliminado correctamente.")
    return redirect('simce_dashboard')


@login_required
def crear_historico_simce_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import PuntajeHistoricoSIMCE

        anio = int(request.POST.get('anio', 2024))
        nivel_escolar = request.POST.get('nivel_escolar', '4° Básico').strip()
        asignatura = request.POST.get('asignatura', 'Lectura').strip()
        puntaje_colegio = int(request.POST.get('puntaje_colegio', 250))
        puntaje_gse = int(request.POST.get('puntaje_gse', 250))
        puntaje_nacional = int(request.POST.get('puntaje_nacional', 250))
        
        try:
            insuf = float(request.POST.get('nivel_insuficiente_pct', 20.0))
            elem = float(request.POST.get('nivel_elemental_pct', 45.0))
            adec = float(request.POST.get('nivel_adecuado_pct', 35.0))
        except ValueError:
            insuf, elem, adec = 20.0, 45.0, 35.0

        PuntajeHistoricoSIMCE.objects.create(
            colegio=colegio,
            anio=anio,
            nivel_escolar=nivel_escolar,
            asignatura=asignatura,
            puntaje_colegio=puntaje_colegio,
            puntaje_gse=puntaje_gse,
            puntaje_nacional=puntaje_nacional,
            nivel_insuficiente_pct=insuf,
            nivel_elemental_pct=elem,
            nivel_adecuado_pct=adec
        )

        messages.success(request, f"¡Puntaje histórico SIMCE {anio} ({nivel_escolar} - {asignatura}) agregado!")
    return redirect('simce_dashboard')


@login_required
def eliminar_historico_simce_view(request, historico_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import PuntajeHistoricoSIMCE
    hist = get_object_or_404(PuntajeHistoricoSIMCE, id=historico_id, colegio=colegio)
    hist.delete()
    messages.info(request, "Registro histórico SIMCE eliminado.")
    return redirect('simce_dashboard')


@login_required
def exportar_simce_excel_view(request, ensayo_id=None):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .models import EnsayoSIMCE, ResultadoEnsayoSIMCE, PuntajeHistoricoSIMCE

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    title_font = Font(name="Arial", size=14, bold=True, color="151A35")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    if ensayo_id:
        # Exportar Ensayo Específico
        ensayo = get_object_or_404(EnsayoSIMCE, id=ensayo_id, colegio=colegio)
        ws = wb.active
        ws.title = "Resultados Ensayo"

        ws.append([f"INFORME DE ENSAYO SIMCE: {ensayo.titulo}"])
        ws.append([f"Establecimiento: {colegio.nombre} | Curso: {ensayo.curso.nombre} | Fecha: {ensayo.fecha.strftime('%d/%m/%Y')}"])
        ws.append([f"Promedio SIMCE: {ensayo.promedio_puntaje} pts | Promedio Logro: {ensayo.promedio_logro}% | Preguntas: {ensayo.total_preguntas}"])
        ws.append([])

        ws.cell(row=1, column=1).font = title_font

        headers = ["N°", "Estudiante", "Sección", "RUT", "Resp. Correctas", "% Logro", "Puntaje SIMCE", "Nivel de Aprendizaje", "Observaciones"]
        ws.append(headers)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=5, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        resultados = ResultadoEnsayoSIMCE.objects.filter(ensayo=ensayo).select_related('estudiante', 'estudiante__seccion').order_by('estudiante__nombre_completo')
        
        for idx, r in enumerate(resultados, 1):
            sec_nombre = r.estudiante.seccion.nombre if r.estudiante.seccion else "-"
            ws.append([
                idx,
                r.estudiante.nombre_completo,
                sec_nombre,
                r.estudiante.rut or "-",
                r.respuestas_correctas,
                f"{r.porcentaje_logro}%",
                r.puntaje_simce,
                r.get_nivel_aprendizaje_display(),
                r.observaciones or ""
            ])

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        filename = f"SIMCE_{ensayo.curso.nombre}_{ensayo.asignatura}_{ensayo.fecha.strftime('%Y%m%d')}.xlsx"
    else:
        # Exportar Resumen General del Colegio
        ws1 = wb.active
        ws1.title = "Ensayos Internos"
        ws1.append([f"RESUMEN GENERAL DE ENSAYOS SIMCE - {colegio.nombre}"])
        ws1.append([f"Fecha de Exportación: {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
        ws1.append([])

        headers1 = ["ID", "Título del Ensayo", "Asignatura", "Curso", "Fecha", "N° Preguntas", "Alumnos Evaluados", "Promedio SIMCE", "% Logro", "% Insuficiente", "% Elemental", "% Adecuado"]
        ws1.append(headers1)
        for col_num in range(1, len(headers1) + 1):
            cell = ws1.cell(row=4, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        ensayos = EnsayoSIMCE.objects.filter(colegio=colegio).select_related('curso').prefetch_related('resultados')
        for e in ensayos:
            d = e.desglose_niveles
            ws1.append([
                e.id,
                e.titulo,
                e.get_asignatura_display(),
                e.curso.nombre,
                e.fecha.strftime('%d/%m/%Y'),
                e.total_preguntas,
                e.total_rendidos,
                e.promedio_puntaje,
                f"{e.promedio_logro}%",
                f"{d['pct_insuficiente']}%",
                f"{d['pct_elemental']}%",
                f"{d['pct_adecuado']}%"
            ])

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Hoja 2: Histórico Oficial
        ws2 = wb.create_sheet(title="Histórico Oficial Agencia")
        ws2.append([f"HISTORIAL OFICIAL DE RESULTADOS SIMCE - {colegio.nombre}"])
        ws2.append([])

        headers2 = ["Año", "Nivel Escolar", "Asignatura", "Puntaje Colegio", "Promedio GSE", "Promedio Nacional", "Diferencia vs GSE", "% Insuficiente", "% Elemental", "% Adecuado"]
        ws2.append(headers2)
        for col_num in range(1, len(headers2) + 1):
            cell = ws2.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        historicos = PuntajeHistoricoSIMCE.objects.filter(colegio=colegio)
        for h in historicos:
            diff = h.puntaje_colegio - h.puntaje_gse
            sign = "+" if diff > 0 else ""
            ws2.append([
                h.anio,
                h.nivel_escolar,
                h.asignatura,
                h.puntaje_colegio,
                h.puntaje_gse,
                h.puntaje_nacional,
                f"{sign}{diff} pts",
                f"{h.nivel_insuficiente_pct}%",
                f"{h.nivel_elemental_pct}%",
                f"{h.nivel_adecuado_pct}%"
            ])

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        filename = f"SIMCE_Informe_General_{colegio.nombre.replace(' ', '_')}.xlsx"

    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ════════════════════════════════════════════════════════════════════════════════
# 📦 VISTAS DE ADQUISICIONES, 3 COTIZACIONES Y ÓRDENES DE COMPRA (SUPEREDUC)
# ════════════════════════════════════════════════════════════════════════════════

@login_required
def procesos_compra_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No estás asociado a ningún colegio.")
        return redirect('solicitar_acceso')

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.is_superuser
        or request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director'])
    )

    from .models import RolPermiso, MiembroPermiso
    tiene_permiso = is_admin
    if not tiene_permiso and miembro and miembro.rol:
        tiene_permiso = RolPermiso.objects.filter(rol=miembro.rol, modulo__nombre__in=['Proveedores', 'Inventario', 'Finanzas'], puede_ver=True).exists()
    if not tiene_permiso and miembro:
        tiene_permiso = MiembroPermiso.objects.filter(miembro=miembro, modulo__nombre__in=['Proveedores', 'Inventario', 'Finanzas'], puede_ver=True).exists()

    if not tiene_permiso:
        messages.error(request, "Acceso restringido. Se requieren permisos de Adquisiciones, Proveedores o Administración.")
        return redirect('dashboard_usuario')

    from .models import ProcesoCompra, ItemProcesoCompra, CotizacionProveedor, ItemInventario, ProveedorColegio
    from django.db.models import Count, Sum, Q

    # Filtros
    estado_filtro = request.GET.get('estado', '')
    fondo_filtro = request.GET.get('fondo', '')
    busqueda = request.GET.get('q', '').strip()

    qs = ProcesoCompra.objects.filter(colegio=colegio).prefetch_related('items', 'cotizaciones__proveedor').select_related('cotizacion_ganadora', 'creado_por')

    if estado_filtro:
        qs = qs.filter(estado=estado_filtro)
    if fondo_filtro:
        qs = qs.filter(tipo_fondo=fondo_filtro)
    if busqueda:
        qs = qs.filter(
            Q(codigo__icontains=busqueda) |
            Q(titulo__icontains=busqueda) |
            Q(centro_costo__icontains=busqueda) |
            Q(numero_orden_compra__icontains=busqueda)
        )

    # Métricas
    todos_procesos = ProcesoCompra.objects.filter(colegio=colegio)
    total_procesos_count = todos_procesos.count()
    en_cotizacion_count = todos_procesos.filter(estado__in=['en_cotizacion', 'evaluacion']).count()
    adjudicados_count = todos_procesos.filter(estado__in=['adjudicado', 'orden_compra_emitida', 'recepcionado']).count()
    
    # Procesos con 3 o más cotizaciones
    con_3_cotizaciones_count = 0
    for p in todos_procesos:
        if p.cumple_normativa_3_cotizaciones:
            con_3_cotizaciones_count += 1

    # Total invertido en compras adjudicadas
    monto_total_adjudicado = CotizacionProveedor.objects.filter(
        proceso__colegio=colegio,
        es_ganadora=True
    ).aggregate(tot=Sum('monto_total'))['tot'] or 0

    items_inventario = ItemInventario.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    proveedores_colegio = ProveedorColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'procesos': qs,
        'total_procesos_count': total_procesos_count,
        'en_cotizacion_count': en_cotizacion_count,
        'adjudicados_count': adjudicados_count,
        'con_3_cotizaciones_count': con_3_cotizaciones_count,
        'monto_total_adjudicado': monto_total_adjudicado,
        'estado_filtro': estado_filtro,
        'fondo_filtro': fondo_filtro,
        'busqueda': busqueda,
        'items_inventario': items_inventario,
        'proveedores_colegio': proveedores_colegio,
        'active_page': 'adquisiciones',
    }
    return render(request, 'colegios/procesos_compra_dashboard.html', context)


@login_required
def crear_proceso_compra_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        messages.error(request, "No tienes un establecimiento asociado.")
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import ProcesoCompra, ItemProcesoCompra, ItemInventario
        
        titulo = request.POST.get('titulo', '').strip()
        tipo_fondo = request.POST.get('tipo_fondo', 'subvencion_general')
        centro_costo = request.POST.get('centro_costo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        fecha_limite = request.POST.get('fecha_limite_cotizacion') or None

        if not titulo:
            messages.error(request, "El título del requerimiento de compra es obligatorio.")
            return redirect('procesos_compra_dashboard')

        # Generar código correlativo
        anio = timezone.now().year
        total_este_anio = ProcesoCompra.objects.filter(colegio=colegio, fecha_creacion__year=anio).count() + 1
        codigo = f"REQ-{anio}-{total_este_anio:03d}"

        proceso = ProcesoCompra.objects.create(
            colegio=colegio,
            codigo=codigo,
            titulo=titulo,
            tipo_fondo=tipo_fondo,
            centro_costo=centro_costo,
            descripcion=descripcion,
            fecha_limite_cotizacion=fecha_limite,
            creado_por=request.user,
            estado='en_cotizacion'
        )

        # Procesar ítems múltiples
        descripciones = request.POST.getlist('item_descripcion[]')
        cantidades = request.POST.getlist('item_cantidad[]')
        unidades = request.POST.getlist('item_unidad[]')
        inventario_ids = request.POST.getlist('item_inventario_id[]')
        especificaciones = request.POST.getlist('item_especificaciones[]')

        for i, desc in enumerate(descripciones):
            desc_clean = desc.strip()
            if desc_clean:
                cant = 1
                try:
                    cant = max(1, int(cantidades[i])) if i < len(cantidades) else 1
                except (ValueError, IndexError):
                    cant = 1
                
                unid = unidades[i].strip() if i < len(unidades) and unidades[i].strip() else 'unidades'
                spec = especificaciones[i].strip() if i < len(especificaciones) else ''
                inv_obj = None
                if i < len(inventario_ids) and inventario_ids[i].isdigit():
                    inv_obj = ItemInventario.objects.filter(colegio=colegio, id=int(inventario_ids[i])).first()

                ItemProcesoCompra.objects.create(
                    proceso=proceso,
                    item_inventario=inv_obj,
                    descripcion=desc_clean,
                    cantidad=cant,
                    unidad_medida=unid,
                    especificaciones_tecnicas=spec
                )

        messages.success(request, f"¡Requerimiento {proceso.codigo} creado con éxito! Ya puedes solicitar y registrar las 3 cotizaciones.")
        return redirect('detalle_proceso_compra', proceso_id=proceso.id)

    return redirect('procesos_compra_dashboard')


@login_required
def detalle_proceso_compra_view(request, proceso_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import ProcesoCompra, CotizacionProveedor, ProveedorColegio
    proceso = get_object_or_404(ProcesoCompra, id=proceso_id, colegio=colegio)
    
    cotizaciones = proceso.cotizaciones.select_related('proveedor', 'registrado_por').order_by('monto_total', 'plazo_entrega_dias')
    items = proceso.items.all()
    proveedores_disponibles = ProveedorColegio.objects.filter(colegio=colegio, activo=True).order_by('nombre')

    # Identificar la mejor oferta económica y la más rápida para el cuadro comparativo
    mejor_precio_id = None
    mas_rapida_id = None
    if cotizaciones.exists():
        mejor_precio_id = cotizaciones.first().id
        mas_rapida_id = cotizaciones.order_by('plazo_entrega_dias').first().id

    context = {
        'colegio': colegio,
        'proceso': proceso,
        'items': items,
        'cotizaciones': cotizaciones,
        'proveedores_disponibles': proveedores_disponibles,
        'mejor_precio_id': mejor_precio_id,
        'mas_rapida_id': mas_rapida_id,
        'active_page': 'adquisiciones',
    }
    return render(request, 'colegios/detalle_proceso_compra.html', context)


@login_required
def registrar_cotizacion_view(request, proceso_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import ProcesoCompra, CotizacionProveedor, ProveedorColegio
    from decimal import Decimal

    proceso = get_object_or_404(ProcesoCompra, id=proceso_id, colegio=colegio)

    if request.method == 'POST':
        proveedor_id = request.POST.get('proveedor_id')
        num_cot = request.POST.get('numero_cotizacion_proveedor', '').strip()
        fecha_cot = request.POST.get('fecha_cotizacion') or timezone.now().date()
        validez = request.POST.get('validez_dias', '30')
        monto_total_str = request.POST.get('monto_total', '0').replace('.', '').replace(',', '.').strip()
        monto_neto_str = request.POST.get('monto_neto', '0').replace('.', '').replace(',', '.').strip()
        plazo_dias = request.POST.get('plazo_entrega_dias', '5')
        condiciones = request.POST.get('condiciones_pago', 'Transferencia 30 días').strip()
        incluye_despacho = request.POST.get('incluye_despacho') == 'on' or request.POST.get('incluye_despacho') == 'true'
        observaciones = request.POST.get('observaciones', '').strip()
        archivo = request.FILES.get('archivo_adjunto')

        proveedor = get_object_or_404(ProveedorColegio, id=proveedor_id, colegio=colegio)

        try:
            monto_total = Decimal(monto_total_str) if monto_total_str else Decimal('0')
            monto_neto = Decimal(monto_neto_str) if monto_neto_str else Decimal('0')
            validez_dias = int(validez) if validez.isdigit() else 30
            plazo_entrega = int(plazo_dias) if plazo_dias.isdigit() else 5
        except Exception:
            monto_total = Decimal('0')
            monto_neto = Decimal('0')
            validez_dias = 30
            plazo_entrega = 5

        # Si solo ingresó monto total, calcular neto e IVA
        if monto_total > 0 and monto_neto == 0:
            monto_neto = round(monto_total / Decimal('1.19'), 2)
            iva = monto_total - monto_neto
        elif monto_neto > 0 and monto_total == 0:
            iva = round(monto_neto * Decimal('0.19'), 2)
            monto_total = monto_neto + iva
        else:
            iva = monto_total - monto_neto if monto_total > monto_neto else Decimal('0')

        cot = CotizacionProveedor.objects.create(
            proceso=proceso,
            proveedor=proveedor,
            numero_cotizacion_proveedor=num_cot,
            fecha_cotizacion=fecha_cot,
            validez_dias=validez_dias,
            monto_neto=monto_neto,
            iva=iva,
            monto_total=monto_total,
            plazo_entrega_dias=plazo_entrega,
            condiciones_pago=condiciones,
            incluye_despacho=incluye_despacho,
            observaciones=observaciones,
            archivo_adjunto=archivo,
            registrado_por=request.user
        )

        # Actualizar estado si ya tiene cotizaciones
        if proceso.estado == 'en_cotizacion' and proceso.cotizaciones.count() >= 3:
            proceso.estado = 'evaluacion'
            proceso.save()

        messages.success(request, f"¡Cotización de '{proveedor.nombre}' por ${monto_total:,.0f} registrada con éxito!")
        return redirect('detalle_proceso_compra', proceso_id=proceso.id)

    return redirect('detalle_proceso_compra', proceso_id=proceso.id)


@login_required
def adjudicar_proceso_compra_view(request, proceso_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import ProcesoCompra, CotizacionProveedor
    proceso = get_object_or_404(ProcesoCompra, id=proceso_id, colegio=colegio)

    if request.method == 'POST':
        cotizacion_id = request.POST.get('cotizacion_id')
        justificacion = request.POST.get('justificacion', '').strip()

        if not cotizacion_id:
            messages.error(request, "Debes seleccionar la cotización ganadora.")
            return redirect('detalle_proceso_compra', proceso_id=proceso.id)

        if not justificacion:
            messages.error(request, "La justificación legal de la compra es obligatoria para rendiciones normativas.")
            return redirect('detalle_proceso_compra', proceso_id=proceso.id)

        cotizacion = get_object_or_404(CotizacionProveedor, id=cotizacion_id, proceso=proceso)

        # Marcar cotizaciones
        proceso.cotizaciones.update(es_ganadora=False)
        cotizacion.es_ganadora = True
        cotizacion.save()

        # Generar número de Orden de Compra correlativa
        anio = timezone.now().year
        total_oc_anio = ProcesoCompra.objects.filter(colegio=colegio, fecha_emision_oc__year=anio).count() + 1
        numero_oc = f"OC-{anio}-{total_oc_anio:03d}"

        proceso.cotizacion_ganadora = cotizacion
        proceso.justificacion_adjudicacion = justificacion
        proceso.fecha_adjudicacion = timezone.now()
        proceso.adjudicado_por = request.user
        proceso.numero_orden_compra = numero_oc
        proceso.fecha_emision_oc = timezone.now()
        proceso.estado = 'orden_compra_emitida'
        proceso.save()

        messages.success(request, f"¡Proceso adjudicado a '{cotizacion.proveedor.nombre}'! Se ha emitido la Orden de Compra {numero_oc}.")
        return redirect('detalle_proceso_compra', proceso_id=proceso.id)

    return redirect('detalle_proceso_compra', proceso_id=proceso.id)


@login_required
def orden_compra_imprimible_view(request, proceso_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import ProcesoCompra
    proceso = get_object_or_404(ProcesoCompra, id=proceso_id, colegio=colegio)

    if not proceso.cotizacion_ganadora:
        messages.warning(request, "Este proceso aún no ha sido adjudicado a un proveedor.")
        return redirect('detalle_proceso_compra', proceso_id=proceso.id)

    cotizacion = proceso.cotizacion_ganadora
    proveedor = cotizacion.proveedor
    items = proceso.items.all()

    context = {
        'colegio': colegio,
        'proceso': proceso,
        'cotizacion': cotizacion,
        'proveedor': proveedor,
        'items': items,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/orden_compra_imprimible.html', context)


@login_required
def recepcionar_compra_inventario_view(request, proceso_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('solicitar_acceso')

    from .models import ProcesoCompra, ItemInventario, CategoriaInventario, MovimientoStock

    proceso = get_object_or_404(ProcesoCompra, id=proceso_id, colegio=colegio)

    if request.method == 'POST':
        numero_guia = request.POST.get('numero_guia_factura', '').strip()
        observaciones = request.POST.get('observaciones_recepcion', '').strip()

        # Ingresar o aumentar stock de cada ítem en el inventario escolar
        cat_default = CategoriaInventario.objects.filter(colegio=colegio).first()

        for item in proceso.items.all():
            item_inv = item.item_inventario
            if not item_inv:
                # Si no estaba enlazado a un producto existente, crearlo o buscar por nombre
                item_inv = ItemInventario.objects.filter(colegio=colegio, nombre__iexact=item.descripcion).first()
                if not item_inv:
                    item_inv = ItemInventario.objects.create(
                        colegio=colegio,
                        categoria=cat_default,
                        proveedor_principal=proceso.cotizacion_ganadora.proveedor if proceso.cotizacion_ganadora else None,
                        nombre=item.descripcion,
                        unidad_medida=item.unidad_medida,
                        stock_actual=0,
                        stock_minimo=1,
                        descripcion=f"Ingresado automáticamente desde compra {proceso.codigo}"
                    )
                item.item_inventario = item_inv
                item.save()

            # Aumentar stock
            stock_anterior = item_inv.stock_actual
            item_inv.stock_actual += item.cantidad
            item_inv.save()

            # Registrar movimiento de stock
            MovimientoStock.objects.create(
                item=item_inv,
                tipo='entrada',
                cantidad=item.cantidad,
                stock_resultante=item_inv.stock_actual,
                motivo=f"Recepción conforme de Orden de Compra {proceso.numero_orden_compra or proceso.codigo}. Doc: {numero_guia}",
                registrado_por=request.user
            )

        proceso.estado = 'recepcionado'
        proceso.fecha_recepcion = timezone.now()
        proceso.recepcionado_por = request.user
        proceso.numero_guia_factura = numero_guia
        proceso.observaciones_recepcion = observaciones
        proceso.save()

        messages.success(request, f"¡Productos recepcionados con éxito! El stock del inventario ha sido actualizado automáticamente.")
        return redirect('detalle_proceso_compra', proceso_id=proceso.id)

    return redirect('detalle_proceso_compra', proceso_id=proceso.id)


# ==============================================================================
# VISTAS DE HORARIO ESCOLAR & HORARIO DOCENTE
# ==============================================================================

@login_required
def horario_docente_imprimible_view(request, docente_id=None):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from django.contrib.auth.models import User
    from .horario_utils import obtener_datos_horario_docente
    from solicitudes.models import MiembroColegio

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if docente_id and is_admin:
        docente = get_object_or_404(User, id=docente_id)
    else:
        # Los profesores regulares solo pueden ver y descargar su propio horario
        docente = request.user

    horario_data = obtener_datos_horario_docente(colegio, docente)

    context = {
        'colegio': colegio,
        'docente': docente,
        'horario_data': horario_data,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/horario_docente_imprimible.html', context)


@login_required
def guardar_horario_clase_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if not is_admin:
        messages.error(request, "No tienes permisos para modificar el horario escolar.")
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import BloqueHorario, HorarioClase, SeccionCurso, Asignatura
        from django.contrib.auth.models import User

        horario_id = request.POST.get('horario_id')
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')
        docente_id = request.POST.get('docente_id')
        bloque_id = request.POST.get('bloque_id')
        dia_semana = request.POST.get('dia_semana')
        sala = request.POST.get('sala', '').strip()
        color = request.POST.get('color', '#7C5CFC')

        seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
        asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)
        bloque = get_object_or_404(BloqueHorario, id=bloque_id, colegio=colegio)
        docente = User.objects.filter(id=docente_id).first() if docente_id else None

        if not docente and asignatura.docente:
            docente = asignatura.docente

        if horario_id and horario_id.isdigit():
            horario = get_object_or_404(HorarioClase, id=int(horario_id), colegio=colegio)
            horario.seccion = seccion
            horario.asignatura = asignatura
            horario.docente = docente
            horario.bloque = bloque
            horario.dia_semana = int(dia_semana)
            horario.sala = sala
            horario.color = color
            horario.save()
            messages.success(request, f"¡Clase actualizada en el horario con éxito!")
        else:
            horario, created = HorarioClase.objects.update_or_create(
                colegio=colegio,
                seccion=seccion,
                bloque=bloque,
                dia_semana=int(dia_semana),
                defaults={
                    'asignatura': asignatura,
                    'docente': docente,
                    'sala': sala,
                    'color': color,
                    'activo': True
                }
            )
            messages.success(request, f"¡Clase de {asignatura.nombre} ({seccion.nombre}) asignada al {bloque.nombre}!")

    next_url = request.POST.get('next') or 'dashboard_usuario'
    return redirect(next_url)


@login_required
def eliminar_horario_clase_view(request, horario_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if not is_admin:
        messages.error(request, "No tienes permisos para eliminar clases del horario escolar.")
        return redirect('dashboard_usuario')

    from .models import HorarioClase
    horario = get_object_or_404(HorarioClase, id=horario_id, colegio=colegio)
    horario.delete()
    messages.info(request, "Bloque de clase eliminado del horario.")

    next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard_usuario'
    return redirect(next_url)


# ==============================================================================
# VISTAS: LECCIONARIO DIGITAL & PLANIFICACIÓN CURRICULAR (MINEDUC COMPLIANT)
# ==============================================================================

@login_required
def leccionario_hub_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import (
        RegistroLeccionario, PlanificacionCurricular, SeccionCurso, 
        Asignatura, BloqueHorario, HorarioClase
    )
    from django.db.models import Q
    import datetime

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )
    is_profesor = (miembro and miembro.rol and miembro.rol.nombre == 'Profesor')

    # Filtros
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        try:
            fecha_filtro = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_filtro = timezone.now().date()
    else:
        fecha_filtro = timezone.now().date()

    seccion_id = request.GET.get('seccion_id')
    asignatura_id = request.GET.get('asignatura_id')
    docente_id = request.GET.get('docente_id')

    leccionarios_qs = RegistroLeccionario.objects.filter(colegio=colegio).select_related(
        'seccion', 'asignatura', 'docente', 'bloque'
    )

    if fecha_filtro:
        leccionarios_qs = leccionarios_qs.filter(fecha=fecha_filtro)
    if seccion_id and seccion_id.isdigit():
        leccionarios_qs = leccionarios_qs.filter(seccion_id=int(seccion_id))
    if asignatura_id and asignatura_id.isdigit():
        leccionarios_qs = leccionarios_qs.filter(asignatura_id=int(asignatura_id))
    if docente_id and docente_id.isdigit() and is_admin:
        leccionarios_qs = leccionarios_qs.filter(docente_id=int(docente_id))
    elif not is_admin:
        # Los profesores ven sus propios leccionarios
        leccionarios_qs = leccionarios_qs.filter(docente=request.user)

    # Clases del día según el horario para verificar qué leccionarios faltan firmar hoy
    dia_semana_num = fecha_filtro.isoweekday()
    clases_programadas = HorarioClase.objects.filter(
        colegio=colegio,
        dia_semana=dia_semana_num,
        activo=True
    ).select_related('seccion', 'asignatura', 'docente', 'bloque')

    if not is_admin:
        clases_programadas = clases_programadas.filter(docente=request.user)

    # Combinar clases programadas con el estado de firma
    clases_con_leccionario = []
    firmados_count = 0
    for c in clases_programadas:
        lecc = RegistroLeccionario.objects.filter(
            colegio=colegio,
            seccion=c.seccion,
            asignatura=c.asignatura,
            bloque=c.bloque,
            fecha=fecha_filtro
        ).first()
        esta_firmado = bool(lecc and lecc.firmado)
        if esta_firmado:
            firmados_count += 1
        clases_con_leccionario.append({
            'clase': c,
            'leccionario': lecc,
            'firmado': esta_firmado
        })

    total_clases_dia = len(clases_programadas)
    pct_cumplimiento = int((firmados_count / total_clases_dia * 100)) if total_clases_dia > 0 else 100

    secciones_colegio = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    asignaturas_colegio = Asignatura.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    bloques_colegio = BloqueHorario.objects.filter(colegio=colegio, activo=True).order_by('orden')
    docentes_colegio = User.objects.filter(
        Q(asignaturas_dictadas__colegio=colegio) | Q(solicitudes_acceso__colegio=colegio, solicitudes_acceso__estado='aprobada')
    ).distinct().order_by('first_name', 'last_name')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'is_profesor': is_profesor,
        'fecha_filtro': fecha_filtro,
        'seccion_id_filtro': int(seccion_id) if seccion_id and seccion_id.isdigit() else None,
        'asignatura_id_filtro': int(asignatura_id) if asignatura_id and asignatura_id.isdigit() else None,
        'docente_id_filtro': int(docente_id) if docente_id and docente_id.isdigit() else None,
        'leccionarios': leccionarios_qs,
        'clases_con_leccionario': clases_con_leccionario,
        'total_clases_dia': total_clases_dia,
        'firmados_count': firmados_count,
        'pct_cumplimiento': pct_cumplimiento,
        'secciones_colegio': secciones_colegio,
        'asignaturas_colegio': asignaturas_colegio,
        'bloques_colegio': bloques_colegio,
        'docentes_colegio': docentes_colegio,
        'active_page': 'leccionario',
    }
    return render(request, 'colegios/leccionario_hub.html', context)


@login_required
def guardar_firma_leccionario_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Colegio no encontrado.'})
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import RegistroLeccionario, SeccionCurso, Asignatura, BloqueHorario, HorarioClase
        import hashlib
        import datetime

        leccionario_id = request.POST.get('leccionario_id')
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')
        bloque_id = request.POST.get('bloque_id')
        horario_clase_id = request.POST.get('horario_clase_id')
        fecha_str = request.POST.get('fecha')
        
        oa_codigo = request.POST.get('oa_codigo', '').strip()
        contenido_tratado = request.POST.get('contenido_tratado', '').strip()
        actividad_tipo = request.POST.get('actividad_tipo', 'catedra')
        observaciones = request.POST.get('observaciones', '').strip()

        if not contenido_tratado:
            messages.error(request, "Debes ingresar la descripción pedagógica del contenido tratado.")
            return redirect(request.POST.get('next') or 'leccionario_hub')

        try:
            fecha_valida = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
        except ValueError:
            fecha_valida = timezone.now().date()

        seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
        asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)
        bloque = BloqueHorario.objects.filter(id=bloque_id, colegio=colegio).first() if bloque_id else None
        horario_clase = HorarioClase.objects.filter(id=horario_clase_id, colegio=colegio).first() if horario_clase_id else None

        # Generar Hash Criptográfico de Seguridad SHA-256
        timestamp_now = timezone.now()
        raw_hash_data = f"EDUTEKA_LECCIONARIO_{colegio.id}_{request.user.id}_{seccion.id}_{asignatura.id}_{fecha_valida}_{timestamp_now.isoformat()}"
        hash_firma = hashlib.sha256(raw_hash_data.encode('utf-8')).hexdigest()

        if leccionario_id and leccionario_id.isdigit():
            lecc = get_object_or_404(RegistroLeccionario, id=int(leccionario_id), colegio=colegio)
            lecc.oa_codigo = oa_codigo
            lecc.contenido_tratado = contenido_tratado
            lecc.actividad_tipo = actividad_tipo
            lecc.observaciones = observaciones
            lecc.firmado = True
            lecc.fecha_firma = timestamp_now
            lecc.hash_firma = hash_firma
            lecc.save()
        else:
            lecc, created = RegistroLeccionario.objects.update_or_create(
                colegio=colegio,
                seccion=seccion,
                asignatura=asignatura,
                bloque=bloque,
                fecha=fecha_valida,
                defaults={
                    'horario_clase': horario_clase,
                    'docente': request.user,
                    'oa_codigo': oa_codigo,
                    'contenido_tratado': contenido_tratado,
                    'actividad_tipo': actividad_tipo,
                    'observaciones': observaciones,
                    'firmado': True,
                    'fecha_firma': timestamp_now,
                    'hash_firma': hash_firma,
                }
            )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'¡Leccionario firmado exitosamente para {asignatura.nombre} ({seccion.nombre})!',
                'leccionario_id': lecc.id,
                'hash_firma': hash_firma[:12] + '...'
            })

        messages.success(request, f"¡Leccionario firmado exitosamente con sello digital para {asignatura.nombre} ({seccion.nombre})!")

    next_url = request.POST.get('next') or 'leccionario_hub'
    return redirect(next_url)


@login_required
def planificaciones_hub_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import PlanificacionCurricular, SeccionCurso, Asignatura
    from django.db.models import Q
    import datetime

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if request.method == 'POST':
        plan_id = request.POST.get('plan_id')
        seccion_id = request.POST.get('seccion_id')
        asignatura_id = request.POST.get('asignatura_id')
        titulo_unidad = request.POST.get('titulo_unidad', '').strip()
        semestre = request.POST.get('semestre', 1)
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_termino_str = request.POST.get('fecha_termino')
        oas_curriculares = request.POST.get('oas_curriculares', '').strip()
        estrategias_metodologicas = request.POST.get('estrategias_metodologicas', '').strip()
        evaluacion_descripcion = request.POST.get('evaluacion_descripcion', '').strip()

        seccion = get_object_or_404(SeccionCurso, id=seccion_id, curso__colegio=colegio)
        asignatura = get_object_or_404(Asignatura, id=asignatura_id, colegio=colegio)

        try:
            f_ini = datetime.datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            f_fin = datetime.datetime.strptime(fecha_termino_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            f_ini = timezone.now().date()
            f_fin = timezone.now().date() + datetime.timedelta(days=30)

        if plan_id and plan_id.isdigit():
            plan = get_object_or_404(PlanificacionCurricular, id=int(plan_id), colegio=colegio)
            plan.seccion = seccion
            plan.asignatura = asignatura
            plan.titulo_unidad = titulo_unidad
            plan.semestre = int(semestre)
            plan.fecha_inicio = f_ini
            plan.fecha_termino = f_fin
            plan.oas_curriculares = oas_curriculares
            plan.estrategias_metodologicas = estrategias_metodologicas
            plan.evaluacion_descripcion = evaluacion_descripcion
            if 'enviar_a_utp' in request.POST:
                plan.estado = 'enviada'
            plan.save()
            messages.success(request, f"¡Planificación '{titulo_unidad}' actualizada con éxito!")
        else:
            estado_inicial = 'enviada' if 'enviar_a_utp' in request.POST else 'borrador'
            PlanificacionCurricular.objects.create(
                colegio=colegio,
                seccion=seccion,
                asignatura=asignatura,
                docente=request.user,
                titulo_unidad=titulo_unidad,
                semestre=int(semestre),
                fecha_inicio=f_ini,
                fecha_termino=f_fin,
                oas_curriculares=oas_curriculares,
                estrategias_metodologicas=estrategias_metodologicas,
                evaluacion_descripcion=evaluacion_descripcion,
                estado=estado_inicial
            )
            messages.success(request, f"¡Planificación '{titulo_unidad}' creada con éxito!")
        return redirect('planificaciones_hub')

    planificaciones_qs = PlanificacionCurricular.objects.filter(colegio=colegio).select_related(
        'seccion', 'asignatura', 'docente', 'revisado_por_utp'
    )
    if not is_admin:
        planificaciones_qs = planificaciones_qs.filter(docente=request.user)

    secciones_colegio = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    asignaturas_colegio = Asignatura.objects.filter(colegio=colegio, activo=True).order_by('nombre')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'planificaciones': planificaciones_qs,
        'secciones_colegio': secciones_colegio,
        'asignaturas_colegio': asignaturas_colegio,
        'active_page': 'planificaciones',
    }
    return render(request, 'colegios/planificaciones_hub.html', context)


@login_required
def cambiar_estado_planificacion_view(request, plan_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import PlanificacionCurricular

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if not is_admin:
        messages.error(request, "Solo el equipo de UTP o Dirección puede auditar y aprobar planificaciones.")
        return redirect('planificaciones_hub')

    plan = get_object_or_404(PlanificacionCurricular, id=plan_id, colegio=colegio)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('nuevo_estado')
        feedback_utp = request.POST.get('feedback_utp', '').strip()

        if nuevo_estado in ['aprobada', 'observada', 'enviada', 'borrador']:
            plan.estado = nuevo_estado
            plan.feedback_utp = feedback_utp
            plan.revisado_por_utp = request.user
            plan.fecha_revision_utp = timezone.now()
            plan.save()
            messages.success(request, f"Planificación actualizada a estado: {plan.get_estado_display()}")

    return redirect('planificaciones_hub')


# ==============================================================================
# VISTAS: MÓDULO PIE (PROGRAMA DE INTEGRACIÓN ESCOLAR / NEE)
# ==============================================================================

@login_required
def pie_dashboard_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import FichaEstudiantePIE, Estudiante, SeccionCurso
    from django.db.models import Count, Q

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP', 'Convivencia'])
    )

    fichas_qs = FichaEstudiantePIE.objects.filter(colegio=colegio, activo=True).select_related(
        'estudiante', 'estudiante__seccion', 'profesional_a_cargo'
    )

    # Filtros
    tipo_nee = request.GET.get('tipo_nee')
    seccion_id = request.GET.get('seccion_id')
    search_q = request.GET.get('q', '').strip()

    if tipo_nee:
        fichas_qs = fichas_qs.filter(tipo_nee=tipo_nee)
    if seccion_id and seccion_id.isdigit():
        fichas_qs = fichas_qs.filter(estudiante__seccion_id=int(seccion_id))
    if search_q:
        fichas_qs = fichas_qs.filter(
            Q(estudiante__nombre_completo__icontains=search_q) |
            Q(estudiante__rut__icontains=search_q) |
            Q(diagnostico_personalizado__icontains=search_q)
        )

    total_pie = FichaEstudiantePIE.objects.filter(colegio=colegio, activo=True).count()
    total_neet = FichaEstudiantePIE.objects.filter(colegio=colegio, activo=True, tipo_nee='transitoria').count()
    total_neep = FichaEstudiantePIE.objects.filter(colegio=colegio, activo=True, tipo_nee='permanente').count()
    total_con_paci = FichaEstudiantePIE.objects.filter(colegio=colegio, activo=True, requiere_paci=True).count()

    # Estudiantes del colegio aún no inscritos en PIE
    alumnos_disponibles = Estudiante.objects.filter(
        colegio=colegio, activo=True
    ).exclude(ficha_pie__isnull=False).order_by('nombre_completo')

    secciones_colegio = SeccionCurso.objects.filter(curso__colegio=colegio, activo=True).select_related('curso').order_by('curso__orden', 'letra')
    especialistas = User.objects.filter(
        Q(solicitudes_acceso__colegio=colegio, solicitudes_acceso__estado='aprobada') |
        Q(asignaturas_dictadas__colegio=colegio)
    ).distinct().order_by('first_name', 'last_name')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'fichas_pie': fichas_qs,
        'total_pie': total_pie,
        'total_neet': total_neet,
        'total_neep': total_neep,
        'total_con_paci': total_con_paci,
        'alumnos_disponibles': alumnos_disponibles,
        'secciones_colegio': secciones_colegio,
        'especialistas': especialistas,
        'active_page': 'pie',
    }
    return render(request, 'colegios/pie_dashboard.html', context)


@login_required
def crear_ficha_pie_view(request):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    if request.method == 'POST':
        from .models import FichaEstudiantePIE, Estudiante
        import datetime

        estudiante_id = request.POST.get('estudiante_id')
        tipo_nee = request.POST.get('tipo_nee', 'transitoria')
        diagnostico = request.POST.get('diagnostico')
        diagnostico_personalizado = request.POST.get('diagnostico_personalizado', '').strip()
        profesional_id = request.POST.get('profesional_id')
        fecha_ingreso_str = request.POST.get('fecha_ingreso')
        fecha_revaluacion_str = request.POST.get('fecha_revaluacion')
        requiere_paci = request.POST.get('requiere_paci') == 'on'
        observaciones_ingreso = request.POST.get('observaciones_ingreso', '').strip()

        estudiante = get_object_or_404(Estudiante, id=estudiante_id, colegio=colegio)
        profesional = User.objects.filter(id=profesional_id).first() if profesional_id else None

        try:
            f_ing = datetime.datetime.strptime(fecha_ingreso_str, '%Y-%m-%d').date() if fecha_ingreso_str else timezone.now().date()
        except ValueError:
            f_ing = timezone.now().date()

        f_rev = None
        if fecha_revaluacion_str:
            try:
                f_rev = datetime.datetime.strptime(fecha_revaluacion_str, '%Y-%m-%d').date()
            except ValueError:
                f_rev = None

        ficha, created = FichaEstudiantePIE.objects.update_or_create(
            estudiante=estudiante,
            defaults={
                'colegio': colegio,
                'tipo_nee': tipo_nee,
                'diagnostico': diagnostico,
                'diagnostico_personalizado': diagnostico_personalizado,
                'profesional_a_cargo': profesional,
                'fecha_ingreso': f_ing,
                'fecha_revaluacion': f_rev,
                'requiere_paci': requiere_paci,
                'observaciones_ingreso': observaciones_ingreso,
                'activo': True
            }
        )
        messages.success(request, f"¡Ficha PIE creada con éxito para {estudiante.nombre_completo}!")
        return redirect('detalle_estudiante_pie', ficha_id=ficha.id)

    return redirect('pie_dashboard')


@login_required
def detalle_estudiante_pie_view(request, ficha_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import FichaEstudiantePIE, AtencionEspecialistaPIE, PlanAdecuacionCurricular, Asignatura

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP', 'Convivencia'])
    )

    ficha = get_object_or_404(FichaEstudiantePIE, id=ficha_id, colegio=colegio)
    atenciones = ficha.atenciones.all().select_related('especialista').order_by('-fecha', '-fecha_registro')
    pacis = ficha.pacis.all().select_related('asignatura', 'aprobado_por').order_by('-anio_lectivo')
    asignaturas = Asignatura.objects.filter(colegio=colegio, activo=True).order_by('nombre')
    especialistas = User.objects.filter(
        Q(solicitudes_acceso__colegio=colegio, solicitudes_acceso__estado='aprobada') |
        Q(asignaturas_dictadas__colegio=colegio)
    ).distinct().order_by('first_name', 'last_name')

    context = {
        'colegio': colegio,
        'miembro': miembro,
        'is_admin': is_admin,
        'ficha': ficha,
        'atenciones': atenciones,
        'pacis': pacis,
        'asignaturas': asignaturas,
        'especialistas': especialistas,
        'hoy': timezone.now(),
        'active_page': 'pie',
    }
    return render(request, 'colegios/detalle_estudiante_pie.html', context)


@login_required
def registrar_sesion_pie_view(request, ficha_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import FichaEstudiantePIE, AtencionEspecialistaPIE
    ficha = get_object_or_404(FichaEstudiantePIE, id=ficha_id, colegio=colegio)

    if request.method == 'POST':
        import datetime

        rol_especialista = request.POST.get('rol_especialista', 'educadora_diferencial')
        tipo_sesion = request.POST.get('tipo_sesion', 'aula_recursos')
        fecha_str = request.POST.get('fecha')
        objetivo_trabajado = request.POST.get('objetivo_trabajado', '').strip()
        resumen_intervencion = request.POST.get('resumen_intervencion', '').strip()
        acuerdos_pedagogicos = request.POST.get('acuerdos_pedagogicos', '').strip()

        try:
            f_sesion = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else timezone.now().date()
        except ValueError:
            f_sesion = timezone.now().date()

        AtencionEspecialistaPIE.objects.create(
            ficha_pie=ficha,
            especialista=request.user,
            rol_especialista=rol_especialista,
            fecha=f_sesion,
            tipo_sesion=tipo_sesion,
            objetivo_trabajado=objetivo_trabajado,
            resumen_intervencion=resumen_intervencion,
            acuerdos_pedagogicos=acuerdos_pedagogicos
        )
        messages.success(request, f"¡Atención de especialista registrada en la bitácora PIE de {ficha.estudiante.nombre_completo}!")

    return redirect('detalle_estudiante_pie', ficha_id=ficha.id)


@login_required
def guardar_paci_view(request, ficha_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import FichaEstudiantePIE, PlanAdecuacionCurricular, Asignatura
    ficha = get_object_or_404(FichaEstudiantePIE, id=ficha_id, colegio=colegio)

    if request.method == 'POST':
        from .models import PlanAdecuacionCurricular, Asignatura

        paci_id = request.POST.get('paci_id')
        asignatura_id = request.POST.get('asignatura_id')
        anio_lectivo = request.POST.get('anio_lectivo', 2026)
        tiempo_adicional = request.POST.get('tiempo_adicional') == 'on'
        adaptacion_materiales = request.POST.get('adaptacion_materiales', '').strip()
        espacio_evaluacion = request.POST.get('espacio_evaluacion', '').strip()
        graduacion_complejidad = request.POST.get('graduacion_complejidad', '').strip()
        priorizacion_objetivos = request.POST.get('priorizacion_objetivos', '').strip()
        estrategias_evaluacion = request.POST.get('estrategias_evaluacion', '').strip()
        observaciones = request.POST.get('observaciones', '').strip()

        asignatura = Asignatura.objects.filter(id=asignatura_id, colegio=colegio).first() if asignatura_id else None

        if paci_id and paci_id.isdigit():
            paci = get_object_or_404(PlanAdecuacionCurricular, id=int(paci_id), ficha_pie=ficha)
            paci.asignatura = asignatura
            paci.anio_lectivo = int(anio_lectivo)
            paci.tiempo_adicional = tiempo_adicional
            paci.adaptacion_materiales = adaptacion_materiales
            paci.espacio_evaluacion = espacio_evaluacion
            paci.graduacion_complejidad = graduacion_complejidad
            paci.priorizacion_objetivos = priorizacion_objetivos
            paci.estrategias_evaluacion = estrategias_evaluacion
            paci.observaciones = observaciones
            paci.save()
            messages.success(request, f"¡PACI actualizado exitosamente para {ficha.estudiante.nombre_completo}!")
        else:
            PlanAdecuacionCurricular.objects.create(
                ficha_pie=ficha,
                asignatura=asignatura,
                anio_lectivo=int(anio_lectivo),
                tiempo_adicional=tiempo_adicional,
                adaptacion_materiales=adaptacion_materiales,
                espacio_evaluacion=espacio_evaluacion,
                graduacion_complejidad=graduacion_complejidad,
                priorizacion_objetivos=priorizacion_objetivos,
                estrategias_evaluacion=estrategias_evaluacion,
                observaciones=observaciones
            )
            messages.success(request, f"¡PACI creado exitosamente para {ficha.estudiante.nombre_completo}!")

    return redirect('detalle_estudiante_pie', ficha_id=ficha.id)


@login_required
def aprobar_paci_view(request, paci_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from solicitudes.models import MiembroColegio
    from .models import PlanAdecuacionCurricular

    miembro = MiembroColegio.objects.filter(usuario=request.user, colegio=colegio, activo=True).first()
    is_admin = (
        request.user.colegios_administrados.filter(id=colegio.id).exists()
        or (miembro and miembro.rol and miembro.rol.nombre in ['Administrador', 'Director', 'UTP'])
    )

    if not is_admin:
        messages.error(request, "Solo el equipo de UTP o Dirección puede aprobar el PACI oficial.")
        return redirect('pie_dashboard')

    paci = get_object_or_404(PlanAdecuacionCurricular, id=paci_id, ficha_pie__colegio=colegio)
    paci.aprobado_utp = True
    paci.fecha_aprobacion = timezone.now().date()
    paci.aprobado_por = request.user
    paci.save()
    messages.success(request, f"¡PACI aprobado oficialmente por UTP para {paci.ficha_pie.estudiante.nombre_completo}!")

    return redirect('detalle_estudiante_pie', ficha_id=paci.ficha_pie.id)


@login_required
def paci_imprimible_view(request, ficha_id):
    colegio = obtener_colegio_usuario(request.user)
    if not colegio:
        return redirect('dashboard_usuario')

    from .models import FichaEstudiantePIE

    ficha = get_object_or_404(FichaEstudiantePIE, id=ficha_id, colegio=colegio)
    pacis = ficha.pacis.all().select_related('asignatura', 'aprobado_por').order_by('-anio_lectivo')
    atenciones_recientes = ficha.atenciones.all().select_related('especialista')[:10]

    context = {
        'colegio': colegio,
        'ficha': ficha,
        'estudiante': ficha.estudiante,
        'pacis': pacis,
        'atenciones_recientes': atenciones_recientes,
        'hoy': timezone.now(),
    }
    return render(request, 'colegios/paci_imprimible.html', context)




